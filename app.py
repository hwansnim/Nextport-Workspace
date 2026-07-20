"""
하루픽스 공동구매 컨텐츠 툴 - 셀러 아카이빙 서버
- Flask 웹 서버
- 셀러별 업데이트 트리거
- 진행 상황 표시
"""
from __future__ import annotations

import json
import logging
import os
import re
import secrets
import sys
import threading
import uuid
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from flask import Flask, jsonify, make_response, redirect, render_template, request, send_from_directory, session
from flask_cors import CORS

# Project paths
ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
LOGS_DIR = ROOT / "logs"
MODULES_DIR = ROOT / "modules"
SELLERS_FILE = DATA_DIR / "sellers.json"
CONFIG_FILE = ROOT / "config.json"
CONFIG_EXAMPLE = ROOT / "config.example.json"

# Ensure dirs
DATA_DIR.mkdir(exist_ok=True)
LOGS_DIR.mkdir(exist_ok=True)

# ─── 구글 드라이브 동기화 부트스트랩 ───────────────────────────
# Render 무료(휘발성 디스크) 대비: 시작 시 드라이브에서 data/ 받아오고,
# 변경분은 워처가 드라이브로 업로드. token.json(로컬) 또는 GOOGLE_TOKEN_JSON(클라우드) 필요.
if os.environ.get("DRIVE_SYNC", "1") != "0":
    try:
        from modules import drive_sync as _drive_sync
        if _drive_sync.enabled():
            _drive_sync.download_all(DATA_DIR)
            _drive_sync.start_watcher(DATA_DIR)
            logging.getLogger("app").info("[app] 구글 드라이브 동기화 ON")
    except Exception as _e:
        logging.getLogger("app").warning(f"[app] 드라이브 동기화 비활성 (무시): {_e}")

# Logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOGS_DIR / "server.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("app")

# Add modules to import path
sys.path.insert(0, str(MODULES_DIR))


def load_config() -> dict[str, Any]:
    cfg = {}
    if CONFIG_FILE.exists():
        try:
            cfg = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        except Exception:
            cfg = {}
    # 환경변수 우선 (Production 배포 시 비밀 박기)
    cfg.setdefault("gemini", {})
    env_key = os.getenv("GEMINI_API_KEY")
    if env_key:
        cfg["gemini"]["api_key"] = env_key
    # 환경 모드 (cloud / local)
    cfg["env_mode"] = os.getenv("ENV_MODE", "local")  # "cloud" or "local"
    return cfg


def load_sellers() -> list[dict[str, Any]]:
    if not SELLERS_FILE.exists():
        return []
    raw = json.loads(SELLERS_FILE.read_text(encoding="utf-8"))
    return raw.get("sellers", [])


def save_sellers(sellers: list[dict[str, Any]]) -> None:
    SELLERS_FILE.write_text(
        json.dumps({"sellers": sellers}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


# In-memory job state. Each job tracks one update run.
JOBS: dict[str, dict[str, Any]] = {}
JOB_LOCK = threading.Lock()


def make_job(seller_id: str) -> str:
    job_id = uuid.uuid4().hex[:12]
    with JOB_LOCK:
        JOBS[job_id] = {
            "id": job_id,
            "seller_id": seller_id,
            "status": "queued",  # queued | running | done | error
            "progress": 0,
            "total": 0,
            "message": "대기 중...",
            "started_at": datetime.now().isoformat(timespec="seconds"),
            "finished_at": None,
            "errors": [],
            "items_added": 0,
        }
    return job_id


def update_job(job_id: str, **kwargs) -> None:
    with JOB_LOCK:
        if job_id in JOBS:
            JOBS[job_id].update(kwargs)


def run_archive_job(job_id: str, seller: dict[str, Any]) -> None:
    """백그라운드 스레드에서 실제 아카이빙 작업 실행."""
    update_job(job_id, status="running", message=f"{seller['name']} 아카이빙 시작...")
    try:
        # 실제 스크래퍼는 modules/scraper.py 에서 import (다음 단계에 구현)
        try:
            from scraper import archive_seller  # type: ignore
        except ImportError:
            update_job(
                job_id,
                status="error",
                message="스크래퍼 모듈이 아직 준비되지 않았어요. (modules/scraper.py)",
                finished_at=datetime.now().isoformat(timespec="seconds"),
            )
            return

        cfg = load_config()
        result = archive_seller(
            seller=seller,
            config=cfg,
            on_progress=lambda **kw: update_job(job_id, **kw),
        )
        update_job(
            job_id,
            status="done",
            message=f"완료! {result.get('items_added', 0)}개 추가됨",
            items_added=result.get("items_added", 0),
            finished_at=datetime.now().isoformat(timespec="seconds"),
        )
    except Exception as e:  # noqa: BLE001
        log.exception("Archive job failed")
        update_job(
            job_id,
            status="error",
            message=f"에러: {e}",
            finished_at=datetime.now().isoformat(timespec="seconds"),
        )


# Flask app
app = Flask(__name__, template_folder=str(ROOT / "templates"), static_folder=str(ROOT / "static"))
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0  # 정적파일 장기 캐시 방지
CORS(app)

# ─────────────────────────────────────────────────────────────────
# 팀원 인증 / 로그인 게이트 / 실시간 접속 현황
# ─────────────────────────────────────────────────────────────────
from modules import team_auth  # noqa: E402

app.secret_key = team_auth.get_secret()
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=bool(os.environ.get("RENDER")),  # Render(HTTPS)에서만 Secure
    PERMANENT_SESSION_LIFETIME=timedelta(days=30),
)

# 로그인 없이도 접근 가능한 엔드포인트 (셀러 외부 뷰 + 인증 화면)
PUBLIC_ENDPOINTS = {
    # 셀러 / 외부 공개 라우트
    "seller_page", "api_seller_data", "api_archive_img",
    "api_campaigns_v2_campaign_share", "api_campaigns_v2_share_link",
    "seller_view", "api_sv_track", "api_sv_slot", "api_seller_track",
    "api_seller_tracking_get", "api_campaigns_v2_patch_slot",
    # 인증 화면 / API
    "static", "setup_page", "login_page", "login_page_token",
    "api_setup", "api_login",
}
# 관리자 전용 엔드포인트 (팀원 관리 + 접속 현황)
ADMIN_ENDPOINTS = {
    "api_team_list", "api_team_add", "api_team_update",
    "api_team_delete", "api_team_regen", "api_presence_list",
}


@app.before_request
def _auth_gate():
    ep = request.endpoint
    path = request.path or ""
    if ep is None:
        return  # 404 등은 그대로
    if ep in PUBLIC_ENDPOINTS or path.startswith("/static/") or path == "/favicon.ico":
        return
    # 관리자 계정이 아직 없으면 → 최초 설정 화면으로
    if not team_auth.has_admin():
        if path.startswith("/api/"):
            return jsonify({"error": "setup_required"}), 401
        return redirect("/setup")
    # 로그인 세션 확인
    member = team_auth.member_by_id(session.get("member_id"))
    if not member:
        if path.startswith("/api/"):
            return jsonify({"error": "auth_required"}), 401
        return redirect("/login")
    # 관리자 전용 보호
    if ep in ADMIN_ENDPOINTS and member.get("role") != "admin":
        return jsonify({"error": "forbidden"}), 403
    request.member = member  # 이후 핸들러에서 사용


def _current_member():
    return getattr(request, "member", None) or team_auth.member_by_id(session.get("member_id"))


def _base_url() -> str:
    return request.host_url.rstrip("/")


def _member_public(m: dict, with_password: bool = False) -> dict:
    out = {
        "id": m["id"], "name": m["name"], "role": m.get("role", "member"),
        "token": m["token"], "created_at": m.get("created_at"),
        "link": f"{_base_url()}/enter/{m['token']}",
    }
    if with_password:
        out["password"] = m.get("password", "")
    return out


# ── 인증 화면 ──
@app.route("/setup")
def setup_page():
    if team_auth.has_admin():
        return redirect("/login")
    return render_template("login.html", mode="setup", ver=_asset_ver())


@app.route("/login")
def login_page():
    if not team_auth.has_admin():
        return redirect("/setup")
    return render_template("login.html", mode="login", token="", member_name="", ver=_asset_ver())


@app.route("/enter/<token>")
def login_page_token(token):
    if not team_auth.has_admin():
        return redirect("/setup")
    m = team_auth.member_by_token(token)
    if not m:
        return render_template("login.html", mode="invalid", token="", member_name="", ver=_asset_ver())
    if session.get("member_id") == m["id"]:
        return redirect("/")
    return render_template("login.html", mode="login", token=token, member_name=m["name"], ver=_asset_ver())


# ── 인증 API ──
@app.route("/api/setup", methods=["POST"])
def api_setup():
    if team_auth.has_admin():
        return jsonify({"error": "이미 관리자가 설정되어 있습니다"}), 400
    p = request.get_json(force=True, silent=True) or {}
    name = (p.get("name") or "관리자").strip()
    pw = (p.get("password") or "").strip()
    if len(pw) < 4:
        return jsonify({"error": "비밀번호는 4자 이상이어야 합니다"}), 400
    m = team_auth.add_member(name, pw, role="admin")
    session.permanent = True
    session["member_id"] = m["id"]
    team_auth.touch_presence(m["id"])
    return jsonify({"ok": True, "link": f"{_base_url()}/enter/{m['token']}"})


@app.route("/api/login", methods=["POST"])
def api_login():
    p = request.get_json(force=True, silent=True) or {}
    m = team_auth.member_by_token(p.get("token") or "")
    if not m or m.get("password", "") != (p.get("password") or ""):
        return jsonify({"error": "링크 또는 비밀번호가 올바르지 않습니다"}), 401
    session.permanent = True
    session["member_id"] = m["id"]
    team_auth.touch_presence(m["id"])
    return jsonify({"ok": True, "role": m.get("role", "member")})


@app.route("/api/logout", methods=["POST"])
def api_logout():
    m = _current_member()
    if m:
        team_auth.drop_presence(m["id"])
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/me")
def api_me():
    m = _current_member()
    if not m:
        return jsonify({"error": "auth_required"}), 401
    return jsonify(_member_public(m, with_password=(m.get("role") == "admin")))


# ── 팀원 관리 (관리자 전용) ──
@app.route("/api/team")
def api_team_list():
    pres = team_auth.presence_status()
    members = []
    for m in team_auth.list_members():
        row = _member_public(m, with_password=True)
        st = pres.get(m["id"], {"online": False, "ago": None})
        row["online"] = st["online"]
        row["ago"] = st["ago"]
        row["is_me"] = (m["id"] == session.get("member_id"))
        members.append(row)
    return jsonify({"members": members})


@app.route("/api/team", methods=["POST"])
def api_team_add():
    p = request.get_json(force=True, silent=True) or {}
    name = (p.get("name") or "").strip()
    pw = (p.get("password") or "").strip()
    role = "admin" if p.get("role") == "admin" else "member"
    if not name:
        return jsonify({"error": "이름을 입력하세요"}), 400
    if len(pw) < 4:
        return jsonify({"error": "비밀번호는 4자 이상이어야 합니다"}), 400
    m = team_auth.add_member(name, pw, role=role)
    return jsonify({"ok": True, "member": _member_public(m, with_password=True)})


@app.route("/api/team/<mid>", methods=["PATCH"])
def api_team_update(mid):
    p = request.get_json(force=True, silent=True) or {}
    m, msg = team_auth.update_member(
        mid, name=p.get("name"), password=p.get("password"), role=p.get("role"))
    if not m:
        return jsonify({"error": msg}), 400
    return jsonify({"ok": True, "member": _member_public(m, with_password=True)})


@app.route("/api/team/<mid>/regen", methods=["POST"])
def api_team_regen(mid):
    m = team_auth.regen_token(mid)
    if not m:
        return jsonify({"error": "멤버를 찾을 수 없습니다"}), 404
    return jsonify({"ok": True, "member": _member_public(m, with_password=True)})


@app.route("/api/team/<mid>", methods=["DELETE"])
def api_team_delete(mid):
    if mid == session.get("member_id"):
        return jsonify({"error": "본인 계정은 삭제할 수 없습니다"}), 400
    ok, msg = team_auth.delete_member(mid)
    return (jsonify({"ok": True}) if ok else (jsonify({"error": msg}), 400))


# ── 실시간 접속 현황 ──
@app.route("/api/presence/ping", methods=["POST"])
def api_presence_ping():
    m = _current_member()
    if m:
        team_auth.touch_presence(m["id"])
    return jsonify({"ok": True})


@app.route("/api/presence/leave", methods=["POST"])
def api_presence_leave():
    """워크스페이스를 나갈 때(탭 닫기/이동) sendBeacon 으로 호출 → 즉시 오프라인."""
    m = _current_member()
    if m:
        team_auth.drop_presence(m["id"])
    return jsonify({"ok": True})


@app.route("/api/presence")
def api_presence_list():
    return jsonify({"presence": team_auth.presence_status()})


def _asset_ver() -> str:
    """정적파일 캐시 버스팅 — style.css/js 가 바뀌면 버전이 바뀌어 브라우저가 새로 받음."""
    try:
        static_dir = ROOT / "static"
        latest = max((f.stat().st_mtime for f in static_dir.glob("*.css")), default=0)
        if any(static_dir.glob("*.js")):
            latest = max(latest, *(f.stat().st_mtime for f in static_dir.glob("*.js")))
        return str(int(latest))
    except Exception:
        return "1"


@app.route("/")
def index():
    ver = _asset_ver()
    resp = make_response(render_template("index.html", ver=ver))
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return resp


@app.route("/api/sellers", methods=["GET"])
def api_sellers():
    sellers = load_sellers()
    return jsonify({"sellers": sellers})


@app.route("/api/sellers", methods=["POST"])
def api_sellers_add():
    payload = request.get_json(force=True)
    sellers = load_sellers()
    new_id = f"{len(sellers) + 1:03d}"
    new_seller = {
        "id": new_id,
        "name": payload.get("name", "").strip(),
        "instagram": payload.get("instagram", "").strip().lstrip("@"),
        "display_name": payload.get("display_name", payload.get("name", "")).strip(),
        "notes": payload.get("notes", "").strip(),
        "active": True,
    }
    if not new_seller["name"] or not new_seller["instagram"]:
        return jsonify({"error": "name 과 instagram 필수"}), 400
    sellers.append(new_seller)
    save_sellers(sellers)
    return jsonify({"seller": new_seller})


@app.route("/api/instagram/login", methods=["POST"])
def api_instagram_login():
    """IG 로그인 전용 작업. Playwright 창을 띄워 사용자가 로그인할 시간 줌."""
    job_id = uuid.uuid4().hex[:12]
    with JOB_LOCK:
        JOBS[job_id] = {
            "id": job_id,
            "seller_id": "_login",
            "status": "running",
            "progress": 0,
            "total": 10,
            "message": "로그인 헬퍼 시작...",
            "started_at": datetime.now().isoformat(timespec="seconds"),
            "finished_at": None,
            "errors": [],
            "items_added": 0,
        }

    def run():
        try:
            from scraper import login_helper  # type: ignore
            cfg = load_config()
            result = login_helper(cfg, on_progress=lambda **kw: update_job(job_id, **kw))
            update_job(
                job_id,
                status="done" if result.get("logged_in") else "error",
                finished_at=datetime.now().isoformat(timespec="seconds"),
            )
        except Exception as e:  # noqa: BLE001
            log.exception("Login helper failed")
            update_job(
                job_id, status="error", message=f"에러: {e}",
                finished_at=datetime.now().isoformat(timespec="seconds"),
            )

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/api/sellers/<seller_id>/update", methods=["POST"])
def api_seller_update(seller_id: str):
    sellers = load_sellers()
    seller = next((s for s in sellers if s["id"] == seller_id), None)
    if not seller:
        return jsonify({"error": "셀러 없음"}), 404
    job_id = make_job(seller_id)
    t = threading.Thread(target=run_archive_job, args=(job_id, seller), daemon=True)
    t.start()
    return jsonify({"job_id": job_id})


@app.route("/api/sellers/<seller_id>", methods=["PATCH"])
def api_seller_patch(seller_id: str):
    payload = request.get_json(force=True)
    sellers = load_sellers()
    seller = next((s for s in sellers if s["id"] == seller_id), None)
    if not seller:
        return jsonify({"error": "셀러 없음"}), 404
    for key in ("name", "instagram", "display_name", "notes", "active"):
        if key in payload:
            v = payload[key]
            if isinstance(v, str):
                v = v.strip().lstrip("@") if key == "instagram" else v.strip()
            seller[key] = v
    save_sellers(sellers)
    return jsonify({"seller": seller})


@app.route("/api/sellers/<seller_id>", methods=["DELETE"])
def api_seller_delete(seller_id: str):
    sellers = load_sellers()
    sellers = [s for s in sellers if s["id"] != seller_id]
    save_sellers(sellers)
    return jsonify({"ok": True})


@app.route("/api/sellers/<seller_id>/stats", methods=["GET"])
def api_seller_stats(seller_id: str):
    """셀러 manifest에서 통계 + 프로필 정보 반환."""
    sellers = load_sellers()
    seller = next((s for s in sellers if s["id"] == seller_id), None)
    if not seller:
        return jsonify({"error": "셀러 없음"}), 404
    seller_folder = f"{seller['id']}.{seller['name']}_@{seller['instagram']}"
    local_root = ROOT / "data" / "local_archive" / seller_folder
    manifest_path = local_root / "_manifest.json"
    if not manifest_path.exists():
        return jsonify({"profile": {}, "stats": {"total_items": 0}, "exists": False})
    try:
        data = json.loads(manifest_path.read_text(encoding="utf-8"))
        items = data.get("items", [])
        by_highlight = {}
        total_size = 0
        for it in items:
            label = it.get("highlight_label", "")
            if label:
                by_highlight[label] = by_highlight.get(label, 0) + 1
            total_size += it.get("size", 0) or 0
        return jsonify({
            "profile": data.get("profile", {}),
            "stats": {
                "total_items": len(items),
                "by_highlight": by_highlight,
                "total_size_bytes": total_size,
            },
            "exists": True,
            "local_path": str(local_root),
        })
    except Exception as e:  # noqa: BLE001
        return jsonify({"error": str(e)}), 500


@app.route("/api/sellers/<seller_id>/open", methods=["POST"])
def api_seller_open_folder(seller_id: str):
    """셀러 아카이브 폴더를 탐색기로 열기."""
    import subprocess

    sellers = load_sellers()
    seller = next((s for s in sellers if s["id"] == seller_id), None)
    if not seller:
        return jsonify({"error": "셀러 없음"}), 404
    seller_folder = f"{seller['id']}.{seller['name']}_@{seller['instagram']}"
    local_root = ROOT / "data" / "local_archive" / seller_folder
    if not local_root.exists():
        local_root.mkdir(parents=True, exist_ok=True)
    try:
        subprocess.Popen(["explorer.exe", str(local_root)])
        return jsonify({"ok": True, "path": str(local_root)})
    except Exception as e:  # noqa: BLE001
        return jsonify({"error": str(e)}), 500


@app.route("/api/jobs/<job_id>", methods=["GET"])
def api_job_status(job_id: str):
    with JOB_LOCK:
        job = JOBS.get(job_id)
    if not job:
        return jsonify({"error": "작업 없음"}), 404
    return jsonify(job)


@app.route("/api/config", methods=["GET"])
def api_config_status():
    cfg = load_config()
    return jsonify({
        "config_exists": CONFIG_FILE.exists(),
        "gemini_set": bool(cfg.get("gemini", {}).get("api_key", "").strip())
                      and not cfg.get("gemini", {}).get("api_key", "").startswith("여기에"),
        "drive_credentials_present": (ROOT / cfg.get("google", {}).get("credentials_file", "credentials.json")).exists(),
    })


@app.route("/health")
def health():
    return {"ok": True, "time": datetime.now().isoformat(timespec="seconds")}


# ─── BACKUP STATUS ────────────────────────────────────────
@app.route("/api/backup_status", methods=["GET"])
def api_backup_status():
    """현재 백업 폴더 위치 + 마지막 백업 시각 반환."""
    candidates = [
        Path("H:/내 드라이브/넥스트포트/공구/백업"),
        Path("G:/내 드라이브/넥스트포트/공구/백업"),
        Path.home() / "Desktop" / "공구" / "백업",
    ]
    base = None
    for c in candidates:
        if c.exists():
            base = c
            break
    if not base:
        return jsonify({
            "connected": False,
            "path": None,
            "last_backup": None,
            "message": "백업 폴더 없음",
        })

    latest_manifest = base / "최신" / "manifest.json"
    last_at = None
    file_count = 0
    if latest_manifest.exists():
        try:
            d = json.loads(latest_manifest.read_text(encoding="utf-8"))
            last_at = d.get("timestamp")
            file_count = d.get("ok", 0)
        except Exception:
            pass

    # 히스토리 개수
    history = base / "히스토리"
    history_count = 0
    if history.exists():
        history_count = sum(1 for _ in history.iterdir() if _.is_dir())

    is_gdrive = "내 드라이브" in str(base)
    return jsonify({
        "connected": True,
        "path": str(base),
        "is_gdrive": is_gdrive,
        "last_backup": last_at,
        "file_count": file_count,
        "history_count": history_count,
        "label": "Google Drive 동기화" if is_gdrive else "로컬 백업",
    })


@app.route("/api/backup_status/open", methods=["POST"])
def api_backup_open():
    """백업 폴더를 탐색기로 열기."""
    import subprocess
    candidates = [
        Path("H:/내 드라이브/넥스트포트/공구/백업"),
        Path("G:/내 드라이브/넥스트포트/공구/백업"),
        Path.home() / "Desktop" / "공구" / "백업",
    ]
    base = next((c for c in candidates if c.exists()), None)
    if not base:
        return jsonify({"error": "백업 폴더 없음"}), 404
    try:
        subprocess.Popen(["explorer.exe", str(base)])
        return jsonify({"ok": True, "path": str(base)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/backup/run", methods=["POST"])
def api_backup_run():
    """수동 백업 트리거."""
    import subprocess
    try:
        result = subprocess.run(
            [sys.executable, str(ROOT / "scripts" / "backup.py")],
            capture_output=True, text=True, timeout=60,
            encoding="utf-8", errors="replace",
        )
        return jsonify({
            "ok": result.returncode == 0,
            "stdout": result.stdout,
            "stderr": result.stderr,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── DASHBOARD (정산 + 매출 통합 뷰) ──────────────────────
@app.route("/api/dashboard", methods=["GET"])
def api_dashboard():
    """캠페인별 정산/매출 요약 + 자동 계산."""
    rows = []
    for c in load_campaigns():
        fi = c.get("financials") or {}
        st = c.get("settlement") or {}

        revenue = float(fi.get("revenue") or 0)
        seller_fee = float(fi.get("seller_fee") or 0)
        pg_fee = float(fi.get("pg_fee") or 0)
        event_cost = float(fi.get("event_cost") or 0)
        cost = float(fi.get("cost") or 0)
        shipping = float(fi.get("shipping") or 0)
        vat = float(fi.get("vat") or 0)
        total_cost = seller_fee + pg_fee + event_cost + cost + shipping + vat
        profit = revenue - total_cost
        rate = (profit / revenue * 100) if revenue > 0 else 0

        rows.append({
            "id": c["id"],
            "label": f"{c.get('seller_name','')} ({c.get('round',1)}차)",
            "seller_name": c.get("seller_name", ""),
            "seller_real_name": c.get("seller_real_name", ""),
            "owner": c.get("owner", ""),
            "brand": c.get("brand", ""),
            "product": c.get("product", ""),
            "round": c.get("round", 1),
            "live_start": c.get("live_start", ""),
            "live_end": c.get("live_end", ""),
            "open_kind": c.get("open_kind", ""),
            "status": c.get("status", ""),
            "settlement": st,
            "financials": fi,
            "calc": {
                "total_cost": total_cost,
                "contribution_profit": profit,
                "contribution_rate": round(rate, 2),
            },
        })
    # 시작일 순 정렬
    rows.sort(key=lambda r: r.get("live_start") or "9999-99-99")
    # 합계 (완료 건만)
    completed = [r for r in rows if r["status"] == "완료"]
    sum_rev = sum(float((r["financials"] or {}).get("revenue") or 0) for r in completed)
    sum_profit = sum(r["calc"]["contribution_profit"] for r in completed)
    return jsonify({
        "campaigns": rows,
        "totals": {
            "completed_count": len(completed),
            "total_revenue": sum_rev,
            "total_profit": sum_profit,
            "avg_rate": round(sum_profit / sum_rev * 100, 2) if sum_rev > 0 else 0,
        },
    })


# ─── HOLIDAYS (한국 공휴일) ────────────────────────────────
HOLIDAYS = {
    "2026-01-01": "신정",
    "2026-02-16": "설날",
    "2026-02-17": "설날",
    "2026-02-18": "설날",
    "2026-03-01": "삼일절",
    "2026-03-02": "삼일절 대체",
    "2026-05-05": "어린이날",
    "2026-05-24": "석가탄신일",
    "2026-05-25": "대체 휴일",
    "2026-06-03": "지방선거일",
    "2026-06-06": "현충일",
    "2026-08-15": "광복절",
    "2026-08-17": "광복절 대체",
    "2026-09-24": "추석",
    "2026-09-25": "추석",
    "2026-09-26": "추석",
    "2026-10-03": "개천절",
    "2026-10-05": "개천절 대체",
    "2026-10-09": "한글날",
    "2026-12-25": "성탄절",
}


@app.route("/api/holidays", methods=["GET"])
def api_holidays():
    return jsonify({"holidays": HOLIDAYS})


# ─── SELLER PUBLIC PAGE (모바일 친화 플레이북) ────────────
@app.route("/s/<token>")
def seller_page(token: str):
    """셀러용 공개 페이지. 토큰만 알면 접근."""
    return render_template("seller_page.html", token=token)


# ─── CLOUDFLARE TUNNEL (셀러 공개 모드) ────────────────────
TUNNEL_STATE = {
    "process": None,
    "url": None,
    "log": [],
    "started_at": None,
}
TUNNEL_LOCK = threading.Lock()
TUNNEL_EXE = ROOT / "cloudflared.exe"


def _tunnel_url_save(url: str | None):
    p = DATA_DIR / "tunnel.json"
    if url:
        p.write_text(json.dumps({"url": url, "saved_at": datetime.now().isoformat(timespec="seconds")},
                                ensure_ascii=False, indent=2), encoding="utf-8")
    elif p.exists():
        p.unlink()


def _tunnel_url_load() -> str | None:
    p = DATA_DIR / "tunnel.json"
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8")).get("url")
    except Exception:
        return None


def _start_tunnel():
    import re
    import subprocess

    with TUNNEL_LOCK:
        if TUNNEL_STATE["process"] and TUNNEL_STATE["process"].poll() is None:
            return  # 이미 실행 중
        if not TUNNEL_EXE.exists():
            TUNNEL_STATE["log"].append(f"[ERR] cloudflared.exe 없음: {TUNNEL_EXE}")
            return

        TUNNEL_STATE["log"] = ["[INFO] cloudflared 시작 중..."]
        TUNNEL_STATE["url"] = None
        TUNNEL_STATE["started_at"] = datetime.now().isoformat(timespec="seconds")

        proc = subprocess.Popen(
            [str(TUNNEL_EXE), "tunnel", "--url", "http://localhost:5000",
             "--no-autoupdate"],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            bufsize=1,
            text=True,
            encoding="utf-8",
            errors="replace",
            creationflags=0x08000000,  # CREATE_NO_WINDOW
        )
        TUNNEL_STATE["process"] = proc

    URL_RE = re.compile(r"(https?://[a-z0-9-]+\.trycloudflare\.com)")

    def read_output():
        for line in proc.stdout:
            line = line.rstrip()
            if line:
                TUNNEL_STATE["log"].append(line)
                if len(TUNNEL_STATE["log"]) > 200:
                    TUNNEL_STATE["log"] = TUNNEL_STATE["log"][-200:]
                m = URL_RE.search(line)
                if m and not TUNNEL_STATE["url"]:
                    TUNNEL_STATE["url"] = m.group(1)
                    _tunnel_url_save(TUNNEL_STATE["url"])
                    log.info(f"Cloudflare Tunnel: {TUNNEL_STATE['url']}")
        TUNNEL_STATE["log"].append("[INFO] cloudflared 종료됨")

    threading.Thread(target=read_output, daemon=True).start()


def _stop_tunnel():
    with TUNNEL_LOCK:
        p = TUNNEL_STATE.get("process")
        if p and p.poll() is None:
            try:
                p.terminate()
            except Exception:
                pass
        TUNNEL_STATE["process"] = None
        TUNNEL_STATE["url"] = None
        _tunnel_url_save(None)


@app.route("/api/tunnel/status", methods=["GET"])
def api_tunnel_status():
    p = TUNNEL_STATE.get("process")
    running = bool(p and p.poll() is None)
    url = TUNNEL_STATE.get("url") or _tunnel_url_load()
    return jsonify({
        "running": running,
        "url": url,
        "installed": TUNNEL_EXE.exists(),
        "started_at": TUNNEL_STATE.get("started_at"),
        "log_tail": TUNNEL_STATE["log"][-15:],
    })


@app.route("/api/tunnel/start", methods=["POST"])
def api_tunnel_start():
    if not TUNNEL_EXE.exists():
        return jsonify({"error": "cloudflared.exe 없음. 프로젝트 폴더에 다운로드 필요."}), 500
    _start_tunnel()
    # URL 생성까지 약간 대기
    import time
    for _ in range(20):
        if TUNNEL_STATE["url"]:
            break
        time.sleep(0.5)
    return jsonify({
        "running": True,
        "url": TUNNEL_STATE.get("url"),
        "started_at": TUNNEL_STATE.get("started_at"),
    })


@app.route("/api/tunnel/stop", methods=["POST"])
def api_tunnel_stop():
    _stop_tunnel()
    return jsonify({"running": False})


# ─── 셀러 아카이브 이미지 추천 ────────────────────────────
@app.route("/api/campaigns/<cid>/recommend_images", methods=["POST"])
def api_recommend_images(cid: str):
    """캠페인의 daily_schedule 각 STORY 슬롯에 어울리는 이미지를 셀러 아카이브에서 자동 추천.

    Gemini가 태깅한 story_slot_fit 점수 기반:
      slot 1 → STORY1_일상노출
      slot 2 → STORY2_고객반응
      slot 3 → STORY3_비포애프터
      slot 4 → STORY4_효능증명
      slot 5 → STORY5_공유어필
    """
    items = load_campaigns()
    c = next((x for x in items if x["id"] == cid), None)
    if not c:
        return jsonify({"error": "캠페인 없음"}), 404

    # 캠페인 셀러의 아카이브 manifest 찾기
    handle = c.get("seller_handle", "")
    seller = None
    for s in load_sellers():
        if s.get("instagram") == handle:
            seller = s
            break
    if not seller:
        return jsonify({"error": f"셀러 아카이브에 @{handle} 없음. 셀러 아카이브 탭에서 먼저 추가 + 업데이트 필요."}), 400

    seller_folder = f"{seller['id']}.{seller['name']}_@{seller['instagram']}"
    manifest_path = DATA_DIR / "local_archive" / seller_folder / "_manifest.json"
    if not manifest_path.exists():
        return jsonify({"error": f"@{handle} 의 manifest 없음. 셀러 아카이브 업데이트 먼저."}), 400

    try:
        mdata = json.loads(manifest_path.read_text(encoding="utf-8"))
        archive_items = mdata.get("items", [])
    except Exception as e:
        return jsonify({"error": f"manifest 읽기 실패: {e}"}), 500

    if not archive_items:
        return jsonify({"error": "아카이브 항목 없음"}), 400

    # 슬롯명 → 매니페스트의 story_slot_fit 키 매핑
    SLOT_KEY = {
        1: "STORY1_일상노출",
        2: "STORY2_고객반응",
        3: "STORY3_비포애프터",
        4: "STORY4_효능증명",
        5: "STORY5_공유어필",
    }

    # 슬롯별로 이미지 점수순 정렬 (이미 사용된 이미지는 제외)
    used = set()
    daily = c.get("daily_schedule", [])
    for d in daily:
        for s in (d.get("stories") or []):
            if s.get("image_url"):
                used.add(s["image_url"])

    matched_count = 0
    for d in daily:
        for s in d.get("stories") or []:
            if s.get("image_url"):
                continue  # 이미 있으면 건드리지 않음
            slot_no = s.get("slot", 1)
            key = SLOT_KEY.get(slot_no, "")
            if not key:
                continue
            # 후보 정렬 (점수 높은 순, 이미 사용된 거 제외)
            scored = []
            for it in archive_items:
                if not it.get("local_path"):
                    continue
                # 로컬 경로를 URL로 (정적 서빙 안 되어 있을 수도)
                local = it["local_path"]
                # 점수
                fit = (it.get("gemini") or {}).get("story_slot_fit") or {}
                score = fit.get(key, 0)
                if score <= 0:
                    continue
                # URL은 셀러 아카이브 폴더 경로
                if local in used:
                    continue
                scored.append((score, local))
            scored.sort(reverse=True)
            if scored:
                _score, path = scored[0]
                s["image_url"] = path  # 로컬 경로 (브라우저에서 직접 못 봐도 일단 채움)
                used.add(path)
                matched_count += 1

    c["updated_at"] = datetime.now().isoformat(timespec="seconds")
    save_campaigns(items)

    return jsonify({
        "ok": True,
        "matched": matched_count,
        "campaign": c,
        "message": f"{matched_count}개 슬롯에 이미지 자동 추천됨"
                   + ("" if matched_count else " — 아카이브 manifest에 story_slot_fit 태그 없음 (구버전). 셀러 아카이브 업데이트 시 자동 태깅됨."),
    })


# ─── GEMINI 자동 멘트 생성 ────────────────────────────────
@app.route("/api/campaigns/<cid>/generate_captions", methods=["POST"])
def api_generate_captions(cid: str):
    """한 일자의 STORY 슬롯 멘트 + 게시물 멘트를 Gemini로 자동 생성.

    컨텍스트:
      - 캠페인 (셀러 / 제품 / 브랜드 / 가이드라인 / 일정)
      - 제품 정보 (USP / 상세 / 가격 / 금지 멘트)
      - 셀러 톤 (셀러 아카이브에서 — 있으면)
      - 그 날의 stage (사전/라이브/사후) + day_label
    """
    items = load_campaigns()
    c = next((x for x in items if x["id"] == cid), None)
    if not c:
        return jsonify({"error": "캠페인 없음"}), 404

    payload = request.get_json(force=True, silent=True) or {}
    day_index = payload.get("day_index")
    day = payload.get("day") or {}
    if day_index is None:
        return jsonify({"error": "day_index 필수"}), 400

    # 제품 정보 가져오기
    product = None
    for p in load_products():
        if c.get("product") and (c["product"] in p.get("name", "") or p.get("name", "") in c.get("product", "")):
            product = p
            break

    # 셀러 아카이브 톤 정보 (있으면)
    seller_archive_info = ""
    handle = c.get("seller_handle")
    if handle:
        for s in load_sellers():
            if s.get("instagram") == handle:
                seller_folder = f"{s['id']}.{s['name']}_@{s['instagram']}"
                manifest_path = DATA_DIR / "local_archive" / seller_folder / "_manifest.json"
                if manifest_path.exists():
                    try:
                        mdata = json.loads(manifest_path.read_text(encoding="utf-8"))
                        prof = mdata.get("profile", {})
                        seller_archive_info = f"\n셀러 인스타 바이오: {prof.get('header_text','')[:300]}\n"
                    except Exception:
                        pass
                break

    # 가이드라인 (캠페인 notes + 제품 정보 합성)
    stories = day.get("stories") or []
    feed = day.get("feed_post") or {}

    prompt = f"""너는 인플루언서 공동구매 마켓의 인스타 스토리/게시물 멘트를 작성하는 카피라이터야.
대화 톤: 한국어 캐주얼 반말. 광고 느낌 절대 X. 친구한테 말하듯 자연스럽게. 이모지 적절히.

[캠페인 정보]
- 셀러: {c.get('seller_name','')} ({c.get('round',1)}차)
- 셀러 인스타 핸들: @{c.get('seller_handle','')}
- 브랜드: {c.get('brand','')}
- 제품: {c.get('product','')}
- 마켓 시작일: {c.get('live_start','')} / 종료일: {c.get('live_end','')}
- 단계: {day.get('kind','')} ({day.get('day_label','')})
- 오늘 날짜: {day.get('date','')}

[제품 상세]
{('USP: ' + product['usp']) if product and product.get('usp') else ''}
{('상세: ' + product['detail']) if product and product.get('detail') else ''}
{('가격/혜택: ' + product['price']) if product and product.get('price') else ''}

[금지 멘트]
{(product.get('avoid','') if product else '광고 느낌 멘트 X / 효능 단정 X')}
{seller_archive_info}
[캠페인 가이드라인]
{c.get('notes','')}

[작성할 슬롯 ({len(stories)}개 STORY + 1개 게시물)]
"""
    for s in stories:
        prompt += f"\nSLOT {s.get('slot', '?')}: {s.get('label', '')}\n  (기존 멘트: {s.get('caption', '')[:80]})\n"
    if feed:
        prompt += f"\nFEED: {feed.get('label', '게시물')}\n  (기존 멘트: {feed.get('caption', '')[:80]})\n"

    prompt += """

각 슬롯에 어울리는 멘트를 작성해. 출력은 다음 JSON 형식으로 (이게 끝):

{
  "stories": [
    {"slot": 1, "label": "[적절한 라벨]", "caption": "셀러가 인스타에 그대로 올릴 멘트. 줄바꿈 자유."},
    ...
  ],
  "feed": {"label": "[게시물 라벨]", "caption": "게시물 멘트"}
}

규칙:
- caption은 인스타 스토리에 그대로 올릴 수 있게. 한 문장씩 줄바꿈 가능.
- 단계별 톤:
  · 사전(pre): 일상 + 호기심 유발. 제품 직접 언급 X 또는 살짝.
  · 라이브(live): 마켓 알림 + 구매 유도. 명확하게.
  · 사후(post): 감사 + 구매자 케어 + 다음 N차 기대.
- 슬롯 라벨도 자연스럽게 (예: "[아침 한 포 루틴]", "[고객 반응]", "[3일째 후기]")
- 금지 멘트 무조건 회피.
- JSON 외 다른 텍스트 X.
"""

    try:
        import google.generativeai as genai
        cfg = load_config()
        api_key = cfg.get("gemini", {}).get("api_key", "")
        if not api_key:
            return jsonify({"error": "Gemini API key 없음"}), 500
        genai.configure(api_key=api_key)
        model_name = cfg.get("gemini", {}).get("caption_model") or "gemini-2.5-flash"
        model = genai.GenerativeModel(model_name, generation_config={
            "response_mime_type": "application/json",
            "max_output_tokens": 4096,
            "temperature": 0.85,
        })
        resp = model.generate_content(prompt)
        text = (resp.text or "").strip()
        # JSON 파싱
        if text.startswith("```"):
            text = text.strip("`").lstrip("json").strip()
        result = json.loads(text)
    except Exception as e:  # noqa: BLE001
        log.exception("caption gen failed")
        return jsonify({"error": f"생성 실패: {e}"}), 500

    # day 객체에 적용
    gen_stories = result.get("stories", []) or []
    for i, s in enumerate(day.get("stories") or []):
        if i < len(gen_stories):
            g = gen_stories[i]
            if g.get("label"): s["label"] = g["label"]
            if g.get("caption"): s["caption"] = g["caption"]

    gen_feed = result.get("feed") or {}
    if day.get("feed_post"):
        if gen_feed.get("label"): day["feed_post"]["label"] = gen_feed["label"]
        if gen_feed.get("caption"): day["feed_post"]["caption"] = gen_feed["caption"]

    # 캠페인 저장
    if day_index < len(c.get("daily_schedule", [])):
        c["daily_schedule"][day_index] = day
    else:
        c.setdefault("daily_schedule", []).append(day)
    c["updated_at"] = datetime.now().isoformat(timespec="seconds")
    save_campaigns(items)

    return jsonify({"ok": True, "day": day})


# ─── 카톡 대화 분석 ────────────────────────────────────────
@app.route("/api/campaigns/<cid>/kakao_log", methods=["POST"])
def api_kakao_log(cid: str):
    items = load_campaigns()
    c = next((x for x in items if x["id"] == cid), None)
    if not c:
        return jsonify({"error": "캠페인 없음"}), 404
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "txt 파일 필수"}), 400
    try:
        raw = f.read().decode("utf-8", errors="replace")
    except Exception as e:
        return jsonify({"error": f"파일 읽기 실패: {e}"}), 500
    if len(raw) > 40000:
        raw = raw[-40000:]

    prompt = f"""너는 비즈니스 카톡 대화 분석가야.
캠페인: {c.get('seller_name','')} {c.get('round',1)}차 / {c.get('brand','')} {c.get('product','')}
라이브: {c.get('live_start','')} ~ {c.get('live_end','')}

[카톡 원문]
{raw}

다음 JSON 출력:
{{
  "summary": "핵심 5~8문장 한국어 요약",
  "decisions": ["..."],
  "action_items": [{{"who":"...","what":"...","when":"..."}}],
  "stage_signals": ["진행 단계 신호"],
  "concerns": ["우려사항"]
}}
JSON만 출력."""

    try:
        import google.generativeai as genai
        cfg = load_config()
        key = cfg.get("gemini", {}).get("api_key", "")
        if not key:
            return jsonify({"error": "Gemini API key 없음"}), 500
        genai.configure(api_key=key)
        model = genai.GenerativeModel(
            cfg.get("gemini", {}).get("model") or "gemini-2.5-flash",
            generation_config={"response_mime_type": "application/json", "max_output_tokens": 4096},
        )
        resp = model.generate_content(prompt)
        text = (resp.text or "").strip()
        if text.startswith("```"):
            text = text.strip("`").lstrip("json").strip()
        result = json.loads(text)
    except Exception as e:
        log.exception("kakao analyze failed")
        return jsonify({"error": f"분석 실패: {e}"}), 500

    logs = c.get("kakao_logs") or []
    logs.append({
        "uploaded_at": datetime.now().isoformat(timespec="seconds"),
        "filename": f.filename,
        "summary": result.get("summary", ""),
        "decisions": result.get("decisions", []),
        "action_items": result.get("action_items", []),
        "stage_signals": result.get("stage_signals", []),
        "concerns": result.get("concerns", []),
    })
    c["kakao_logs"] = logs
    c["updated_at"] = datetime.now().isoformat(timespec="seconds")
    save_campaigns(items)
    return jsonify({"ok": True, "summary": result.get("summary", ""), "log": logs[-1]})


# ─── GOOGLE SHEETS IMPORT (셀러 스케줄 자동 가져오기) ────
@app.route("/api/campaigns/<cid>/import_sheet", methods=["POST"])
def api_import_sheet(cid: str):
    """캠페인의 시트 URL에서 daily_schedule 자동 import.

    시트는 '링크가 있는 모든 사람이 볼 수 있음' 권한 필요.
    인식하는 컬럼 (한글/영문):
      - 날짜 / date
      - D-라벨 / day_label / D
      - 종류 / kind / type
      - 제목 / title
      - 부제 / subtitle
      - 메모 / notes
    """
    import re
    import urllib.request
    import csv as csv_mod
    from io import StringIO

    items = load_campaigns()
    c = next((x for x in items if x["id"] == cid), None)
    if not c:
        return jsonify({"error": "캠페인 없음"}), 404

    payload = request.get_json(silent=True) or {}
    sheet_url = (payload.get("sheet_url") or c.get("sheet_url") or "").strip()
    if not sheet_url:
        return jsonify({"error": "캠페인에 시트 URL이 없음. 먼저 시트 URL 박으세요."}), 400

    # 시트 ID + GID 추출
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", sheet_url)
    if not m:
        return jsonify({"error": "올바른 Google Sheets URL이 아님"}), 400
    sheet_id = m.group(1)
    gid_m = re.search(r"[#&?]gid=(\d+)", sheet_url)
    gid = gid_m.group(1) if gid_m else "0"

    # CSV export URL
    csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"

    try:
        req = urllib.request.Request(csv_url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as r:
            raw = r.read().decode("utf-8")
    except Exception as e:
        return jsonify({
            "error": f"시트 가져오기 실패: {e}. 시트 공유 설정 확인 (링크 있는 사람 보기 가능).",
        }), 500

    # CSV 파싱
    reader = csv_mod.reader(StringIO(raw))
    rows = [row for row in reader]
    if len(rows) < 2:
        return jsonify({"error": "시트가 비어있음"}), 400

    # 시트 구조 자동 감지 — STORY 그리드형이 우선
    header_idx = None
    is_story_grid = False
    # 1단계: STORY 셀이 2개 이상인 행 (확실한 그리드형)
    for i, row in enumerate(rows[:20]):
        story_cells = sum(1 for c in row if "STORY" in (c or "").upper())
        if story_cells >= 2:
            header_idx = i
            is_story_grid = True
            break
    # 2단계: 단순 컬럼형 — 첫 셀이 정확히 "날짜"/"일자"/"date" 인 행
    if header_idx is None:
        for i, row in enumerate(rows[:15]):
            if not row:
                continue
            first = (row[0] or "").strip().lower()
            if first in ("날짜", "일자", "date"):
                header_idx = i
                break

    if header_idx is None:
        return jsonify({
            "error": "시트 헤더를 찾지 못함. STORY 1~5 같은 컬럼 또는 첫 컬럼에 '날짜' 헤더 필요.",
            "first_rows": [r[:5] for r in rows[:8]],
        }), 400

    # 캠페인 연도 추출 (날짜 파싱용)
    base_year = datetime.now().year
    if c.get("live_start"):
        try:
            base_year = int(c["live_start"][:4])
        except Exception:
            pass

    new_schedule = []

    if is_story_grid:
        # ── STORY 그리드형 (지나/양미라 시트 같은 양식) ──
        # header_idx: "날짜 / STORY 1 / STORY 2 / ... / 게시물 / 비고"
        headers = [(c or "").strip() for c in rows[header_idx]]
        # header_idx + 1 은 보통 슬롯별 가이드 (스킵)
        # header_idx + 2 부터 실제 날짜 행
        data_start = header_idx + 2

        # 한글 날짜 패턴: "4월 4일 (화)" / "4월 4일" / "4/4"
        import re as _re
        date_pat = _re.compile(r"(\d{1,2})\s*[월/]\s*(\d{1,2})")
        d_label_pat = _re.compile(r"(D[-+]?\d+|D[-]?Day|디데이)", _re.IGNORECASE)

        for row in rows[data_start:]:
            if not row or not any((c or "").strip() for c in row):
                continue
            date_cell = (row[0] if len(row) > 0 else "").strip()
            if not date_cell:
                continue

            # 날짜 추출
            m = date_pat.search(date_cell)
            if not m:
                continue
            month_n = int(m.group(1))
            day_n = int(m.group(2))
            date_str = f"{base_year}-{month_n:02d}-{day_n:02d}"

            # D-라벨 추출
            dl_m = d_label_pat.search(date_cell)
            day_label = dl_m.group(0) if dl_m else ""

            # STORY 1~5 + 게시물 + 비고 추출
            slot_parts = []
            for col_i in range(1, min(len(row), len(headers))):
                col_header = headers[col_i].strip()
                val = (row[col_i] or "").strip()
                if not val:
                    continue
                slot_parts.append(f"[{col_header}]\n{val}")

            # 카드 제목 = "이 날 할 일" 또는 첫 STORY 첫 줄
            first_line = ""
            if slot_parts:
                # 첫 STORY 내용 첫 줄 (라벨 줄 빼고)
                first_content = slot_parts[0].split("\n", 1)
                if len(first_content) > 1:
                    first_line = first_content[1].split("\n")[0]

            new_schedule.append({
                "date": date_str,
                "day_label": day_label,
                "kind": "content",
                "title": first_line[:60] if first_line else f"{month_n}월 {day_n}일 콘텐츠",
                "subtitle": f"STORY {len([s for s in slot_parts if s.startswith('[STORY')])}개 + 게시물 + 비고",
                "notes": "\n\n".join(slot_parts),
                "is_new": False,
            })
    else:
        # ── 단순 컬럼형 (기존 generic 파서) ──
        headers = [h.strip().lower() for h in rows[header_idx]]

        def find_col(*candidates):
            for cand in candidates:
                for i, h in enumerate(headers):
                    if cand in h:
                        return i
            return -1

        col_date = find_col("날짜", "일자", "date")
        col_label = find_col("d-라벨", "d-label", "라벨", "day_label", "차수일", "디데이", "d-")
        col_kind = find_col("종류", "구분", "kind", "type", "분류")
        col_title = find_col("제목", "타이틀", "title", "할일", "액션", "내용")
        col_sub = find_col("부제", "subtitle", "설명", "desc")
        col_notes = find_col("메모", "notes", "비고")

        kind_map = {
            "사전": "pre", "발송": "shipment", "콘텐츠": "content",
            "라이브": "live", "사후": "post", "기타": "other",
        }

        for row in rows[header_idx + 1:]:
            def cell(i):
                if i < 0 or i >= len(row):
                    return ""
                return (row[i] or "").strip()
            date_str = cell(col_date)
            if date_str:
                date_str = date_str.replace(".", "-").replace("/", "-").replace(" ", "")
                parts = date_str.split("-")
                if len(parts) == 3 and all(p.isdigit() for p in parts):
                    date_str = f"{parts[0]}-{int(parts[1]):02d}-{int(parts[2]):02d}"
            title = cell(col_title)
            if not date_str and not title:
                continue
            kind_raw = cell(col_kind).lower()
            new_schedule.append({
                "date": date_str,
                "day_label": cell(col_label),
                "kind": kind_map.get(kind_raw, "other"),
                "title": title,
                "subtitle": cell(col_sub),
                "notes": cell(col_notes),
                "is_new": False,
            })

    if not new_schedule:
        return jsonify({"error": "유효한 데이터 행이 없음", "headers_found": rows[header_idx] if header_idx is not None else []}), 400

    c["daily_schedule"] = new_schedule
    c["updated_at"] = datetime.now().isoformat(timespec="seconds")
    save_campaigns(items)

    return jsonify({
        "ok": True,
        "imported": len(new_schedule),
        "campaign": c,
        "headers_found": headers,
    })


@app.route("/api/seller/<token>", methods=["GET"])
def api_seller_data(token: str):
    """토큰 또는 slug로 캠페인 조회."""
    campaigns = load_campaigns()
    c = next((x for x in campaigns
              if x.get("seller_token") == token or x.get("seller_slug") == token), None)
    if not c:
        return jsonify({"error": "유효하지 않은 링크"}), 404

    # 브랜드 정보
    brand = None
    if c.get("brand_id"):
        brand = next((b for b in load_brands() if b["id"] == c["brand_id"]), None)
    # 제품 정보 (브랜드 매칭)
    product = None
    for p in load_products():
        if p.get("name") and (c.get("product") in p["name"] or p["name"] in c.get("product", "")):
            product = p
            break

    # D-N 계산
    from datetime import date
    today = datetime.now().date()
    d_label = "—"
    if c.get("live_start"):
        try:
            ls = datetime.strptime(c["live_start"], "%Y-%m-%d").date()
            le = datetime.strptime(c["live_end"], "%Y-%m-%d").date() if c.get("live_end") else None
            if today < ls:
                days = (ls - today).days
                d_label = f"D-{days}"
            elif le and today > le:
                d_label = "종료"
            else:
                d_label = "라이브 중 🔴"
        except ValueError:
            pass

    # 같은 셀러의 다른 차수들 (N차 스위치용) — slug 우선, fallback token
    same_seller = [
        {"id": x["id"], "round": x.get("round", 1),
         "token": x.get("seller_slug") or x.get("seller_token", ""),
         "status": x.get("status", ""), "live_start": x.get("live_start", "")}
        for x in campaigns
        if x.get("seller_name") == c.get("seller_name") and x.get("brand_id") == c.get("brand_id")
    ]
    same_seller.sort(key=lambda x: x.get("round") or 0)

    # 이 캠페인 관련 이벤트 (캘린더 직접 추가 + 캠페인 파생)
    related_events = []
    for ev in load_events():
        if ev.get("ref_id") == c["id"] or ev.get("ref_kind") == "":
            # 무관한 이벤트는 제외 (ref_id가 이 캠페인 또는 아예 비어있고 날짜가 라이브 기간 근처)
            if ev.get("ref_id") == c["id"]:
                related_events.append(ev)
    # 캠페인 자체에서 파생되는 이벤트
    derived = _derive_campaign_events(c)
    # 발송일/라이브 시작/종료가 의미있음 — 합치기
    for ev in derived:
        related_events.append(ev)
    related_events.sort(key=lambda x: x.get("date", ""))

    # 매출만 노출 — 🔒 원가/공헌이익/수수료는 셀러용 응답에 절대 포함하지 않음.
    fi = c.get("financials") or {}
    revenue = float(fi.get("revenue") or 0)

    return jsonify({
        "campaign": {
            "id": c["id"],
            "seller_name": c.get("seller_name", ""),
            "seller_real_name": c.get("seller_real_name", ""),
            "round": c.get("round", 1),
            "brand": c.get("brand", ""),
            "product": c.get("product", ""),
            "live_start": c.get("live_start", ""),
            "live_end": c.get("live_end", ""),
            "open_kind": c.get("open_kind", ""),
            "stage": c.get("stage", ""),
            "stage_label": STAGE_LABEL.get(c.get("stage", ""), c.get("stage", "")),
            "status": c.get("status", ""),
            "notes": c.get("notes", ""),
            "sheet_url": c.get("sheet_url", ""),
            "daily_schedule": c.get("daily_schedule") or [],
            "faq": c.get("faq") or [],
            "d_label": d_label,
            "shipment": c.get("shipment") or {"qty": 0, "date": ""},
            "settlement": c.get("settlement") or {},
        },
        "brand": brand,
        "product": product,
        "other_rounds": same_seller,
        "events": related_events,
        "financials": {
            "revenue": revenue,
            "has_data": revenue > 0,
        },
    })


# ─── PRODUCTS ─────────────────────────────────────────────
PRODUCTS_FILE = DATA_DIR / "products.json"


def load_products() -> list[dict[str, Any]]:
    if not PRODUCTS_FILE.exists():
        return []
    raw = json.loads(PRODUCTS_FILE.read_text(encoding="utf-8"))
    return raw.get("products", [])


def save_products(products: list[dict[str, Any]]) -> None:
    PRODUCTS_FILE.write_text(
        json.dumps({"products": products}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


@app.route("/api/products", methods=["GET"])
def api_products_list():
    return jsonify({"products": load_products()})


@app.route("/api/products", methods=["POST"])
def api_products_add():
    payload = request.get_json(force=True)
    products = load_products()
    new_id = f"p{len(products) + 1:03d}"
    new = {
        "id": new_id,
        "name": payload.get("name", "").strip(),
        "usp": payload.get("usp", "").strip(),
        "detail": payload.get("detail", "").strip(),
        "price": payload.get("price", "").strip(),
        "avoid": payload.get("avoid", "").strip(),
        "mall_url": (payload.get("mall_url") or "").strip(),
        "tiers": payload.get("tiers") if isinstance(payload.get("tiers"), list) else [],
    }
    if not new["name"]:
        return jsonify({"error": "name 필수"}), 400
    products.append(new)
    save_products(products)
    return jsonify({"product": new})


@app.route("/api/products/<pid>", methods=["PATCH"])
def api_products_patch(pid: str):
    payload = request.get_json(force=True)
    products = load_products()
    p = next((x for x in products if x["id"] == pid), None)
    if not p:
        return jsonify({"error": "not found"}), 404
    for k in ("name", "usp", "detail", "price", "avoid", "mall_url"):
        if k in payload:
            p[k] = (payload[k] or "").strip()
    if "tiers" in payload and isinstance(payload["tiers"], list):
        p["tiers"] = payload["tiers"]
    save_products(products)
    return jsonify({"product": p})


@app.route("/api/products/<pid>", methods=["DELETE"])
def api_products_delete(pid: str):
    products = [x for x in load_products() if x["id"] != pid]
    save_products(products)
    return jsonify({"ok": True})


# ─── 파일 업로드 (릴스 영상 등) → 드라이브 uploads 폴더 + 프록시 스트림 ───
@app.route("/api/upload", methods=["POST"])
def api_upload():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "file 없음"}), 400
    data = f.read()
    if len(data) > 200 * 1024 * 1024:
        return jsonify({"error": "200MB 초과"}), 413
    try:
        from modules import drive_sync
        r = drive_sync.upload_blob(f.filename or "file", data, f.mimetype)
        if not r:
            return jsonify({"error": "드라이브 미연결 (token 없음)"}), 503
        return jsonify({"ok": True, "file_id": r["id"], "name": r["name"], "url": f"/api/file/{r['id']}"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/backup/status", methods=["GET"])
def api_drive_backup_status():
    """백업(드라이브 동기화) 상태 — 마지막 저장 시각 등."""
    try:
        from modules import drive_sync
        return jsonify(drive_sync.status())
    except Exception as e:
        return jsonify({"enabled": False, "error": str(e)})


@app.route("/api/backup/now", methods=["POST"])
def api_drive_backup_now():
    """지금 즉시 모든 데이터를 드라이브로 백업."""
    try:
        from modules import drive_sync
        if not drive_sync.enabled():
            return jsonify({"ok": False, "error": "드라이브 미연결 (token 없음)"}), 503
        r = drive_sync.force_sync_all(DATA_DIR)
        return jsonify({"ok": True, **r})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/file/<file_id>", methods=["GET"])
def api_file(file_id):
    try:
        from modules import drive_sync
        got = drive_sync.get_file_bytes(file_id)
        if not got:
            return jsonify({"error": "드라이브 미연결"}), 503
        data, mime, _name = got
        resp = make_response(data)
        resp.headers["Content-Type"] = mime
        resp.headers["Cache-Control"] = "public, max-age=86400"
        return resp
    except Exception as e:
        return jsonify({"error": str(e)}), 404


# ─── SCHEDULE GENERATOR ───────────────────────────────────
@app.route("/api/schedule/generate", methods=["POST"])
def api_schedule_generate():
    payload = request.get_json(force=True)
    job_id = uuid.uuid4().hex[:12]
    with JOB_LOCK:
        JOBS[job_id] = {
            "id": job_id,
            "kind": "schedule",
            "status": "running",
            "progress": 0, "total": 10,
            "message": "스케줄 생성 시작…",
            "started_at": datetime.now().isoformat(timespec="seconds"),
            "finished_at": None,
            "errors": [],
            "result_url": None,
        }

    def run():
        try:
            from schedule_gen import generate_schedule
            cfg = load_config()
            result = generate_schedule(payload, cfg, on_progress=lambda **kw: update_job(job_id, **kw))
            if result.get("error"):
                update_job(job_id, status="error", message=result["error"], finished_at=datetime.now().isoformat(timespec="seconds"))
                return
            update_job(
                job_id,
                status="done",
                message=f"완료! {result.get('rows_count', 0)}일 × {result.get('slots_per_row', 0)}슬롯",
                result_url=result.get("result_url"),
                finished_at=datetime.now().isoformat(timespec="seconds"),
            )
        except Exception as e:  # noqa: BLE001
            log.exception("Schedule gen failed")
            update_job(job_id, status="error", message=f"에러: {e}", finished_at=datetime.now().isoformat(timespec="seconds"))

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/generated/<path:filename>")
def serve_generated(filename: str):
    return send_from_directory(DATA_DIR / "generated", filename)


# ─── CAMPAIGNS (셀러 캠페인) ──────────────────────────────
CAMPAIGNS_FILE = DATA_DIR / "campaigns.json"
MEETINGS_FILE = DATA_DIR / "meetings.json"
EVENTS_FILE = DATA_DIR / "events.json"
BRANDS_FILE = DATA_DIR / "brands.json"
MEETINGS_DIR = DATA_DIR / "meetings"
MEETINGS_DIR.mkdir(exist_ok=True)

# Google Drive 동기화 폴더 (자동 이관 위치)
DRIVE_MEETING_DIRS = [
    Path("H:/내 드라이브/넥스트포트/공구/녹취"),
    Path("G:/내 드라이브/넥스트포트/공구/녹취"),
]


def _find_drive_meeting_dir() -> Path | None:
    """녹취 자동 이관 가능한 Drive 폴더 찾기."""
    for d in DRIVE_MEETING_DIRS:
        if d.exists() or d.parent.exists():
            d.mkdir(parents=True, exist_ok=True)
            return d
    return None


def _safe_filename(s: str, fallback: str = "untitled") -> str:
    """파일명에 못 쓰는 글자 제거 + 공백 → _."""
    import re
    s = (s or "").strip()
    if not s:
        return fallback
    # Windows 금지 문자 + 제어 문자 제거
    s = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "", s)
    s = re.sub(r"\s+", "_", s)
    # 양옆 ., _ 제거
    s = s.strip("._")
    return s[:60] or fallback


STAGE_ORDER = [
    "contact",           # 컨택
    "confirmed",         # 셀러 컨펌
    "shipped",           # 제품 발송
    "received",          # 수령 확인
    "sheet_drafted",     # 시트 작성
    "sheet_confirmed",   # 시트 컨펌
    "live",              # 라이브 중
    "complete",          # 종료/정산
]
STAGE_LABEL = {
    "contact": "컨택",
    "confirmed": "셀러 컨펌",
    "shipped": "제품 발송",
    "received": "수령 확인",
    "sheet_drafted": "시트 작성",
    "sheet_confirmed": "시트 컨펌",
    "live": "라이브",
    "complete": "완료",
}


def _read_json(path: Path, key: str) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    raw = json.loads(path.read_text(encoding="utf-8"))
    return raw.get(key, [])


def _write_json(path: Path, key: str, items: list[dict[str, Any]]) -> None:
    path.write_text(json.dumps({key: items}, ensure_ascii=False, indent=2), encoding="utf-8")


def _next_id(items: list[dict[str, Any]], prefix: str) -> str:
    nums = []
    for it in items:
        s = str(it.get("id", ""))
        if s.startswith(prefix):
            try:
                nums.append(int(s[len(prefix):]))
            except ValueError:
                pass
    n = (max(nums) if nums else 0) + 1
    return f"{prefix}{n:03d}"


def load_campaigns() -> list[dict[str, Any]]:
    return _read_json(CAMPAIGNS_FILE, "campaigns")


def save_campaigns(items: list[dict[str, Any]]) -> None:
    _write_json(CAMPAIGNS_FILE, "campaigns", items)


def load_meetings() -> list[dict[str, Any]]:
    return _read_json(MEETINGS_FILE, "meetings")


def save_meetings(items: list[dict[str, Any]]) -> None:
    _write_json(MEETINGS_FILE, "meetings", items)


def load_events() -> list[dict[str, Any]]:
    return _read_json(EVENTS_FILE, "events")


def save_events(items: list[dict[str, Any]]) -> None:
    _write_json(EVENTS_FILE, "events", items)


CAMPAIGN_FIELDS = (
    "seller_name", "seller_handle", "seller_real_name", "owner",
    "brand", "brand_id", "product", "round",
    "contact_type", "stage", "live_start", "live_end", "open_kind",
    "sheet_url", "status", "notes", "seller_slug",
    "daily_schedule", "faq",
    "reels", "banner", "plan", "kakao_logs",
)


SELLER_SLUG_MAP = {
    "지나": "jina", "양미라": "yangmira", "한연아": "hanyeona",
    "윰니": "yumni", "김희연": "kimheeyeon", "오늘희": "onulhui", "느루": "neuru",
    "야곰": "yagom", "박예은": "parkyeeun", "이지비메": "easydiet",
}


def _make_slug(seller_name: str, seller_handle: str, round_no: int, fallback: str) -> str:
    """캠페인 slug 자동 생성."""
    import re
    base = SELLER_SLUG_MAP.get((seller_name or "").strip())
    if not base and seller_handle:
        base = re.sub(r"[^a-zA-Z0-9_]", "", seller_handle).lower()
    if not base:
        base = fallback
    return f"{base}-{round_no}"


def load_brands() -> list[dict[str, Any]]:
    return _read_json(BRANDS_FILE, "brands")


def save_brands(items: list[dict[str, Any]]) -> None:
    _write_json(BRANDS_FILE, "brands", items)


def _brand_id_from_name(brand_name: str) -> str:
    """기존 brand 텍스트 → brand_id 추정 (마이그레이션용)."""
    if not brand_name:
        return ""
    for b in load_brands():
        if b["name"] == brand_name or b["id"] == brand_name:
            return b["id"]
    return ""


@app.route("/api/brands", methods=["GET"])
def api_brands_list():
    items = load_brands()
    # 캠페인 개수 카운트
    camps = load_campaigns()
    for b in items:
        bid = b["id"]
        b["campaign_count"] = sum(
            1 for c in camps if c.get("brand_id") == bid or _brand_id_from_name(c.get("brand", "")) == bid
        )
    return jsonify({"brands": items})


@app.route("/api/brands", methods=["POST"])
def api_brands_add():
    payload = request.get_json(force=True)
    items = load_brands()
    import re
    name = (payload.get("name") or "").strip()
    if not name:
        return jsonify({"error": "name 필수"}), 400
    # id 자동 생성 — 영문/숫자만 + 짧게
    bid = payload.get("id") or re.sub(r"[^a-zA-Z0-9]", "", name.lower()) or f"b{len(items)+1}"
    if any(b["id"] == bid for b in items):
        bid = f"{bid}_{len(items)+1}"
    new = {
        "id": bid,
        "name": name,
        "category": (payload.get("category") or "").strip(),
        "color": (payload.get("color") or "#888").strip(),
        "emoji": (payload.get("emoji") or "🏷️").strip(),
        "active": payload.get("active", True),
        "notes": (payload.get("notes") or "").strip(),
    }
    items.append(new)
    save_brands(items)
    return jsonify({"brand": new})


@app.route("/api/brands/<bid>", methods=["PATCH"])
def api_brands_patch(bid: str):
    payload = request.get_json(force=True)
    items = load_brands()
    b = next((x for x in items if x["id"] == bid), None)
    if not b:
        return jsonify({"error": "not found"}), 404
    for k in ("name", "category", "color", "emoji", "active", "notes"):
        if k in payload:
            v = payload[k]
            if isinstance(v, str):
                v = v.strip()
            b[k] = v
    save_brands(items)
    return jsonify({"brand": b})


@app.route("/api/brands/<bid>", methods=["DELETE"])
def api_brands_delete(bid: str):
    items = [b for b in load_brands() if b["id"] != bid]
    save_brands(items)
    return jsonify({"ok": True})


@app.route("/api/campaigns", methods=["GET"])
def api_campaigns_list():
    return jsonify({"campaigns": load_campaigns()})


@app.route("/api/campaigns", methods=["POST"])
def api_campaigns_add():
    payload = request.get_json(force=True)
    items = load_campaigns()
    now = datetime.now().isoformat(timespec="seconds")
    brand_str = (payload.get("brand") or "").strip()
    brand_id = (payload.get("brand_id") or "").strip() or _brand_id_from_name(brand_str)
    # brand_id 있으면 brand 이름은 마스터에서 가져옴
    if brand_id:
        b = next((x for x in load_brands() if x["id"] == brand_id), None)
        if b:
            brand_str = b["name"]
    new = {
        "id": _next_id(items, "c"),
        "seller_name": (payload.get("seller_name") or "").strip(),
        "seller_handle": (payload.get("seller_handle") or "").strip().lstrip("@"),
        "brand": brand_str,
        "brand_id": brand_id,
        "product": (payload.get("product") or "").strip(),
        "round": int(payload.get("round") or 1),
        "contact_type": payload.get("contact_type") or "direct",
        "stage": payload.get("stage") or "contact",
        "live_start": payload.get("live_start") or "",
        "live_end": payload.get("live_end") or "",
        "shipment": payload.get("shipment") or {"qty": 0, "date": ""},
        "sheet_url": (payload.get("sheet_url") or "").strip(),
        "status": payload.get("status") or "예정",
        "notes": (payload.get("notes") or "").strip(),
        "created_at": now,
        "updated_at": now,
    }
    if not new["seller_name"]:
        return jsonify({"error": "seller_name 필수"}), 400
    # seller_token + slug 자동 생성
    import secrets
    new["seller_token"] = secrets.token_urlsafe(8).replace("-", "a").replace("_", "b")[:10]
    explicit_slug = (payload.get("seller_slug") or "").strip()
    base_slug = explicit_slug or _make_slug(new["seller_name"], new["seller_handle"], new["round"], new["id"])
    # 중복 회피
    existing = {c.get("seller_slug") for c in items}
    slug = base_slug
    i = 2
    while slug in existing:
        slug = f"{base_slug}-{i}"
        i += 1
    new["seller_slug"] = slug
    items.append(new)
    save_campaigns(items)
    return jsonify({"campaign": new})


@app.route("/api/campaigns/<cid>", methods=["PATCH"])
def api_campaigns_patch(cid: str):
    payload = request.get_json(force=True)
    items = load_campaigns()
    c = next((x for x in items if x["id"] == cid), None)
    if not c:
        return jsonify({"error": "not found"}), 404
    for k in CAMPAIGN_FIELDS:
        if k in payload:
            v = payload[k]
            if k == "round":
                v = int(v or 1)
            elif isinstance(v, str):
                v = v.strip()
                if k == "seller_handle":
                    v = v.lstrip("@")
            c[k] = v
    if "shipment" in payload:
        sh = payload["shipment"] or {}
        c["shipment"] = {
            "qty": int(sh.get("qty") or 0),
            "date": (sh.get("date") or "").strip(),
        }
    if "daily_schedule" in payload:
        c["daily_schedule"] = payload["daily_schedule"] or []
    if "faq" in payload:
        c["faq"] = payload["faq"] or []
    for asset_field in ("reels", "banner", "plan"):
        if asset_field in payload:
            v = payload[asset_field] or {}
            c[asset_field] = {
                "stage": (v.get("stage") or "").strip() if isinstance(v.get("stage"), str) else str(v.get("stage") or ""),
                "notes": (v.get("notes") or "").strip() if isinstance(v.get("notes"), str) else str(v.get("notes") or ""),
            }
    if "settlement" in payload:
        st = payload["settlement"] or {}
        c["settlement"] = {
            "rs_percent": float(st.get("rs_percent") or 0) if st.get("rs_percent") not in (None, "") else None,
            "type": (st.get("type") or "").strip(),
            "pg_logistics": (st.get("pg_logistics") or "").strip(),
            "completed_date": (st.get("completed_date") or "").strip(),
            "base_date": (st.get("base_date") or "").strip(),
        }
    if "financials" in payload:
        fi = payload["financials"] or {}
        # 숫자 필드만 받아서 깔끔히 저장
        cleaned = {}
        for k in ("revenue", "seller_fee", "pg_fee", "event_cost", "cost", "shipping", "vat"):
            v = fi.get(k)
            if v in (None, ""):
                continue
            try:
                cleaned[k] = float(v)
            except (TypeError, ValueError):
                pass
        c["financials"] = cleaned
    c["updated_at"] = datetime.now().isoformat(timespec="seconds")
    save_campaigns(items)
    return jsonify({"campaign": c})


@app.route("/api/campaigns/<cid>", methods=["DELETE"])
def api_campaigns_delete(cid: str):
    items = [c for c in load_campaigns() if c["id"] != cid]
    save_campaigns(items)
    return jsonify({"ok": True})


# ─── EVENTS (캘린더 직접 추가 이벤트) ─────────────────────
EVENT_FIELDS = ("date", "time", "kind", "title", "ref_id", "ref_kind", "color", "notes")
EVENT_KINDS = ("meeting", "shipment", "live_start", "live_end", "deadline", "other")


@app.route("/api/events", methods=["GET"])
def api_events_list():
    return jsonify({"events": load_events()})


@app.route("/api/events", methods=["POST"])
def api_events_add():
    payload = request.get_json(force=True)
    items = load_events()
    new = {
        "id": _next_id(items, "e"),
        "date": (payload.get("date") or "").strip(),
        "time": (payload.get("time") or "").strip(),
        "kind": payload.get("kind") or "other",
        "title": (payload.get("title") or "").strip(),
        "ref_id": payload.get("ref_id") or "",
        "ref_kind": payload.get("ref_kind") or "",
        "color": payload.get("color") or "",
        "notes": (payload.get("notes") or "").strip(),
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    if not new["date"] or not new["title"]:
        return jsonify({"error": "date 와 title 필수"}), 400
    items.append(new)
    save_events(items)
    return jsonify({"event": new})


@app.route("/api/events/<eid>", methods=["PATCH"])
def api_events_patch(eid: str):
    payload = request.get_json(force=True)
    items = load_events()
    e = next((x for x in items if x["id"] == eid), None)
    if not e:
        return jsonify({"error": "not found"}), 404
    for k in EVENT_FIELDS:
        if k in payload:
            v = payload[k]
            if isinstance(v, str):
                v = v.strip()
            e[k] = v
    save_events(items)
    return jsonify({"event": e})


@app.route("/api/events/<eid>", methods=["DELETE"])
def api_events_delete(eid: str):
    items = [x for x in load_events() if x["id"] != eid]
    save_events(items)
    return jsonify({"ok": True})


# ─── CALENDAR (캠페인 + 미팅 + events 합쳐서 반환) ────────
def _derive_campaign_events(campaign: dict[str, Any]) -> list[dict[str, Any]]:
    out = []
    seller = campaign.get("seller_name", "")
    rd = campaign.get("round", 1)
    cid = campaign.get("id", "")
    # 라이브 시작
    if campaign.get("live_start"):
        out.append({
            "id": f"{cid}.live_start",
            "date": campaign["live_start"],
            "time": "",
            "kind": "live_start",
            "title": f"🔴 {seller} {rd}차 라이브 시작",
            "ref_id": cid, "ref_kind": "campaign",
            "auto": True,
        })
    # 라이브 종료
    if campaign.get("live_end"):
        out.append({
            "id": f"{cid}.live_end",
            "date": campaign["live_end"],
            "time": "",
            "kind": "live_end",
            "title": f"⚪ {seller} {rd}차 라이브 종료",
            "ref_id": cid, "ref_kind": "campaign",
            "auto": True,
        })
    # 발송
    sh = campaign.get("shipment") or {}
    if sh.get("date"):
        qty = sh.get("qty", 0)
        out.append({
            "id": f"{cid}.shipment",
            "date": sh["date"],
            "time": "",
            "kind": "shipment",
            "title": f"🎁 {seller} 발송{' (' + str(qty) + '개)' if qty else ''}",
            "ref_id": cid, "ref_kind": "campaign",
            "auto": True,
        })
    return out


def _derive_meeting_events(meeting: dict[str, Any]) -> list[dict[str, Any]]:
    if not meeting.get("date"):
        return []
    return [{
        "id": f"{meeting['id']}.meeting",
        "date": meeting["date"],
        "time": meeting.get("time", ""),
        "kind": "meeting",
        "title": f"🎤 {meeting.get('title', '미팅')}",
        "ref_id": meeting["id"], "ref_kind": "meeting",
        "auto": True,
    }]


@app.route("/api/calendar", methods=["GET"])
def api_calendar():
    """모든 이벤트(캠페인 파생 + 미팅 파생 + 직접 추가) 합쳐서 반환."""
    out = []
    for c in load_campaigns():
        out.extend(_derive_campaign_events(c))
    for m in load_meetings():
        out.extend(_derive_meeting_events(m))
    for e in load_events():
        item = dict(e)
        item["auto"] = False
        out.append(item)
    # 날짜 기준 정렬
    out.sort(key=lambda x: (x.get("date", ""), x.get("time", "")))
    return jsonify({"events": out})


# ─── TODAY (오늘/이번주 할 일) ────────────────────────────
@app.route("/api/today", methods=["GET"])
def api_today():
    today = datetime.now().date()
    from datetime import timedelta
    week_end = today + timedelta(days=7)
    urgent, this_week, overdue, undated = [], [], [], []

    for c in load_campaigns():
        if c.get("status") in ("완료",):
            continue
        ls = c.get("live_start") or ""
        seller = c.get("seller_name", "")
        rd = c.get("round", 1)
        stage_label = STAGE_LABEL.get(c.get("stage", ""), c.get("stage", ""))
        if not ls:
            undated.append({
                "campaign_id": c["id"],
                "title": f"{seller} {rd}차",
                "stage": stage_label,
                "status": c.get("status", ""),
            })
            continue
        try:
            d = datetime.strptime(ls, "%Y-%m-%d").date()
        except ValueError:
            continue
        days = (d - today).days
        item = {
            "campaign_id": c["id"],
            "title": f"{seller} {rd}차 라이브 시작",
            "live_start": ls,
            "days": days,
            "stage": stage_label,
            "status": c.get("status", ""),
        }
        if days < 0 and c.get("stage") != "complete":
            overdue.append(item)
        elif days <= 3:
            urgent.append(item)
        elif days <= 14:
            this_week.append(item)

    return jsonify({
        "today": today.isoformat(),
        "urgent": urgent,
        "this_week": this_week,
        "overdue": overdue,
        "undated": undated,
    })


# ─── MEETINGS (미팅 + 녹취 분석) ──────────────────────────
MEETING_FIELDS = (
    "campaign_id", "title", "date", "time",
    "attendees", "agenda", "manual_notes",
    "transcript", "summary", "decisions",
    "action_items", "key_points", "follow_up_topics",
)


@app.route("/api/meetings", methods=["GET"])
def api_meetings_list():
    return jsonify({"meetings": load_meetings()})


@app.route("/api/meetings", methods=["POST"])
def api_meetings_add():
    payload = request.get_json(force=True)
    items = load_meetings()
    now = datetime.now().isoformat(timespec="seconds")
    new = {
        "id": _next_id(items, "m"),
        "campaign_id": payload.get("campaign_id") or "",
        "title": (payload.get("title") or "").strip() or "(제목 없음)",
        "date": (payload.get("date") or "").strip(),
        "time": (payload.get("time") or "").strip(),
        "attendees": payload.get("attendees") or [],
        "agenda": (payload.get("agenda") or "").strip(),
        "audio_file": "",
        "transcript": "",
        "summary": "",
        "decisions": [],
        "action_items": [],
        "key_points": [],
        "follow_up_topics": [],
        "manual_notes": "",
        "created_at": now,
        "updated_at": now,
        "analysis_status": "none",
    }
    items.append(new)
    save_meetings(items)
    return jsonify({"meeting": new})


@app.route("/api/meetings/<mid>", methods=["PATCH"])
def api_meetings_patch(mid: str):
    payload = request.get_json(force=True)
    items = load_meetings()
    m = next((x for x in items if x["id"] == mid), None)
    if not m:
        return jsonify({"error": "not found"}), 404
    for k in MEETING_FIELDS:
        if k in payload:
            v = payload[k]
            if isinstance(v, str):
                v = v.strip()
            m[k] = v
    m["updated_at"] = datetime.now().isoformat(timespec="seconds")
    save_meetings(items)
    return jsonify({"meeting": m})


@app.route("/api/meetings/<mid>", methods=["DELETE"])
def api_meetings_delete(mid: str):
    items = load_meetings()
    m = next((x for x in items if x["id"] == mid), None)
    if m and m.get("audio_file"):
        try:
            Path(m["audio_file"]).unlink(missing_ok=True)
        except Exception:
            pass
    items = [x for x in items if x["id"] != mid]
    save_meetings(items)
    return jsonify({"ok": True})


@app.route("/api/meetings/<mid>/upload", methods=["POST"])
def api_meeting_upload(mid: str):
    """녹취 업로드 → Drive 동기화 폴더로 자동 이관 + 이름 자동 변경.

    파일명 규칙: {YYYYMMDD}_{미팅제목}_{미팅ID}.{ext}
    Drive 폴더 없으면 fallback으로 data/meetings/ 에 저장.
    """
    items = load_meetings()
    m = next((x for x in items if x["id"] == mid), None)
    if not m:
        return jsonify({"error": "not found"}), 404
    f = request.files.get("audio")
    if not f:
        return jsonify({"error": "audio 파일 필수"}), 400

    # 확장자 정리
    ext = (Path(f.filename or "audio.m4a").suffix or ".m4a").lower()
    if ext not in (".m4a", ".mp3", ".wav", ".aac", ".ogg", ".flac", ".webm", ".mp4"):
        ext = ".m4a"

    # 의미있는 파일명 생성
    date_part = (m.get("date") or datetime.now().strftime("%Y-%m-%d")).replace("-", "")
    title_part = _safe_filename(m.get("title", ""), fallback="미팅")
    new_name = f"{date_part}_{title_part}_{mid}{ext}"

    # 저장 위치: Drive 폴더 우선, fallback 로컬
    drive_dir = _find_drive_meeting_dir()
    if drive_dir:
        save_path = drive_dir / new_name
        m["audio_location"] = "drive"
    else:
        save_path = MEETINGS_DIR / new_name
        m["audio_location"] = "local"

    # 같은 이름 있으면 _2, _3... 붙임
    if save_path.exists():
        i = 2
        while True:
            cand = save_path.with_stem(f"{save_path.stem}_{i}")
            if not cand.exists():
                save_path = cand
                break
            i += 1

    try:
        f.save(str(save_path))
    except Exception as e:
        return jsonify({"error": f"파일 저장 실패: {e}"}), 500

    m["audio_file"] = str(save_path)
    m["audio_filename"] = save_path.name
    m["audio_folder"] = str(save_path.parent)
    m["analysis_status"] = "uploaded"
    m["updated_at"] = datetime.now().isoformat(timespec="seconds")
    save_meetings(items)
    return jsonify({"meeting": m, "saved_to": str(save_path), "drive_synced": drive_dir is not None})


@app.route("/api/meetings/open_folder", methods=["POST"])
def api_meeting_open_folder():
    """녹취 폴더를 탐색기로 열기."""
    import subprocess
    drive_dir = _find_drive_meeting_dir()
    target = drive_dir or MEETINGS_DIR
    if not target.exists():
        target.mkdir(parents=True, exist_ok=True)
    try:
        subprocess.Popen(["explorer.exe", str(target)])
        return jsonify({"ok": True, "path": str(target), "is_drive": drive_dir is not None})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/meetings/<mid>/analyze_text", methods=["POST"])
def api_meeting_analyze_text(mid: str):
    """클로바노트 등에서 받은 텍스트를 붙여넣어 Gemini가 요약/액션 추출."""
    payload = request.get_json(force=True)
    transcript = (payload.get("transcript") or "").strip()
    if not transcript:
        return jsonify({"error": "transcript 필수"}), 400

    items = load_meetings()
    m = next((x for x in items if x["id"] == mid), None)
    if not m:
        return jsonify({"error": "not found"}), 404

    job_id = uuid.uuid4().hex[:12]
    with JOB_LOCK:
        JOBS[job_id] = {
            "id": job_id, "kind": "meeting_analyze_text", "meeting_id": mid,
            "status": "running", "progress": 0, "total": 10,
            "message": "텍스트 분석 시작…",
            "started_at": datetime.now().isoformat(timespec="seconds"),
            "finished_at": None, "errors": [],
        }

    def run():
        try:
            from meeting_analyzer import MeetingAnalyzer  # type: ignore
            cfg = load_config()
            analyzer = MeetingAnalyzer(cfg)
            result = analyzer.analyze_text(transcript, on_progress=lambda **kw: update_job(job_id, **kw))
            if result.get("_error"):
                update_job(job_id, status="error", message=result["_error"],
                           finished_at=datetime.now().isoformat(timespec="seconds"))
                return
            items2 = load_meetings()
            mm = next((x for x in items2 if x["id"] == mid), None)
            if mm:
                mm["transcript"] = result.get("transcript", transcript)
                mm["summary"] = result.get("summary", "")
                mm["decisions"] = result.get("decisions", [])
                mm["action_items"] = result.get("action_items", [])
                mm["key_points"] = result.get("key_points", [])
                mm["follow_up_topics"] = result.get("follow_up_topics", [])
                mm["analysis_status"] = "done"
                mm["updated_at"] = datetime.now().isoformat(timespec="seconds")
                save_meetings(items2)
            update_job(job_id, status="done", message="분석 완료!",
                       progress=10, total=10,
                       finished_at=datetime.now().isoformat(timespec="seconds"))
        except Exception as e:  # noqa: BLE001
            log.exception("meeting analyze_text failed")
            update_job(job_id, status="error", message=f"에러: {e}",
                       finished_at=datetime.now().isoformat(timespec="seconds"))

    threading.Thread(target=run, daemon=True).start()

    m["analysis_status"] = "analyzing"
    save_meetings(items)
    return jsonify({"job_id": job_id})


@app.route("/api/meetings/<mid>/analyze", methods=["POST"])
def api_meeting_analyze(mid: str):
    items = load_meetings()
    m = next((x for x in items if x["id"] == mid), None)
    if not m:
        return jsonify({"error": "not found"}), 404
    if not m.get("audio_file"):
        return jsonify({"error": "오디오 파일 먼저 업로드"}), 400

    job_id = uuid.uuid4().hex[:12]
    with JOB_LOCK:
        JOBS[job_id] = {
            "id": job_id,
            "kind": "meeting_analyze",
            "meeting_id": mid,
            "status": "running",
            "progress": 0, "total": 10,
            "message": "분석 시작…",
            "started_at": datetime.now().isoformat(timespec="seconds"),
            "finished_at": None,
            "errors": [],
        }

    def run():
        try:
            from meeting_analyzer import MeetingAnalyzer  # type: ignore
            cfg = load_config()
            analyzer = MeetingAnalyzer(cfg)
            result = analyzer.analyze_audio(
                Path(m["audio_file"]),
                on_progress=lambda **kw: update_job(job_id, **kw),
            )
            if result.get("_error"):
                # 마킹
                items2 = load_meetings()
                mm = next((x for x in items2 if x["id"] == mid), None)
                if mm:
                    mm["analysis_status"] = "error"
                    mm["updated_at"] = datetime.now().isoformat(timespec="seconds")
                    save_meetings(items2)
                update_job(job_id, status="error", message=result["_error"],
                           finished_at=datetime.now().isoformat(timespec="seconds"))
                return

            # 미팅 업데이트
            items2 = load_meetings()
            mm = next((x for x in items2 if x["id"] == mid), None)
            if mm:
                mm["transcript"] = result.get("transcript", "")
                mm["summary"] = result.get("summary", "")
                mm["decisions"] = result.get("decisions", [])
                mm["action_items"] = result.get("action_items", [])
                mm["key_points"] = result.get("key_points", [])
                mm["follow_up_topics"] = result.get("follow_up_topics", [])
                mm["analysis_status"] = "done"
                mm["updated_at"] = datetime.now().isoformat(timespec="seconds")
                save_meetings(items2)

            update_job(job_id, status="done", message="분석 완료!",
                       progress=10, total=10,
                       finished_at=datetime.now().isoformat(timespec="seconds"))
        except Exception as e:  # noqa: BLE001
            log.exception("meeting analyze failed")
            update_job(job_id, status="error", message=f"에러: {e}",
                       finished_at=datetime.now().isoformat(timespec="seconds"))

    threading.Thread(target=run, daemon=True).start()

    # 미팅 상태 분석 중으로 마킹
    m["analysis_status"] = "analyzing"
    save_meetings(items)

    return jsonify({"job_id": job_id})


# ─── AI CHAT WIDGET (격리 모듈) ──────────────────────────
CHAT_FILE = DATA_DIR / "chat_history.json"
CHAT_UPLOADS_DIR = DATA_DIR / "chat_uploads"
CHAT_UPLOADS_DIR.mkdir(exist_ok=True)


def load_chat() -> dict:
    if not CHAT_FILE.exists():
        return {"messages": []}
    try:
        return json.loads(CHAT_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {"messages": []}


def save_chat(data: dict) -> None:
    CHAT_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


@app.route("/api/chat/messages", methods=["GET"])
def api_chat_messages():
    return jsonify(load_chat())


@app.route("/api/chat/new", methods=["POST"])
def api_chat_new():
    save_chat({"messages": []})
    return jsonify({"ok": True})


@app.route("/api/chat/send", methods=["POST"])
def api_chat_send():
    text = (request.form.get("text") or "").strip()
    image_file = request.files.get("image")

    if not text and not image_file:
        return jsonify({"error": "메시지 또는 이미지 필수"}), 400

    data = load_chat()

    # 이미지 저장
    image_url = None
    image_bytes = None
    image_mime = None
    if image_file:
        ext = (Path(image_file.filename or "img.png").suffix or ".png").lower()
        if ext not in (".png", ".jpg", ".jpeg", ".gif", ".webp"):
            ext = ".png"
        fname = f"img_{uuid.uuid4().hex[:10]}{ext}"
        fpath = CHAT_UPLOADS_DIR / fname
        image_file.save(str(fpath))
        try:
            image_bytes = fpath.read_bytes()
        except Exception:
            image_bytes = None
        image_mime = image_file.mimetype or "image/png"
        image_url = f"/chat_uploads/{fname}"

    # 사용자 메시지 누적 저장
    now = datetime.now().isoformat(timespec="seconds")
    user_msg = {"role": "user", "text": text, "image": image_url, "at": now}
    data["messages"].append(user_msg)
    save_chat(data)

    # Gemini history 형식 변환 (현재 메시지 빼고)
    history_for_gemini = []
    for m in data["messages"][:-1]:
        role = "user" if m["role"] == "user" else "model"
        history_for_gemini.append({"role": role, "parts": [m.get("text", "") or "(이미지)"]})

    # Gemini 호출 (자동 도구 호출)
    tool_calls = []
    try:
        from chat_agent import ChatAgent  # type: ignore
        agent = ChatAgent(load_config())
        result = agent.send(history_for_gemini, text,
                            image_bytes=image_bytes, image_mime=image_mime)
        if isinstance(result, dict):
            reply_text = result.get("text", "")
            tool_calls = result.get("tool_calls", [])
        else:
            reply_text = str(result)
    except Exception as e:  # noqa: BLE001
        log.exception("chat send failed")
        reply_text = f"❌ 에러: {e}"

    assistant_msg = {
        "role": "assistant",
        "text": reply_text,
        "image": None,
        "tool_calls": tool_calls,
        "at": datetime.now().isoformat(timespec="seconds"),
    }
    data["messages"].append(assistant_msg)
    save_chat(data)

    # 어떤 데이터가 바뀌었는지 클라이언트에게 알려줌 (UI 새로고침용)
    changed = set()
    for tc in tool_calls:
        n = tc.get("name", "")
        if "campaign" in n: changed.update({"campaigns", "calendar", "dashboard", "today"})
        if "event" in n: changed.add("calendar")
        if "meeting" in n: changed.update({"meetings", "calendar"})
        if "brand" in n: changed.add("brands")

    return jsonify({"reply": assistant_msg, "changed": list(changed)})


@app.route("/chat_uploads/<path:filename>")
def serve_chat_upload(filename: str):
    return send_from_directory(CHAT_UPLOADS_DIR, filename)


# ═══════════════════════════════════════════════════════════
# 📨 DM 영업 시스템 — v2 (인플루언서 5000 + 계정 100+)
# ═══════════════════════════════════════════════════════════

INFLUENCERS_FILE = DATA_DIR / "influencers.json"
OUR_ACCOUNTS_FILE = DATA_DIR / "our_accounts.json"
IMPORTS_DIR = DATA_DIR / "imports"
IMPORTS_DIR.mkdir(exist_ok=True)


def _load_influencers() -> list[dict]:
    if not INFLUENCERS_FILE.exists():
        return []
    return json.loads(INFLUENCERS_FILE.read_text(encoding="utf-8")).get("influencers", [])


def _save_influencers(items: list[dict]) -> None:
    INFLUENCERS_FILE.write_text(
        json.dumps({"influencers": items, "schema_version": 2,
                    "updated_at": datetime.now().isoformat(timespec="seconds")},
                   ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _load_our_accounts() -> list[dict]:
    if not OUR_ACCOUNTS_FILE.exists():
        return []
    return json.loads(OUR_ACCOUNTS_FILE.read_text(encoding="utf-8")).get("accounts", [])


def _save_our_accounts(items: list[dict]) -> None:
    OUR_ACCOUNTS_FILE.write_text(
        json.dumps({"accounts": items, "schema_version": 2,
                    "updated_at": datetime.now().isoformat(timespec="seconds")},
                   ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


# ─── 인플루언서 마스터 (5,000명) ─────────────────────────
@app.route("/api/influencers", methods=["GET"])
def api_influencers_list():
    items = _load_influencers()
    # 페이지네이션
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("page_size", 50))
    q = (request.args.get("q") or "").strip().lower()
    status_filter = (request.args.get("status") or "").strip()

    filtered = items
    if q:
        filtered = [it for it in filtered if
                    q in (it.get("instagram_id", "") or "").lower() or
                    q in (it.get("seller_name", "") or "").lower()]
    if status_filter:
        filtered = [it for it in filtered if it.get("status") == status_filter]

    total = len(filtered)
    start = (page - 1) * page_size
    end = start + page_size
    return jsonify({
        "influencers": filtered[start:end],
        "total": total,
        "page": page,
        "page_size": page_size,
        "summary": _influencer_summary(items),
    })


def _influencer_summary(items: list[dict]) -> dict:
    by_status: dict[str, int] = {}
    for it in items:
        s = it.get("status") or "미발송"
        by_status[s] = by_status.get(s, 0) + 1
    return {"total": len(items), "by_status": by_status}


@app.route("/api/influencers/<iid>", methods=["PATCH"])
def api_influencers_patch(iid: str):
    payload = request.get_json(force=True)
    items = _load_influencers()
    it = next((x for x in items if x["id"] == iid), None)
    if not it:
        return jsonify({"error": "not found"}), 404
    for k in ("seller_name", "status", "notes", "history_text"):
        if k in payload:
            it[k] = (payload[k] or "").strip()
    _save_influencers(items)
    return jsonify({"influencer": it})


@app.route("/api/influencers/<iid>", methods=["DELETE"])
def api_influencers_delete(iid: str):
    items = [x for x in _load_influencers() if x["id"] != iid]
    _save_influencers(items)
    return jsonify({"ok": True})


# ─── 우리 계정 마스터 (100+) ─────────────────────────────
@app.route("/api/our_accounts", methods=["GET"])
def api_our_accounts_list():
    items = _load_our_accounts()
    # 비번 마스킹
    out = []
    for a in items:
        masked = dict(a)
        if masked.get("login_pw"): masked["login_pw"] = "*" * 8
        if masked.get("linked_email_pw"): masked["linked_email_pw"] = "*" * 8
        out.append(masked)
    # 기기별 그룹 카운트
    by_device = {}
    for it in items:
        d = it.get("device") or "(미지정)"
        by_device[d] = by_device.get(d, 0) + 1
    by_status = {}
    for it in items:
        s = it.get("status") or "활성"
        by_status[s] = by_status.get(s, 0) + 1
    return jsonify({
        "accounts": out,
        "summary": {
            "total": len(items),
            "by_status": by_status,
            "by_device": by_device,
        }
    })


@app.route("/api/our_accounts/<aid>", methods=["PATCH"])
def api_our_accounts_patch(aid: str):
    payload = request.get_json(force=True)
    items = _load_our_accounts()
    a = next((x for x in items if x["id"] == aid), None)
    if not a:
        return jsonify({"error": "not found"}), 404
    for k in ("status", "daily_limit", "notes", "phone", "device", "account_owner"):
        if k in payload:
            v = payload[k]
            if k == "daily_limit":
                v = int(v or 50)
            elif isinstance(v, str):
                v = v.strip()
            a[k] = v
    if "login_pw" in payload and payload["login_pw"]:
        a["login_pw"] = payload["login_pw"]
    _save_our_accounts(items)
    return jsonify({"account": {**a, "login_pw": "*" * 8}})


@app.route("/api/our_accounts/<aid>", methods=["DELETE"])
def api_our_accounts_delete(aid: str):
    items = [x for x in _load_our_accounts() if x["id"] != aid]
    _save_our_accounts(items)
    return jsonify({"ok": True})


# ─── 엑셀 자동 임포트 ────────────────────────────────────
@app.route("/api/dm/import/preview", methods=["GET"])
def api_dm_import_preview():
    """data/imports/ 폴더의 xlsx 파일 목록 + 미리보기 정보."""
    files = []
    if IMPORTS_DIR.exists():
        for p in sorted(IMPORTS_DIR.glob("*.xlsx")):
            files.append({
                "name": p.name,
                "size_kb": round(p.stat().st_size / 1024, 1),
                "modified": datetime.fromtimestamp(p.stat().st_mtime).isoformat(timespec="seconds"),
            })
    return jsonify({"files": files, "imports_dir": str(IMPORTS_DIR)})


@app.route("/api/dm/import/influencers", methods=["POST"])
def api_dm_import_influencers():
    payload = request.get_json(force=True, silent=True) or {}
    fname = (payload.get("filename") or "").strip()
    if not fname:
        return jsonify({"error": "filename 필수"}), 400
    fpath = IMPORTS_DIR / fname
    if not fpath.exists():
        return jsonify({"error": f"파일 없음: {fpath}"}), 404

    try:
        from dm_importer import import_influencers  # type: ignore
    except ImportError as e:
        return jsonify({"error": f"importer 모듈 실패: {e}"}), 500

    mode = (payload.get("mode") or "add").strip().lower()
    if mode not in ("add", "update"):
        mode = "add"

    existing = _load_influencers()
    result = import_influencers(fpath, existing, mode=mode)
    if result.get("error"):
        return jsonify(result), 400

    new_items = result.pop("items", [])
    existing.extend(new_items)
    _save_influencers(existing)

    return jsonify({**result, "mode": mode, "total_after": len(existing)})


@app.route("/api/dm/import/accounts", methods=["POST"])
def api_dm_import_accounts():
    payload = request.get_json(force=True, silent=True) or {}
    fname = (payload.get("filename") or "").strip()
    if not fname:
        return jsonify({"error": "filename 필수"}), 400
    fpath = IMPORTS_DIR / fname
    if not fpath.exists():
        return jsonify({"error": f"파일 없음: {fpath}"}), 404

    try:
        from dm_importer import import_accounts  # type: ignore
    except ImportError as e:
        return jsonify({"error": f"importer 모듈 실패: {e}"}), 500

    mode = (payload.get("mode") or "add").strip().lower()
    if mode not in ("add", "update"):
        mode = "add"

    existing = _load_our_accounts()
    result = import_accounts(fpath, existing, mode=mode)
    if result.get("error"):
        return jsonify(result), 400

    new_items = result.pop("items", [])
    existing.extend(new_items)
    _save_our_accounts(existing)

    return jsonify({**result, "mode": mode, "total_after": len(existing)})


@app.route("/api/dm/import/template/<kind>", methods=["GET"])
def api_dm_import_template(kind):
    """양식 엑셀 다운로드. kind = 'influencers' | 'accounts'"""
    tpl_dir = IMPORTS_DIR / "templates"
    mapping = {
        "influencers": "influencers_template.xlsx",
        "accounts": "accounts_template.xlsx",
    }
    fname = mapping.get(kind)
    if not fname:
        return jsonify({"error": "kind는 influencers / accounts"}), 400
    fpath = tpl_dir / fname
    if not fpath.exists():
        return jsonify({
            "error": "양식 파일 없음",
            "hint": "python scripts/make_templates.py 실행",
        }), 404
    return send_from_directory(tpl_dir, fname, as_attachment=True)


# ═══════════════════════════════════════════════════════════
# 🚀 Phase B — 발송 큐 (오늘 보낼 후보 자동 산출)
# ═══════════════════════════════════════════════════════════

DM_TEMPLATES_V2_FILE = DATA_DIR / "dm_templates_v2.json"

# 라이브 발송 상태 (in-memory) — 서버 재시작 시 초기화
DM_LIVE_STATE = {
    "current": None,       # {influencer_handle, account_handle, started_at, message_preview}
    "log": [],             # [{ts, account, target, status, message}], 최근 200개
    "paused": False,
    "running": False,
}
DM_LIVE_LOG_MAX = 200


def _dm_live_emit(event: dict) -> None:
    """라이브 로그에 1건 추가 (오래된 항목 잘라냄)."""
    event["ts"] = datetime.now().isoformat(timespec="seconds")
    DM_LIVE_STATE["log"].insert(0, event)
    if len(DM_LIVE_STATE["log"]) > DM_LIVE_LOG_MAX:
        DM_LIVE_STATE["log"] = DM_LIVE_STATE["log"][:DM_LIVE_LOG_MAX]


def _load_dm_templates_v2() -> list[dict]:
    if not DM_TEMPLATES_V2_FILE.exists():
        return [{"id": "tpl_default", "name": "기본 1차", "body": "안녕하세요! 넥스트포트에서 공동구매 제안드립니다 :)"}]
    try:
        with open(DM_TEMPLATES_V2_FILE, encoding="utf-8") as f:
            data = json.load(f)
        return data.get("templates", [])
    except Exception:
        return []


def _save_dm_templates_v2(items: list[dict]) -> None:
    DM_TEMPLATES_V2_FILE.write_text(
        json.dumps({"templates": items, "updated_at": datetime.now().isoformat(timespec="seconds")},
                   ensure_ascii=False, indent=2), encoding="utf-8")


@app.route("/api/dm/queue/today", methods=["GET"])
def api_dm_queue_today():
    """오늘 발송 가능한 (인플루언서, 추천계정) 페어 산출."""
    try:
        from dm_scheduler import build_queue  # type: ignore
    except ImportError as e:
        return jsonify({"error": f"scheduler 모듈 실패: {e}"}), 500
    max_per_run = int(request.args.get("max", 100))
    influencers = _load_influencers()
    accounts = _load_our_accounts()
    return jsonify(build_queue(influencers, accounts, max_per_run=max_per_run))


@app.route("/api/dm/templates_v2", methods=["GET", "POST"])
def api_dm_templates_v2():
    if request.method == "POST":
        payload = request.get_json(force=True, silent=True) or {}
        items = payload.get("templates") or []
        _save_dm_templates_v2(items)
        return jsonify({"ok": True, "count": len(items)})
    return jsonify({"templates": _load_dm_templates_v2()})


@app.route("/api/dm/send", methods=["POST"])
def api_dm_send_one():
    """큐의 한 페어를 즉시 발송. body: {influencer_id, account_id, message}"""
    # 클라우드 환경에서는 발송 차단
    if (load_config() or {}).get("env_mode") == "cloud":
        return jsonify({"error": "클라우드 환경에서는 DM 발송 비활성화"}), 403

    payload = request.get_json(force=True, silent=True) or {}
    inf_id = payload.get("influencer_id")
    acc_id = payload.get("account_id")
    message = (payload.get("message") or "").strip()
    template_id = (payload.get("template_id") or "").strip()
    if not (inf_id and acc_id and message):
        return jsonify({"error": "influencer_id, account_id, message 필수"}), 400

    influencers = _load_influencers()
    accounts = _load_our_accounts()
    inf = next((x for x in influencers if x["id"] == inf_id), None)
    acc = next((x for x in accounts if x["id"] == acc_id), None)
    if not inf or not acc:
        return jsonify({"error": "인플루언서 또는 계정 없음"}), 404

    try:
        from dm_sender import DMSender  # type: ignore
        from dm_scheduler import record_send  # type: ignore
        from dm_inbox import upsert_conversation, append_message  # type: ignore
    except ImportError as e:
        return jsonify({"error": f"모듈 실패: {e}"}), 500

    # 템플릿 통계 — sent_count +1 (즉시; ok/fail은 별도)
    if template_id:
        try:
            tpls = _load_dm_templates_v2()
            tpl = next((t for t in tpls if t.get("id") == template_id), None)
            if tpl:
                tpl["sent_count"] = (tpl.get("sent_count") or 0) + 1
                _save_dm_templates_v2(tpls)
        except Exception as e:
            log.warning(f"템플릿 통계 갱신 실패: {e}")

    # 라이브 상태 업데이트 — UI 가 polling
    DM_LIVE_STATE["current"] = {
        "influencer_handle": inf.get("instagram_id"),
        "seller_name": inf.get("seller_name"),
        "account_handle": acc.get("instagram_id"),
        "started_at": datetime.now().isoformat(timespec="seconds"),
        "message_preview": message[:80],
        "send_count": (inf.get("send_count") or 0) + 1,
    }
    DM_LIVE_STATE["running"] = True
    _dm_live_emit({
        "type": "start",
        "account": acc.get("instagram_id"),
        "target": inf.get("instagram_id"),
        "message": message[:80],
    })

    sender = DMSender()
    ok = False
    err_msg = ""
    try:
        from playwright.sync_api import sync_playwright  # type: ignore
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=False)
            ctx = browser.new_context()
            page = ctx.new_page()
            if not sender.login(page, acc):
                err_msg = "로그인 실패"
            else:
                ok, err_msg = sender.send_dm(page, {"instagram_id": inf["instagram_id"]}, message)
            ctx.close()
            browser.close()
    except ImportError:
        DM_LIVE_STATE["current"] = None
        DM_LIVE_STATE["running"] = False
        return jsonify({"error": "Playwright 미설치"}), 500
    except Exception as e:
        err_msg = str(e)

    # 라이브 로그 — 종료
    _dm_live_emit({
        "type": "done",
        "account": acc.get("instagram_id"),
        "target": inf.get("instagram_id"),
        "status": "ok" if ok else "fail",
        "error": err_msg if not ok else "",
    })
    DM_LIVE_STATE["current"] = None
    DM_LIVE_STATE["running"] = False

    # 결과 기록
    status = "ok" if ok else "fail"
    record_send(inf, acc, message, status=status, note=err_msg)
    if ok:
        # 답장 회수 매칭용 — 마지막 발송 정보 박음
        inf["last_sent_template_id"] = template_id
        inf["last_sent_message_full"] = message
        # 계정에도 누적 발송
        acc["sent_count_total"] = (acc.get("sent_count_total") or 0) + 1
    _save_influencers(influencers)
    _save_our_accounts(accounts)

    # 인박스에도 우리가 보낸 메시지로 기록 (template_id 포함)
    if ok:
        inbox = _load_inbox()
        conv = upsert_conversation(inbox, acc, inf["instagram_id"], inf)
        append_message(conv, "us", message)
        # 우리 측 메시지에 template_id 박음 (답장 매칭용)
        if conv.get("messages"):
            conv["messages"][-1]["template_id"] = template_id
        conv["last_sent_template_id"] = template_id
        _save_inbox(inbox)

    return jsonify({"ok": ok, "error": err_msg if not ok else None, "send_count": inf.get("send_count")})


# ═══════════════════════════════════════════════════════════
# 📊 회신율 통계 — 템플릿별 / 계정별
# ═══════════════════════════════════════════════════════════

def _recompute_reply_stats():
    """모든 conversation을 훑어서 템플릿별/계정별 회신율 재계산.
    인플루언서마다 1차례라도 답장이 있으면 그 인플루언서의 last_sent_template_id로 +1.
    답장이 온 conversation의 our_account_id로 계정 카운트 +1."""
    influencers = _load_influencers()
    inf_by_handle = {x.get("instagram_id", "").lower(): x for x in influencers}
    accounts = _load_our_accounts()
    acc_by_id = {a["id"]: a for a in accounts}
    templates = _load_dm_templates_v2()
    tpl_by_id = {t["id"]: t for t in templates}

    # 카운터 초기화
    for t in templates:
        t["reply_count"] = 0
    for a in accounts:
        a["reply_count"] = 0

    conversations = _load_inbox()
    for c in conversations:
        has_them = any(m.get("from") == "them" for m in c.get("messages", []))
        if not has_them:
            continue
        # template 통계 (인플루언서의 last_sent_template_id 또는 conv의 last_sent_template_id)
        handle = (c.get("their_handle") or "").lower()
        inf = inf_by_handle.get(handle)
        tpl_id = (inf.get("last_sent_template_id") if inf else "") or c.get("last_sent_template_id") or ""
        tpl = tpl_by_id.get(tpl_id)
        if tpl:
            tpl["reply_count"] = (tpl.get("reply_count") or 0) + 1
        # 계정 통계
        acc_id = c.get("our_account_id")
        acc = acc_by_id.get(acc_id)
        if acc:
            acc["reply_count"] = (acc.get("reply_count") or 0) + 1

    # 비율 계산
    for t in templates:
        sent = t.get("sent_count") or 0
        rep = t.get("reply_count") or 0
        t["reply_rate"] = round(rep / sent * 100, 1) if sent else 0
    for a in accounts:
        sent = a.get("sent_count_total") or a.get("total_sent") or 0
        rep = a.get("reply_count") or 0
        a["reply_rate"] = round(rep / sent * 100, 1) if sent else 0

    _save_dm_templates_v2(templates)
    _save_our_accounts(accounts)


@app.route("/api/dm/stats/recompute", methods=["POST"])
def api_dm_stats_recompute():
    _recompute_reply_stats()
    return jsonify({"ok": True})


@app.route("/api/dm/stats/templates", methods=["GET"])
def api_dm_stats_templates():
    """템플릿별 회신율 + 발송수."""
    templates = _load_dm_templates_v2()
    # 자동 재계산 — 매 조회시 동기화 (느리면 캐시)
    _recompute_reply_stats()
    templates = _load_dm_templates_v2()
    return jsonify({
        "templates": sorted(templates, key=lambda t: -(t.get("reply_rate") or 0)),
    })


@app.route("/api/dm/stats/accounts", methods=["GET"])
def api_dm_stats_accounts():
    """계정별 회신율 + 발송수 (회신 많이 오는 계정 분석)."""
    accounts = _load_our_accounts()
    return jsonify({
        "accounts": [
            {
                "id": a["id"],
                "instagram_id": a.get("instagram_id"),
                "device": a.get("device"),
                "owner": a.get("account_owner"),
                "sent_count": a.get("sent_count_total") or a.get("total_sent") or 0,
                "reply_count": a.get("reply_count") or 0,
                "reply_rate": a.get("reply_rate") or 0,
                "status": a.get("status"),
            }
            for a in sorted(accounts, key=lambda a: -(a.get("reply_rate") or 0))
        ],
    })


# ═══════════════════════════════════════════════════════════
# 📥 Phase C — 통합 DM 답장 인박스
# ═══════════════════════════════════════════════════════════

INBOX_FILE = DATA_DIR / "inbox_messages.json"


def _load_inbox() -> list[dict]:
    if not INBOX_FILE.exists():
        return []
    try:
        with open(INBOX_FILE, encoding="utf-8") as f:
            return json.load(f).get("conversations", [])
    except Exception:
        return []


def _save_inbox(conversations: list[dict]) -> None:
    INBOX_FILE.write_text(json.dumps({
        "conversations": conversations,
        "schema_version": 1,
        "synced_at": datetime.now().isoformat(timespec="seconds"),
    }, ensure_ascii=False, indent=2), encoding="utf-8")


@app.route("/api/dm/inbox", methods=["GET"])
def api_dm_inbox_list():
    try:
        from dm_inbox import list_conversations, summarize  # type: ignore
    except ImportError as e:
        return jsonify({"error": f"인박스 모듈 실패: {e}"}), 500
    conversations = _load_inbox()
    result = list_conversations(
        conversations,
        q=request.args.get("q", "").strip(),
        only_unread=request.args.get("only_unread", "").lower() in ("1", "true", "yes"),
        account_id=request.args.get("account_id", "").strip(),
        page=int(request.args.get("page", 1)),
        page_size=int(request.args.get("page_size", 50)),
    )
    result["summary"] = summarize(conversations)
    return jsonify(result)


@app.route("/api/dm/inbox/<conv_id>", methods=["GET"])
def api_dm_inbox_conversation(conv_id):
    conversations = _load_inbox()
    conv = next((c for c in conversations if c.get("id") == conv_id), None)
    if not conv:
        return jsonify({"error": "대화 없음"}), 404
    return jsonify(conv)


@app.route("/api/dm/inbox/<conv_id>/read", methods=["POST"])
def api_dm_inbox_mark_read(conv_id):
    try:
        from dm_inbox import mark_read  # type: ignore
    except ImportError as e:
        return jsonify({"error": f"인박스 모듈 실패: {e}"}), 500
    conversations = _load_inbox()
    conv = next((c for c in conversations if c.get("id") == conv_id), None)
    if not conv:
        return jsonify({"error": "대화 없음"}), 404
    mark_read(conv)
    _save_inbox(conversations)
    return jsonify({"ok": True})


@app.route("/api/dm/inbox/<conv_id>/reply", methods=["POST"])
def api_dm_inbox_reply(conv_id):
    """답장 보내기 — 그 conversation 의 our_account 로."""
    if (load_config() or {}).get("env_mode") == "cloud":
        return jsonify({"error": "클라우드 환경에서는 DM 발송 비활성화"}), 403

    payload = request.get_json(force=True, silent=True) or {}
    message = (payload.get("message") or "").strip()
    if not message:
        return jsonify({"error": "message 필수"}), 400

    conversations = _load_inbox()
    conv = next((c for c in conversations if c.get("id") == conv_id), None)
    if not conv:
        return jsonify({"error": "대화 없음"}), 404

    accounts = _load_our_accounts()
    acc = next((a for a in accounts if a["id"] == conv.get("our_account_id")), None)
    if not acc:
        return jsonify({"error": "발송 계정 없음"}), 404

    try:
        from dm_sender import DMSender  # type: ignore
        from dm_inbox import append_message  # type: ignore
        from playwright.sync_api import sync_playwright  # type: ignore
    except ImportError as e:
        return jsonify({"error": f"모듈 실패: {e}"}), 500

    sender = DMSender()
    ok = False
    err_msg = ""
    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=False)
            ctx = browser.new_context()
            page = ctx.new_page()
            if not sender.login(page, acc):
                err_msg = "로그인 실패"
            else:
                ok, err_msg = sender.send_dm(page, {"instagram_id": conv["their_handle"]}, message)
            ctx.close()
            browser.close()
    except Exception as e:
        err_msg = str(e)

    if ok:
        append_message(conv, "us", message)
        _save_inbox(conversations)

    return jsonify({"ok": ok, "error": err_msg if not ok else None})


@app.route("/api/dm/inbox/sync", methods=["POST"])
def api_dm_inbox_sync():
    """모든 활성 계정에서 받은 DM 동기화 (Playwright). 무거움 — 백그라운드 실행 권장.
    body 옵션: {account_ids: ["acc0001",...]} 비우면 전체 활성 계정
    """
    if (load_config() or {}).get("env_mode") == "cloud":
        return jsonify({"error": "클라우드 환경에서는 비활성화"}), 403

    payload = request.get_json(silent=True) or {}
    filter_ids = set(payload.get("account_ids") or [])
    accounts = _load_our_accounts()
    influencers = _load_influencers()
    inf_by_handle = {x.get("instagram_id", "").lower(): x for x in influencers}
    targets = [a for a in accounts if a.get("status") == "활성"]
    if filter_ids:
        targets = [a for a in targets if a["id"] in filter_ids]

    if not targets:
        return jsonify({"error": "활성 계정 없음"}), 400

    try:
        from dm_sender import DMSender  # type: ignore
        from dm_inbox import upsert_conversation, append_message  # type: ignore
        from playwright.sync_api import sync_playwright  # type: ignore
    except ImportError as e:
        return jsonify({"error": f"모듈 실패: {e}"}), 500

    conversations = _load_inbox()
    total_new = 0
    per_account = []
    sender = DMSender()

    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            for acc in targets:
                ctx = browser.new_context()
                page = ctx.new_page()
                new_for_acc = 0
                try:
                    if not sender.login(page, acc):
                        per_account.append({"account": acc.get("instagram_id"), "error": "로그인 실패"})
                        ctx.close()
                        continue
                    # 인박스 페이지로 가서 최근 대화 목록 긁기
                    page.goto("https://www.instagram.com/direct/inbox/", timeout=30000)
                    page.wait_for_timeout(3000)
                    threads = page.locator('div[role="listitem"]').all()
                    for t in threads[:20]:
                        try:
                            handle = t.locator('span').first.text_content() or ""
                            preview = t.locator('span').nth(1).text_content() or ""
                            handle = handle.strip()
                            if not handle:
                                continue
                            inf = inf_by_handle.get(handle.lower())
                            conv = upsert_conversation(conversations, acc, handle, inf)
                            # 마지막 미리보기와 다르면 새 메시지로 간주
                            if preview and preview != conv.get("last_message_preview"):
                                append_message(conv, "them", preview)
                                new_for_acc += 1
                        except Exception:
                            continue
                    per_account.append({"account": acc.get("instagram_id"), "new": new_for_acc})
                    total_new += new_for_acc
                except Exception as e:
                    per_account.append({"account": acc.get("instagram_id"), "error": str(e)})
                finally:
                    ctx.close()
            browser.close()
    except Exception as e:
        return jsonify({"error": f"브라우저 실패: {e}"}), 500

    _save_inbox(conversations)
    return jsonify({"ok": True, "total_new": total_new, "per_account": per_account})


# ═══════════════════════════════════════════════════════════
# ▶️ Phase D — 라이브 발송 상태 & 회신 인플루언서 현황 & 데일리 통계
# ═══════════════════════════════════════════════════════════

@app.route("/api/dm/live", methods=["GET"])
def api_dm_live():
    """현재 발송 중 상태 + 최근 로그. UI가 2초 polling."""
    return jsonify({
        "current": DM_LIVE_STATE.get("current"),
        "running": DM_LIVE_STATE.get("running", False),
        "paused": DM_LIVE_STATE.get("paused", False),
        "log": DM_LIVE_STATE.get("log", [])[:100],
        "log_count": len(DM_LIVE_STATE.get("log", [])),
    })


@app.route("/api/dm/live/clear", methods=["POST"])
def api_dm_live_clear():
    DM_LIVE_STATE["log"] = []
    return jsonify({"ok": True})


@app.route("/api/dm/live/pause", methods=["POST"])
def api_dm_live_pause():
    DM_LIVE_STATE["paused"] = not DM_LIVE_STATE.get("paused", False)
    return jsonify({"paused": DM_LIVE_STATE["paused"]})


@app.route("/api/dm/replies", methods=["GET"])
def api_dm_replies():
    """답장 받은 인플루언서 현황 — 인박스 + 인플루언서 매칭 + 전체 컬럼."""
    q = request.args.get("q", "").strip().lower()
    status_filter = request.args.get("status", "").strip()
    conversations = _load_inbox()
    influencers = _load_influencers()
    accounts = _load_our_accounts()
    inf_by_handle = {x.get("instagram_id", "").lower(): x for x in influencers}
    acc_by_handle = {a.get("instagram_id", "").lower(): a for a in accounts}

    rows = []
    seen = set()
    for c in conversations:
        has_them = any(m.get("from") == "them" for m in c.get("messages", []))
        if not has_them:
            continue
        handle = (c.get("their_handle") or "").lower()
        key = (handle, c.get("our_account_id"))
        if key in seen:
            continue
        seen.add(key)
        inf = inf_by_handle.get(handle) or {}
        our_handle = c.get("our_account_handle") or ""
        our_acc = acc_by_handle.get(our_handle.lower()) or {}

        row = {
            "conv_id": c.get("id"),
            "influencer_id": inf.get("id"),
            "influencer_handle": c.get("their_handle"),
            "influencer_url": inf.get("url") or f"https://www.instagram.com/{c.get('their_handle','')}/",
            "seller_name": c.get("seller_name") or inf.get("seller_name") or "",
            # ─── 스크린샷 컬럼 ───
            "first_reply_date": inf.get("first_reply_date") or (c.get("last_message_at") or "")[:10],
            "status": inf.get("status") or "dm 소통중",
            "owner": inf.get("owner") or "",
            "follower_count": inf.get("follower_count") or "",
            "reply_account": inf.get("reply_account") or our_handle,
            "device": our_acc.get("device") or "",
            "email": inf.get("email") or "",
            "phone": inf.get("phone") or "",
            "kakao_id": inf.get("kakao_id") or "",
            "notes": inf.get("notes") or "",
            # ─── 발송 정보 ───
            "send_count": inf.get("send_count"),
            "last_sent_date": inf.get("last_sent_date"),
            "our_account_handle": our_handle,
            "last_reply_at": c.get("last_message_at"),
            "last_message_preview": c.get("last_message_preview"),
            "unread_count": c.get("unread_count") or 0,
            "pipeline_stage": inf.get("pipeline_stage") or "",
        }
        if status_filter and row["status"] != status_filter:
            continue
        if q:
            blob = " ".join(str(v or "").lower() for v in row.values())
            if q not in blob:
                continue
        rows.append(row)

    rows.sort(key=lambda r: r.get("last_reply_at") or "", reverse=True)
    return jsonify({"replies": rows, "total": len(rows)})


@app.route("/api/dm/replies/<inf_id>", methods=["PATCH"])
def api_dm_replies_patch(inf_id):
    """회신 현황 표에서 인라인 편집 → influencer record 업데이트.
    status가 '미팅 fix' / '회신중' / '카톡 소통중' 으로 바뀌면 pipeline_stage 자동 전이."""
    payload = request.get_json(force=True, silent=True) or {}
    allowed = {"status", "owner", "first_reply_date", "reply_account", "email",
               "phone", "kakao_id", "notes", "follower_count", "pipeline_stage"}
    influencers = _load_influencers()
    inf = next((x for x in influencers if x.get("id") == inf_id), None)
    if not inf:
        return jsonify({"error": "인플루언서 없음"}), 404
    for k, v in payload.items():
        if k in allowed:
            inf[k] = v

    # 자동 파이프라인 전이 — status 변경 시
    new_status = payload.get("status")
    if new_status:
        cur_stage = inf.get("pipeline_stage") or ""
        if new_status in ("회신중", "dm 소통중", "카톡 소통중") and not cur_stage:
            inf["pipeline_stage"] = "진행예정"
        elif new_status == "미팅 fix" and cur_stage in ("", "진행예정"):
            inf["pipeline_stage"] = "미팅예약"
        elif new_status == "컨펌":
            inf["pipeline_stage"] = "캠페인진행중"
        elif new_status == "이탈":
            inf["pipeline_stage"] = "종료"

    _save_influencers(influencers)
    return jsonify({"ok": True, "influencer": inf, "auto_stage": inf.get("pipeline_stage")})


@app.route("/api/dm/replies/<inf_id>/meeting", methods=["POST"])
def api_dm_replies_add_meeting(inf_id):
    """회신 현황에서 [📅 미팅 박기] → 인플루언서.meetings 추가 + pipeline=미팅예약 + 캘린더 등록."""
    payload = request.get_json(force=True, silent=True) or {}
    date = (payload.get("date") or "").strip()
    note = (payload.get("note") or "").strip()
    if not date:
        return jsonify({"error": "date 필수"}), 400

    influencers = _load_influencers()
    inf = next((x for x in influencers if x.get("id") == inf_id), None)
    if not inf:
        return jsonify({"error": "인플루언서 없음"}), 404

    meetings = inf.setdefault("meetings", [])
    meetings.append({
        "date": date,
        "round": len(meetings) + 1,
        "note": note,
        "transcript": "",
        "audio_file": "",
        "created_at": datetime.now().isoformat(timespec="seconds"),
    })
    inf["pipeline_stage"] = inf.get("pipeline_stage") or "미팅예약"
    inf["status"] = "미팅 fix"
    _save_influencers(influencers)

    # 캘린더 자동
    try:
        events = load_events()
        events.append({
            "id": f"meet_{inf_id}_{int(datetime.now().timestamp())}",
            "title": f"[{inf.get('seller_name') or inf.get('instagram_id')}] {len(meetings)}차 미팅",
            "date": date,
            "type": "meeting",
            "linked_influencer_id": inf_id,
            "note": note,
        })
        save_events(events)
    except Exception as e:
        log.warning(f"미팅 캘린더 등록 실패: {e}")

    return jsonify({"ok": True, "meeting_round": len(meetings)})


@app.route("/api/pipeline/<inf_id>/meeting/<int:idx>", methods=["PATCH", "DELETE"])
def api_pipeline_meeting_one(inf_id, idx):
    """진행 예정 셀러의 특정 미팅 수정/삭제 + 녹취/메모."""
    influencers = _load_influencers()
    inf = next((x for x in influencers if x.get("id") == inf_id), None)
    if not inf:
        return jsonify({"error": "인플루언서 없음"}), 404
    meetings = inf.get("meetings") or []
    if idx < 0 or idx >= len(meetings):
        return jsonify({"error": "미팅 인덱스 범위 초과"}), 404

    if request.method == "DELETE":
        meetings.pop(idx)
        for i, m in enumerate(meetings):
            m["round"] = i + 1
        _save_influencers(influencers)
        return jsonify({"ok": True})

    payload = request.get_json(force=True, silent=True) or {}
    for k in ("date", "note", "transcript", "audio_file", "outcome"):
        if k in payload:
            meetings[idx][k] = payload[k]
    _save_influencers(influencers)
    return jsonify({"ok": True, "meeting": meetings[idx]})


@app.route("/api/pipeline/<inf_id>/meeting/<int:idx>/audio", methods=["POST"])
def api_pipeline_meeting_audio(inf_id, idx):
    """녹취 파일 업로드 (m4a/mp3/wav) → data/meetings/ 에 저장 + meetings[idx].audio_file에 경로."""
    influencers = _load_influencers()
    inf = next((x for x in influencers if x.get("id") == inf_id), None)
    if not inf:
        return jsonify({"error": "인플루언서 없음"}), 404
    meetings = inf.get("meetings") or []
    if idx < 0 or idx >= len(meetings):
        return jsonify({"error": "미팅 인덱스 범위 초과"}), 404

    if "file" not in request.files:
        return jsonify({"error": "file 필드 필수"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "파일명 없음"}), 400

    safe_ext = Path(f.filename).suffix.lower() or ".m4a"
    if safe_ext not in (".m4a", ".mp3", ".wav", ".ogg"):
        return jsonify({"error": "지원 형식: m4a/mp3/wav/ogg"}), 400

    meetings_dir = DATA_DIR / "meetings"
    meetings_dir.mkdir(exist_ok=True)
    fname = f"{inf.get('seller_name') or inf_id}_{meetings[idx].get('date','')}_m{idx+1}{safe_ext}"
    fname = re.sub(r"[\\/:*?\"<>|]+", "_", fname)
    fpath = meetings_dir / fname
    f.save(fpath)

    meetings[idx]["audio_file"] = str(fpath)
    _save_influencers(influencers)
    return jsonify({"ok": True, "path": str(fpath), "size_kb": round(fpath.stat().st_size / 1024, 1)})


@app.route("/api/pipeline/<inf_id>/detail", methods=["GET"])
def api_pipeline_detail(inf_id):
    """진행 예정 셀러 상세 — 모든 정보 + 미팅 리스트."""
    influencers = _load_influencers()
    inf = next((x for x in influencers if x.get("id") == inf_id), None)
    if not inf:
        return jsonify({"error": "인플루언서 없음"}), 404
    return jsonify(inf)


# ═══════════════════════════════════════════════════════════
# 🎯 진행 예정 셀러 (Pipeline)
# ═══════════════════════════════════════════════════════════

PIPELINE_STAGES = ["진행예정", "미팅예약", "미팅완료", "캠페인진행중", "종료"]


@app.route("/api/pipeline", methods=["GET"])
def api_pipeline_list():
    """pipeline_stage가 빈 값이 아닌 인플루언서만 반환."""
    q = request.args.get("q", "").strip().lower()
    stage_filter = request.args.get("stage", "").strip()
    influencers = _load_influencers()

    rows = []
    counts = {s: 0 for s in PIPELINE_STAGES}
    for inf in influencers:
        stage = inf.get("pipeline_stage") or ""
        if not stage:
            continue
        counts[stage] = counts.get(stage, 0) + 1
        if stage_filter and stage != stage_filter:
            continue
        meetings = inf.get("meetings") or []
        last_meeting = meetings[-1] if meetings else None
        row = {
            "influencer_id": inf.get("id"),
            "instagram_id": inf.get("instagram_id"),
            "url": inf.get("url"),
            "seller_name": inf.get("seller_name"),
            "follower_count": inf.get("follower_count"),
            "owner": inf.get("owner"),
            "pipeline_stage": stage,
            "meeting_count": len(meetings),
            "last_meeting_date": last_meeting.get("date") if last_meeting else None,
            "next_action": inf.get("next_action") or "",
            "campaign_id": inf.get("campaign_id"),
            "campaign_name": inf.get("campaign_name"),
        }
        if q:
            blob = " ".join(str(v or "").lower() for v in row.values())
            if q not in blob:
                continue
        rows.append(row)

    rows.sort(key=lambda r: (r.get("last_meeting_date") or "", r.get("seller_name") or ""), reverse=True)
    return jsonify({
        "pipeline": rows,
        "total": len(rows),
        "counts": counts,
        "stages": PIPELINE_STAGES,
    })


@app.route("/api/pipeline/manual", methods=["POST"])
def api_pipeline_manual_add():
    """수동으로 진행 예정 셀러 추가 — 인플루언서 DB에 없어도 박을 수 있음."""
    payload = request.get_json(force=True, silent=True) or {}
    seller_name = (payload.get("seller_name") or "").strip()
    instagram_id = (payload.get("instagram_id") or "").strip().lstrip("@")
    if not seller_name and not instagram_id:
        return jsonify({"error": "셀러명 또는 인스타ID 중 1개 필수"}), 400

    influencers = _load_influencers()
    # 인스타ID 중복이면 기존 record 사용
    existing = None
    if instagram_id:
        existing = next((x for x in influencers
                         if (x.get("instagram_id") or "").lower() == instagram_id.lower()), None)
    if existing:
        existing["pipeline_stage"] = payload.get("pipeline_stage") or "진행예정"
        for k in ("seller_name", "follower_count", "owner", "email", "phone",
                  "kakao_id", "notes", "category"):
            if payload.get(k):
                existing[k] = payload[k]
        _save_influencers(influencers)
        return jsonify({"ok": True, "influencer": existing, "created": False})

    nums = []
    for it in influencers:
        m = re.match(r"inf(\d+)", str(it.get("id", "")))
        if m:
            nums.append(int(m.group(1)))
    new_id = f"inf{(max(nums) if nums else 0) + 1:05d}"

    new_inf = {
        "id": new_id,
        "instagram_id": instagram_id or f"manual_{new_id}",
        "url": f"https://www.instagram.com/{instagram_id}/" if instagram_id else "",
        "seller_name": seller_name,
        "status": payload.get("status") or "수동등록",
        "follower_count": payload.get("follower_count") or "",
        "category": payload.get("category") or "",
        "owner": payload.get("owner") or "",
        "email": payload.get("email") or "",
        "phone": payload.get("phone") or "",
        "kakao_id": payload.get("kakao_id") or "",
        "notes": payload.get("notes") or "수동 추가",
        "first_reply_date": "",
        "reply_account": "",
        "last_sent_date": "",
        "last_sent_account_id": "",
        "send_count": 0,
        "history_text": "",
        "history": [],
        "used_account_ids": [],
        "reply_received": False,
        "last_reply_at": None,
        "pipeline_stage": payload.get("pipeline_stage") or "진행예정",
        "meetings": [],
        "imported_at": datetime.now().isoformat(timespec="seconds"),
        "manually_added": True,
    }
    influencers.append(new_inf)
    _save_influencers(influencers)
    return jsonify({"ok": True, "influencer": new_inf, "created": True})


@app.route("/api/pipeline/<inf_id>", methods=["PATCH"])
def api_pipeline_patch(inf_id):
    """파이프라인 단계 변경 / next_action 수정."""
    payload = request.get_json(force=True, silent=True) or {}
    allowed = {"pipeline_stage", "next_action", "owner", "campaign_id", "campaign_name"}
    influencers = _load_influencers()
    inf = next((x for x in influencers if x.get("id") == inf_id), None)
    if not inf:
        return jsonify({"error": "인플루언서 없음"}), 404
    for k, v in payload.items():
        if k in allowed:
            inf[k] = v
    _save_influencers(influencers)
    return jsonify({"ok": True, "influencer": inf})


@app.route("/api/pipeline/<inf_id>/meeting", methods=["POST"])
def api_pipeline_add_meeting(inf_id):
    """미팅 1건 추가 + 캘린더에도 박기."""
    payload = request.get_json(force=True, silent=True) or {}
    date = (payload.get("date") or "").strip()
    note = (payload.get("note") or "").strip()
    if not date:
        return jsonify({"error": "date 필수 (YYYY-MM-DD)"}), 400

    influencers = _load_influencers()
    inf = next((x for x in influencers if x.get("id") == inf_id), None)
    if not inf:
        return jsonify({"error": "인플루언서 없음"}), 404

    meetings = inf.setdefault("meetings", [])
    meetings.append({
        "date": date,
        "round": len(meetings) + 1,
        "note": note,
        "created_at": datetime.now().isoformat(timespec="seconds"),
    })
    # 단계 자동 전이: 첫 미팅이면 미팅예약 → 첫 미팅 지나면 미팅완료
    if not inf.get("pipeline_stage"):
        inf["pipeline_stage"] = "미팅예약"
    _save_influencers(influencers)

    # 캘린더 자동 등록
    try:
        events = load_events()
        events.append({
            "id": f"meet_{inf_id}_{int(datetime.now().timestamp())}",
            "title": f"[{inf.get('seller_name') or inf.get('instagram_id')}] {len(meetings)}차 미팅",
            "date": date,
            "type": "meeting",
            "linked_influencer_id": inf_id,
            "note": note,
        })
        save_events(events)
    except Exception as e:
        log.warning(f"미팅 캘린더 자동등록 실패 (무시): {e}")

    return jsonify({"ok": True, "meeting_round": len(meetings)})


# ═══════════════════════════════════════════════════════════
# 📣 Campaigns v2 — 메타 광고관리자 스타일 (캠페인 > 세트 > 공동구매)
# ═══════════════════════════════════════════════════════════

CAMPAIGNS_V2_FILE = DATA_DIR / "campaigns_v2.json"


def _load_campaigns_v2() -> list[dict]:
    if not CAMPAIGNS_V2_FILE.exists():
        return []
    try:
        return json.loads(CAMPAIGNS_V2_FILE.read_text(encoding="utf-8")).get("campaigns", [])
    except Exception:
        return []


def _save_campaigns_v2(items: list[dict]) -> None:
    CAMPAIGNS_V2_FILE.write_text(json.dumps({
        "campaigns": items,
        "schema_version": 2,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }, ensure_ascii=False, indent=2), encoding="utf-8")


def _next_id(items: list, prefix: str) -> str:
    nums = []
    for it in items:
        m = re.match(rf"{prefix}_(\d+)", str(it.get("id", "")))
        if m:
            nums.append(int(m.group(1)))
    return f"{prefix}_{(max(nums) if nums else 0) + 1:04d}"


# ─── 자동 콘텐츠 스케줄 (스토리 가이드) ───
# 사용자가 보낸 스프레드시트 그대로: D-day 기준 거꾸로, 각 날짜에 STORY 1~5 + 피드 1
STORY_CONCEPTS_BY_PHASE = {
    "도입": [   # D-10 ~ D-7
        {"title": "아침 한 포 루틴", "caption": "요즘 아침에 꼭 챙기는 게 생겼는데... 한 번도 빠진 적 없네요 ㅋㅋ 원지 궁금하시죠? 곧 다 말씀드릴게요 👀"},
        {"title": "2달 됐음 고백", "caption": "사실 이거 시작한 게 2달 됐어요. 우연히 정말 좋은 기회가 와서 샘플 받고 꾸준히 써보고 있었거든요... 이제는 말할 수 있을 것 같아요 ㅎㅎ"},
        {"title": "예전 vs 지금 비교", "caption": "예전 사진 보다가 좀 충격 받았는데... 살이 확 찐 건 아닌데 라인이 다 뭉개져 있더라구요 ㅠㅠ 특히 턱선이랑 팔뚝..."},
        {"title": "최근 좋아진 점", "caption": "근데 요즘은 아는 게 뜨는 게 확실히 훨씬 가볍고 편해요! 화장실도 잘 가고 ㅋㅋ 특히 변비... 진짜 너무 오래 고민이었는데"},
        {"title": "공유 결심", "caption": "진짜 진심하게 이거 먹고 나서부터 원가 달라진 것 같아서... 이런 게 처음이라 말하기 많이 망설였는데 이 좋은 거 나만 알기 너무 아까워서 슬슬 공유해 드릴게요! 기대해 주세요 🙏"},
        {"title": "느낌의 변화", "caption": "예전엔 점심 먹고 나면 1시간은 못 일어났는데, 요즘은 진짜 멀쩡해요. 졸음도 적고 머리도 맑은 느낌"},
    ],
    "교감": [   # D-6 ~ D-4
        {"title": "주말 아침 한 포", "caption": "주말 아침도 어김없이 한 포! 일요일에도 빠짐없이 하는 루틴 됐어요 ㅋㅋ 좀 있다가 브런치 먹으러 나가야 하는데..."},
        {"title": "DM 폭발 + 공유 결심", "caption": "어제 올리고 나서 진짜 궁금하다고 DM이랑 연락이 입정 왔다고 ㅋㅋ 감사해요!! 소파 눕자마자 답장 드리는데... 결국 이사님께 하기로 했다!! 카톡 보낸 이거 공유하기로 결심한 거 맞죠?? ㅋㅋ"},
        {"title": "식곤증 사라짐", "caption": "오늘 점심 먹고도 멀쩡한 저... 지금 식곤증 장난 아닐 시간일 텐데 소파 눕자마자 잠 자던 게 진짜인지 요즘은 그런 거 하나도 없어요 ㅋㅋㅋ 이렇게 해결될 줄 몰랐음"},
        {"title": "피부 속건조 고민", "caption": "아... 그래도 화장은 많이 떠서 속건조 때문인가? 싶었거든요 ㅠㅠ 겉만 바른다고 잘 될 게 아니구나 싶었는데 몸 안에서 채워져야 진짜 중요한 거였더라고요..."},
        {"title": "근본 해결 결심", "caption": "내 몸은 내가 챙겨야 한다!! 주변 전문가 지인분들한테 조언 많이 구했어요. 근본적인 해결책을 드릴 수 있을 것 같아서 내일부터 하나하나 더 풀어드릴게요!"},
    ],
    "정보": [   # D-3 ~ D-1
        {"title": "성분 설명", "caption": "오늘은 제가 먹는 이 제품 성분 한번 풀어볼게요. 굳이 제품명 안 까고도 어떻게 작용하는지 설명할 수 있어요 ✨"},
        {"title": "비교/리뷰", "caption": "기존에 먹던 거랑 비교해봤는데, 진짜 차이가 명확해요. 후기 사진 함께 올려요 👀"},
        {"title": "FAQ", "caption": "DM에 가장 많이 들어왔던 질문들 정리했어요. 효능 / 부작용 / 복용법 / 보관법 다 답변!"},
        {"title": "USP 강조", "caption": "이 제품이 다른 거랑 다른 점 딱 3가지만! 그래서 제가 끝까지 못 끊은 이유 ✨"},
    ],
    "임박": [   # D-day
        {"title": "🎉 공구 OPEN", "caption": "드디어 오늘! 공동구매 시작했어요 🎉 링크는 프로필 / 하이라이트 / DM 가능! 이번이 진짜 마지막 기회예요 🔥"},
        {"title": "라이브 공지", "caption": "오후 8시에 인스타 라이브로 직접 보여드릴게요! 질문 미리 받아요"},
        {"title": "주문 인증", "caption": "벌써 주문하신 분들 인증샷 받았어요! 다들 빠르세요... 재고 빨리 빠질 듯 🚨"},
    ],
    "마감": [   # D+1 ~ D+N
        {"title": "후기 인증", "caption": "벌써 받으신 분들 후기 올라오기 시작! 너무 행복한 후기들 ㅠㅠ 감사합니다 💖"},
        {"title": "재구매 안내", "caption": "한 번 드신 분들 재구매 문의 너무 많아서... 마감 전에 꼭 챙기세요 🙏"},
        {"title": "마감 D-1", "caption": "내일 자정 마감입니다! 마지막 D-1 알람 🔔 놓치지 마세요"},
        {"title": "🔴 공구 마감", "caption": "오늘 자정으로 공구 마감됩니다. 마지막 기회예요. 진짜 후회 없으실 거예요 🙏"},
    ],
}

FEED_CONCEPTS = [
    {"title": "라이프스타일 컷", "caption": "(고화질 라이프스타일 컷 + 자연스러운 제품 노출)"},
    {"title": "거울 셀카", "caption": "(피부/몸 변화 인증 + 진솔한 후기)"},
    {"title": "결심 / 에너지", "caption": "(에너지 넘치는 모습 + 변화 강조)"},
    {"title": "비교 콘텐츠", "caption": "(예전 vs 현재 변화 강조 - 인포그래픽 또는 사진)"},
]


def _generate_content_schedule(start_date: str, end_date: str = "", prep_start: str = "", per_day: int = 5) -> list[dict]:
    """콘텐츠 가이드(스케줄링) 자동 생성 — 스토리만 (피드 없음).
    - 사전 기간(prep_start ~ D-11): 주차별로 묶어 일주일에 per_day개 ("N월 M주차").
    - 집중 기간(D-10 ~ 마켓 D-day ~ 마감 D+N): 하루 per_day개 무조건.
    """
    try:
        sd = datetime.strptime(start_date[:10], "%Y-%m-%d")
    except (ValueError, TypeError):
        return []
    try:
        ed = datetime.strptime(end_date[:10], "%Y-%m-%d") if end_date else sd
    except (ValueError, TypeError):
        ed = sd
    weekday_kr = ["월", "화", "수", "목", "금", "토", "일"]
    intensive_start = sd - timedelta(days=10)  # D-10
    try:
        ps = datetime.strptime(prep_start[:10], "%Y-%m-%d") if prep_start else intensive_start
    except (ValueError, TypeError):
        ps = intensive_start

    def story_slots(phase: str, seed: int) -> list[dict]:
        pool = STORY_CONCEPTS_BY_PHASE.get(phase, [])
        out = []
        for i in range(max(1, per_day)):
            cp = pool[(seed + i) % len(pool)] if pool else {"title": "", "caption": ""}
            out.append({
                "type": "story", "title": f"STORY {i + 1}",
                "concept": cp["title"], "caption": cp["caption"],
                "image_url": "", "posted": False, "posted_at": "",
            })
        return out

    days = []

    # 1) 사전 기간 — 주차별 (prep_start ~ D-11): 주당 per_day개
    if ps < intensive_start:
        cursor = ps
        wk = 0
        while cursor < intensive_start:
            wom = ((cursor.day - 1) // 7) + 1
            days.append({
                "date": cursor.strftime("%Y-%m-%d"),
                "weekday": "",
                "d_label": f"{cursor.month}월 {wom}주차",
                "phase": "사전",
                "weekly": True,
                "slots": story_slots("도입", wk),
            })
            cursor += timedelta(days=7)
            wk += 1

    # 2) 집중 기간 — 일별 (D-10 ~ D-day ~ 마감 D+N): 하루 per_day개
    post_days = (ed - sd).days
    for offset in range(-10, max(post_days, 0) + 1):
        d = sd + timedelta(days=offset)
        if offset < 0:
            d_label = f"D-{-offset}"
            phase = "정보" if offset >= -3 else ("교감" if offset >= -6 else "도입")
        elif offset == 0:
            d_label = "D-day"; phase = "임박"
        else:
            d_label = f"D+{offset}"; phase = "마감"
        days.append({
            "date": d.strftime("%Y-%m-%d"),
            "weekday": weekday_kr[d.weekday()],
            "d_label": d_label,
            "phase": phase,
            "slots": story_slots(phase, offset),
        })

    return days


@app.route("/api/campaigns_v2", methods=["GET", "POST"])
def api_campaigns_v2_list():
    items = _load_campaigns_v2()
    if request.method == "POST":
        payload = request.get_json(force=True, silent=True) or {}
        # linked_handle 로 인플루언서 매칭
        linked_inf_id = payload.get("linked_influencer_id")
        linked_handle = (payload.get("linked_handle") or "").strip().lstrip("@")
        linked_inf_handle = None
        if linked_handle:
            influencers = _load_influencers()
            inf = next((x for x in influencers
                        if (x.get("instagram_id") or "").lower() == linked_handle.lower()), None)
            if inf:
                linked_inf_id = inf["id"]
                linked_inf_handle = inf.get("instagram_id")

        start_date = payload.get("start_date") or ""
        end_date = payload.get("end_date") or ""
        # 콘텐츠는 자동 생성 X — 캠페인 디테일에서 [생성] 버튼으로 제품정보+소구점 입력 후 생성
        content_days = []

        cam = {
            "id": _next_id(items, "cam"),
            "seller_name": (payload.get("seller_name") or "").strip(),
            "brand": (payload.get("brand") or "").strip(),
            "product": (payload.get("product") or "").strip(),
            "product_id": (payload.get("product_id") or "").strip(),
            "type": payload.get("type") or "마이크로",
            "market_schedule": start_date,
            "instagram_url": payload.get("instagram_url") or "",
            "seller_traits": payload.get("seller_traits") or "",
            "linked_influencer_id": linked_inf_id,
            "linked_influencer_handle": linked_inf_handle or linked_handle or None,
            "status": payload.get("status") or "준비중",
            "notes": payload.get("notes") or "",
            "product_shipping": {  # 셀러 단위 (캠페인 레벨로 이동)
                "sent_date": None,
                "tracking_no": "",
                "carrier": "",
                "note": "",
            },
            "sets": [{
                "id": "set_0001",
                "round": 1,
                "label": "1차",
                "memo": "",
                "features": payload.get("features") or {"schedule": True, "events": True, "drive": True, "banners": True, "reels": True},
                "ads": [{
                    "id": "ad_0001",
                    "name": "공동구매 1차",
                    "scheduling": {
                        "start_date": start_date or None,
                        "end_date": end_date or None,
                        "items": [],
                    },
                    "content_days": content_days,
                    "events": [],
                    "drive_links": [],
                    "banners": {
                        "openfeed": {"checked": False, "draft_url": "", "final_url": "", "note": ""},
                        "price": {"checked": False, "draft_url": "", "final_url": "", "note": ""},
                        "event": {"checked": False, "draft_url": "", "final_url": "", "note": ""},
                    },
                    "reels": [],
                    "revenue": 0,
                    "cost": 0,
                    "status": "준비중",
                    "created_at": datetime.now().isoformat(timespec="seconds"),
                }],
                "created_at": datetime.now().isoformat(timespec="seconds"),
            }],
            "created_at": datetime.now().isoformat(timespec="seconds"),
        }
        items.append(cam)
        _save_campaigns_v2(items)
        return jsonify({"ok": True, "campaign": cam})
    return jsonify({"campaigns": items, "total": len(items)})


@app.route("/api/campaigns_v2/<cam_id>", methods=["GET", "PATCH", "DELETE"])
def api_campaigns_v2_one(cam_id):
    items = _load_campaigns_v2()
    cam = next((c for c in items if c["id"] == cam_id), None)
    if not cam:
        return jsonify({"error": "캠페인 없음"}), 404
    if request.method == "GET":
        return jsonify(cam)
    if request.method == "DELETE":
        items = [c for c in items if c["id"] != cam_id]
        _save_campaigns_v2(items)
        return jsonify({"ok": True})
    payload = request.get_json(force=True, silent=True) or {}
    for k in ["seller_name", "brand", "product", "type", "market_schedule", "status",
              "linked_influencer_id", "instagram_url", "seller_traits", "notes",
              "settlement_done"]:
        if k in payload:
            cam[k] = payload[k]
    # 제품 발송 (캠페인 레벨)
    if "product_shipping" in payload and isinstance(payload["product_shipping"], dict):
        cam.setdefault("product_shipping", {}).update(payload["product_shipping"])
    # 건수 정산 (제품 티어별 판매 건수)
    if "tier_counts" in payload and isinstance(payload["tier_counts"], dict):
        cam["tier_counts"] = payload["tier_counts"]
    # 제품 연결 변경
    if "product_id" in payload:
        cam["product_id"] = payload["product_id"]
    if "product" in payload:
        cam["product"] = payload["product"]
    _save_campaigns_v2(items)

    # 마켓 일정 → 캘린더 sync
    try:
        events = load_events()
        tag = f"camcal_{cam['id']}_market"
        events = [e for e in events if e.get("id") != tag]
        mk = cam.get("market_schedule")
        if mk and len(mk) >= 8:
            # YYYY-MM-DD 형식이면 직접 박고, 아니면 텍스트 그대로 type=campaign_market
            date_match = re.match(r"^(\d{4}-\d{2}-\d{2})", mk)
            if date_match:
                events.append({
                    "id": tag,
                    "title": f"[{cam.get('seller_name') or ''}] 마켓 시작",
                    "date": date_match.group(1),
                    "type": "campaign_market",
                    "linked_campaign_id": cam["id"],
                    "note": mk,
                })
                save_events(events)
    except Exception as e:
        log.warning(f"마켓 일정 캘린더 sync 실패 (무시): {e}")

    return jsonify({"ok": True, "campaign": cam})


@app.route("/api/campaigns_v2/<cam_id>/settlement", methods=["PATCH"])
def api_campaigns_v2_settlement(cam_id):
    """캠페인 정산 — settings(사업자/오픈타입/RS/수수료율/PG반반/보이는열) + rows(일자별).
    body: {settings:{...}, rows:[...]}"""
    items = _load_campaigns_v2()
    cam = next((c for c in items if c["id"] == cam_id), None)
    if not cam:
        return jsonify({"error": "캠페인 없음"}), 404
    payload = request.get_json(force=True, silent=True) or {}
    stl = cam.setdefault("settlement", {"settings": {}, "rows": []})
    if "settings" in payload and isinstance(payload["settings"], dict):
        stl.setdefault("settings", {}).update(payload["settings"])
    if "rows" in payload and isinstance(payload["rows"], list):
        stl["rows"] = payload["rows"]
        # 캠페인 revenue/cost roll-up (정산 있으면 우선)
        cam["settlement_revenue"] = sum(int(r.get("revenue") or 0) for r in payload["rows"])
        cam["settlement_contribution"] = sum(_row_contribution(r) for r in payload["rows"])
    _save_campaigns_v2(items)
    return jsonify({"ok": True, "settlement": stl,
                    "revenue": cam.get("settlement_revenue", 0),
                    "contribution": cam.get("settlement_contribution", 0)})


def _row_contribution(r: dict) -> int:
    rev = int(r.get("revenue") or 0)
    cost = sum(int(r.get(k) or 0) for k in ("cogs", "shipping", "event_cost", "seller_fee", "pg_fee"))
    return rev - cost


@app.route("/api/campaigns_v2/totals", methods=["GET"])
def api_campaigns_v2_totals():
    """전체 캠페인 매출/공헌 합."""
    items = _load_campaigns_v2()
    rows = []
    total_rev = 0
    total_contrib = 0
    for cam in items:
        stl = cam.get("settlement") or {}
        srows = stl.get("rows") or []
        if srows:
            rev = sum(int(r.get("revenue") or 0) for r in srows)
            contrib = sum(_row_contribution(r) for r in srows)
        else:
            # 정산 없으면 마켓 sales 합산
            rev = 0
            cost = 0
            for st in cam.get("sets", []):
                for ad in st.get("ads", []):
                    rev += int(ad.get("revenue") or 0)
                    cost += int(ad.get("cost") or 0)
            contrib = rev - cost
        total_rev += rev
        total_contrib += contrib
        rows.append({
            "id": cam["id"], "seller_name": cam.get("seller_name"),
            "brand": cam.get("brand"), "type": cam.get("type"),
            "status": cam.get("status"),
            "revenue": rev, "contribution": contrib,
            "contribution_pct": round(contrib / rev * 100, 1) if rev else None,
            "has_settlement": bool(srows),
        })
    rows.sort(key=lambda x: -(x["revenue"] or 0))
    return jsonify({
        "campaigns": rows,
        "total_revenue": total_rev,
        "total_contribution": total_contrib,
        "total_contribution_pct": round(total_contrib / total_rev * 100, 1) if total_rev else None,
        "count": len(rows),
    })


@app.route("/api/campaigns_v2/<cam_id>/sets", methods=["POST"])
def api_campaigns_v2_add_set(cam_id):
    """세트 = 공구 차수. {round: 1} 박으면 자동 ad 1개 생성."""
    items = _load_campaigns_v2()
    cam = next((c for c in items if c["id"] == cam_id), None)
    if not cam:
        return jsonify({"error": "캠페인 없음"}), 404
    payload = request.get_json(force=True, silent=True) or {}
    sets = cam.setdefault("sets", [])
    round_num = payload.get("round") or len(sets) + 1
    # 기능 토글 (기본 전부 ON)
    feats = payload.get("features")
    if not isinstance(feats, dict):
        feats = {"schedule": True, "events": True, "drive": True, "banners": True, "reels": True}
    new_set = {
        "id": _next_id(sets, "set"),
        "round": round_num,
        "label": payload.get("label") or f"{round_num}차",
        "memo": "",
        "features": feats,
        "ads": [{
            "id": "ad_0001",
            "name": payload.get("ad_name") or f"공동구매 {round_num}차",
            "product_sent_date": None,
            "scheduling": {"start_date": None, "end_date": None, "items": []},
            "content_days": [],
            "events": [],
            "drive_links": [],
            "banners": {
                "openfeed": {"checked": False, "draft_url": "", "final_url": "", "note": ""},
                "price": {"checked": False, "draft_url": "", "final_url": "", "note": ""},
                "event": {"checked": False, "draft_url": "", "final_url": "", "note": ""},
            },
            "reels": [],
            "sales": [],
            "revenue": 0,
            "cost": 0,
            "status": "준비중",
            "created_at": datetime.now().isoformat(timespec="seconds"),
        }],
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    sets.append(new_set)
    _save_campaigns_v2(items)
    return jsonify({"ok": True, "set": new_set})


def _sync_ad_to_calendar(cam: dict, st: dict, ad: dict) -> int:
    """광고의 scheduling.items + product_sent_date를 events.json에 sync. 기존 ad_id 매칭 이벤트 제거 후 재생성."""
    try:
        events = load_events()
    except Exception:
        return 0
    tag = f"adcal_{cam['id']}_{st['id']}_{ad['id']}_"
    # 기존 자동 등록 제거
    events = [e for e in events if not str(e.get("id", "")).startswith(tag)]
    added = 0
    seller = cam.get("seller_name") or "셀러"
    brand = cam.get("brand") or ""
    round_label = st.get("label") or f"{st.get('round', 1)}차"

    # 스케줄링 시작/종료를 라이브 기간으로
    sched = ad.get("scheduling") or {}
    if sched.get("start_date") and sched.get("end_date"):
        events.append({
            "id": f"{tag}live",
            "title": f"[{seller}] {round_label} 라이브",
            "date": sched["start_date"],
            "end_date": sched["end_date"],
            "type": "campaign_live",
            "linked_campaign_id": cam["id"],
            "linked_set_id": st["id"],
            "linked_ad_id": ad["id"],
            "brand": brand,
        })
        added += 1

    # 세부 일정 각각
    for i, item in enumerate(sched.get("items") or []):
        if not item.get("date"):
            continue
        events.append({
            "id": f"{tag}sched_{i}",
            "title": f"[{seller}] {item.get('label') or '일정'}",
            "date": item["date"],
            "type": "campaign_schedule",
            "linked_campaign_id": cam["id"],
            "linked_set_id": st["id"],
            "linked_ad_id": ad["id"],
        })
        added += 1

    # 제품 발송일
    if ad.get("product_sent_date"):
        events.append({
            "id": f"{tag}ship",
            "title": f"[{seller}] 제품 발송 ({round_label})",
            "date": ad["product_sent_date"],
            "type": "campaign_ship",
            "linked_campaign_id": cam["id"],
            "linked_set_id": st["id"],
            "linked_ad_id": ad["id"],
        })
        added += 1

    save_events(events)
    return added


@app.route("/api/campaigns_v2/<cam_id>/sets/<set_id>", methods=["PATCH"])
def api_campaigns_v2_patch_set(cam_id, set_id):
    """세트 단위 필드 수정 (label / memo)."""
    items = _load_campaigns_v2()
    cam = next((c for c in items if c["id"] == cam_id), None)
    if not cam:
        return jsonify({"error": "캠페인 없음"}), 404
    st = next((s for s in cam.get("sets", []) if s["id"] == set_id), None)
    if not st:
        return jsonify({"error": "세트 없음"}), 404
    payload = request.get_json(force=True, silent=True) or {}
    for k in ["label", "memo", "last_ship_date"]:
        if k in payload:
            st[k] = payload[k]
    if "features" in payload and isinstance(payload["features"], dict):
        st.setdefault("features", {}).update(payload["features"])
    _save_campaigns_v2(items)
    return jsonify({"ok": True, "set": st})


@app.route("/api/campaigns_v2/<cam_id>/sets/<set_id>/ads/<ad_id>", methods=["PATCH"])
def api_campaigns_v2_patch_ad(cam_id, set_id, ad_id):
    """공동구매(광고) 단위 필드 수정 — 제품발송/스케줄링/이벤트/드라이브/배너/릴스 전부.
    수정 후 캘린더 자동 sync."""
    items = _load_campaigns_v2()
    cam = next((c for c in items if c["id"] == cam_id), None)
    if not cam:
        return jsonify({"error": "캠페인 없음"}), 404
    st = next((s for s in cam.get("sets", []) if s["id"] == set_id), None)
    if not st:
        return jsonify({"error": "세트 없음"}), 404
    ad = next((a for a in st.get("ads", []) if a["id"] == ad_id), None)
    if not ad:
        return jsonify({"error": "광고 없음"}), 404

    payload = request.get_json(force=True, silent=True) or {}
    for k in ["name", "product_sent_date", "status", "revenue", "cost", "expected_revenue"]:
        if k in payload:
            ad[k] = payload[k]
    for k in ["scheduling", "banners", "banner_cats"]:
        if k in payload and isinstance(payload[k], dict):
            ad[k] = payload[k] if k == "banner_cats" else {**ad.get(k, {}), **payload[k]}
    for k in ["events", "drive_links", "reels", "content_days", "sales", "banner_images", "event_costs", "other_costs"]:
        if k in payload and isinstance(payload[k], list):
            ad[k] = payload[k]
    # 날짜별 매출 입력 시 → revenue/cost 자동 합산
    if "sales" in payload and isinstance(payload["sales"], list):
        ad["revenue"] = sum(int(r.get("revenue") or 0) for r in payload["sales"])
        ad["cost"] = sum(int(r.get("cost") or 0) for r in payload["sales"])
    # 스케줄링 시작일 바뀌면 콘텐츠 슬롯 자동 재생성 (옵션)
    if payload.get("regenerate_content_schedule"):
        sd = (ad.get("scheduling") or {}).get("start_date")
        ed = (ad.get("scheduling") or {}).get("end_date")
        if sd:
            ad["content_days"] = _generate_content_schedule(sd, ed or "", payload.get("prep_start") or "", int(payload.get("per_day") or 5))
    _save_campaigns_v2(items)

    # 캘린더 자동 sync (스케줄링/제품발송 변경 시)
    try:
        _sync_ad_to_calendar(cam, st, ad)
    except Exception as e:
        log.warning(f"캘린더 sync 실패 (무시): {e}")

    return jsonify({"ok": True, "ad": ad})


# ─── 콘텐츠 스케줄 생성 (입력 기반 + 셀러 톤 학습) ───
@app.route("/api/campaigns_v2/<cam_id>/sets/<set_id>/ads/<ad_id>/generate", methods=["POST"])
def api_campaigns_v2_generate_content(cam_id, set_id, ad_id):
    """제품 정보 + 소구점 + 길이 입력 → Gemini 셀러 톤 매칭 생성.
    body: {product:{name,usp,detail,price,avoid}, selling_points:[..], length:"short|medium|long",
           reference_handles:[..], attach_images:bool}"""
    items = _load_campaigns_v2()
    cam = next((c for c in items if c["id"] == cam_id), None)
    if not cam:
        return jsonify({"error": "캠페인 없음"}), 404
    st = next((s for s in cam.get("sets", []) if s["id"] == set_id), None)
    ad = next((a for a in (st or {}).get("ads", []) if a["id"] == ad_id), None) if st else None
    if not ad:
        return jsonify({"error": "세트/광고 없음"}), 404

    payload = request.get_json(force=True, silent=True) or {}
    sched = ad.get("scheduling") or {}
    start_date = payload.get("start_date") or sched.get("start_date") or cam.get("market_schedule") or ""
    if not start_date:
        return jsonify({"error": "시작일이 없음 — 캠페인/광고 시작일 먼저 박아"}), 400
    end_date = payload.get("end_date") or sched.get("end_date") or ""

    try:
        from content_gen import generate_content_schedule  # type: ignore
    except ImportError as e:
        return jsonify({"error": f"content_gen 모듈 실패: {e}"}), 500

    handle = cam.get("linked_influencer_handle") or ""
    if not handle and cam.get("instagram_url"):
        m = re.search(r"instagram\.com/([^/?\s]+)", cam["instagram_url"])
        handle = m.group(1) if m else cam["instagram_url"].lstrip("@")

    result = generate_content_schedule(
        seller_handle=handle,
        product=payload.get("product") or {
            "name": cam.get("product") or "",
            "usp": "", "detail": "", "price": "", "avoid": "",
        },
        selling_points=payload.get("selling_points") or [],
        length=payload.get("length") or "medium",
        start_date=start_date,
        end_date=end_date,
        config=load_config() or {},
        reference_handles=payload.get("reference_handles") or [],
        attach_images=payload.get("attach_images", True),
    )

    # 결정형 스케줄 뼈대(주차별 사전 + 일별 D-10~마감, 피드 없음)에 Gemini 캡션을 순서대로 채움
    prep_start = payload.get("prep_start") or ""
    skeleton = _generate_content_schedule(start_date, end_date, prep_start, 5)
    gen_slots = [sl for d in result["content_days"] for sl in (d.get("slots") or []) if sl.get("type") != "feed"]
    gi = 0
    for d in skeleton:
        for sl in d["slots"]:
            if gi < len(gen_slots):
                g = gen_slots[gi]
                if g.get("concept"):
                    sl["concept"] = g["concept"]
                if g.get("caption"):
                    sl["caption"] = g["caption"]
                if g.get("image_url"):
                    sl["image_url"] = g["image_url"]
                gi += 1
    ad["content_days"] = skeleton
    ad.setdefault("scheduling", {})["start_date"] = start_date
    if end_date:
        ad["scheduling"]["end_date"] = end_date
    # 생성 메타 저장
    ad["content_gen_meta"] = {
        "product": payload.get("product") or {},
        "selling_points": payload.get("selling_points") or [],
        "length": payload.get("length") or "medium",
        "gemini_used": result["gemini_used"],
        "tone_samples_count": result["tone_samples_count"],
        "images_attached": result["images_attached"],
        "generated_at": result["generated_at"],
    }
    _save_campaigns_v2(items)
    return jsonify({"ok": True, **{k: result[k] for k in ("gemini_used", "tone_samples_count", "images_attached")},
                    "days": len(result["content_days"])})


# ─── 아카이브 이미지 서빙 ───
@app.route("/archive-img/<seller_folder>/<path:img_path>", methods=["GET"])
def api_archive_img(seller_folder, img_path):
    base = DATA_DIR / "local_archive" / seller_folder
    full = (base / img_path).resolve()
    # 경로 탈출 방지
    if not str(full).startswith(str(base.resolve())):
        return jsonify({"error": "잘못된 경로"}), 400
    if not full.exists():
        return jsonify({"error": "이미지 없음"}), 404
    return send_from_directory(full.parent, full.name)


# ─── 아카이브 이미지 풀 (피커용) ───
@app.route("/api/archive/images", methods=["GET"])
def api_archive_images():
    """이미지 피커용 — 모든 셀러 아카이브 이미지 목록 (선택적 핸들 필터)."""
    handle_filter = request.args.get("handle", "").strip().lower()
    sellers = load_sellers()
    out = []
    for s in sellers:
        if handle_filter and (s.get("instagram") or "").lower() != handle_filter:
            continue
        folder = f"{s['id']}.{s['name']}_@{s['instagram']}"
        mp = DATA_DIR / "local_archive" / folder / "_manifest.json"
        if not mp.exists():
            continue
        try:
            data = json.loads(mp.read_text(encoding="utf-8"))
        except Exception:
            continue
        for it in data.get("items", []):
            if it.get("media") != "image" or not it.get("file_path"):
                continue
            out.append({
                "url": f"/archive-img/{folder}/{it['file_path']}",
                "source": f"@{s['instagram']}",
                "seller_name": s["name"],
                "highlight": it.get("highlight_label", ""),
                "alt": (it.get("alt_text") or "")[:80],
                "fit": it.get("story_slot_fit", {}),
            })
    return jsonify({"images": out, "total": len(out)})


# ─── 셀러 공개 뷰 (캠페인 단위 큐레이션) ─────────────────────
#  🔒 보안 원칙: 셀러에게는 _seller_safe_view 화이트리스트 필드만 노출.
#  원가/공헌이익/수수료(seller_fee·pg_fee·vat)/event_costs/배너 레퍼런스·내부메모는
#  절대 페이로드에 담지 않는다. (서버 사이드 차단 = 단일 보안 경계)
BANNER_CAT_LABELS = {"open": "오픈 배너", "price": "가격구성 배너", "event": "이벤트 배너"}


def _seller_token(cam_id, set_id, ad_id) -> str:
    import base64 as _b64
    raw = f"{cam_id}|{set_id}|{ad_id}".encode("utf-8")
    return _b64.urlsafe_b64encode(raw).decode("ascii").rstrip("=")


def _decode_seller_token(token: str):
    import base64 as _b64
    pad = "=" * (-len(token) % 4)
    try:
        raw = _b64.urlsafe_b64decode(token + pad).decode("utf-8")
        parts = raw.split("|")
        if len(parts) == 3:
            return parts[0], parts[1], parts[2]
    except Exception:
        pass
    return None, None, None


def _ensure_access_code(cam: dict, items: list[dict]) -> str:
    """캠페인에 셀러 접속코드 부여(없으면 생성·충돌회피). 호출 측에서 save 필요."""
    code = cam.get("seller_access_code")
    if code:
        return code
    existing = {c.get("seller_access_code") for c in items if c.get("seller_access_code")}
    code = secrets.token_urlsafe(6).replace("-", "x").replace("_", "y")[:9]
    while code in existing:
        code = secrets.token_urlsafe(6).replace("-", "x").replace("_", "y")[:9]
    cam["seller_access_code"] = code
    return code


def _seller_safe_view(cam: dict) -> dict:
    """셀러 노출용 큐레이션 페이로드 — 화이트리스트 필드만. (보안 경계)
    캠페인 단위로 전 차수(1·2·3차) 스토리/최종배너/이벤트를 누적해서 묶고,
    출고일·일자별 매출(매출액만)·정산 완료 여부를 노출한다."""
    rounds = []
    for st in sorted(cam.get("sets", []), key=lambda s: (s.get("round") or 0)):
        feats = st.get("features") or {}
        ads = st.get("ads") or []
        ad = ads[0] if ads else {}
        sch = ad.get("scheduling") or {}
        # 콘텐츠 가이드 — 스토리만(피드 제외). raw 인덱스 보존(_di/_si) → 체크 PATCH 매핑용.
        days = []
        for di, d in enumerate(ad.get("content_days") or []):
            slots = []
            for si, sl in enumerate(d.get("slots") or []):
                if sl.get("type") == "feed":
                    continue
                slots.append({
                    "title": sl.get("title", ""),
                    "concept": sl.get("concept", ""),
                    "caption": sl.get("caption", ""),
                    "image_url": sl.get("image_url", ""),
                    "posted": bool(sl.get("posted")),
                    "live_url": sl.get("live_url", ""),
                    "_si": si,
                })
            if not slots:
                continue
            days.append({
                "date": d.get("date", ""),
                "weekday": d.get("weekday", ""),
                "d_label": d.get("d_label", ""),
                "phase": d.get("phase", ""),
                "weekly": bool(d.get("weekly")),
                "_di": di,
                "slots": slots,
            })
        # 최종 배너만 (레퍼런스/내부메모 제외). 표시는 썸네일, 다운로드는 드라이브 원본.
        banners = []
        cats = ad.get("banner_cats") or {}
        for key, label in BANNER_CAT_LABELS.items():
            cv = cats.get(key) or {}
            for im in (cv.get("finals") or []):
                if not isinstance(im, dict):
                    continue
                orig = (f"/api/file/{im['file_id']}" if im.get("file_id") else "") or im.get("url") or im.get("data")
                disp = im.get("thumb") or im.get("url") or im.get("data") or orig
                if disp:
                    banners.append({"label": label, "url": disp, "download": orig or disp, "name": im.get("name") or ""})
        # 이벤트 — 이름/날짜/설명만 (비용 event_costs 제외)
        events = []
        for ev in (ad.get("events") or []):
            name = ev.get("label") or ev.get("name") or ev.get("title") or ""
            if not name and not ev.get("date"):
                continue
            events.append({
                "name": name,
                "date": ev.get("date") or "",
                "desc": ev.get("desc") or ev.get("note") or ev.get("memo") or "",
            })
        rounds.append({
            "set_id": st.get("id"),
            "round": st.get("round"),
            "label": st.get("label") or (f"{st.get('round')}차" if st.get("round") else "공구"),
            "start_date": sch.get("start_date") or "",
            "end_date": sch.get("end_date") or "",
            "show_schedule": bool(feats.get("schedule", True)),
            "days": days,
            "banners": banners,
            "events": events,
        })
    # 일자별 매출 — 매출(revenue)만. 원가/공헌이익/수수료 전부 제외.
    daily_sales = []
    for r in ((cam.get("settlement") or {}).get("rows") or []):
        rev = int(r.get("revenue") or 0)
        if not r.get("date") and not rev:
            continue
        daily_sales.append({"date": r.get("date") or "", "revenue": rev})
    daily_sales.sort(key=lambda x: x["date"])
    total_rev = sum(x["revenue"] for x in daily_sales)
    ship = cam.get("product_shipping") or {}
    return {
        "seller_name": cam.get("seller_name", ""),
        "brand": cam.get("brand", ""),
        "product": cam.get("product", ""),
        "ship_date": ship.get("sent_date") or "",
        "ship_carrier": ship.get("carrier") or "",
        "ship_tracking": ship.get("tracking_no") or "",
        "rounds": rounds,
        "daily_sales": daily_sales,
        "total_revenue": total_rev,
        "settlement_done": bool(cam.get("settlement_done")),
        "has_sales": bool(daily_sales),
    }


@app.route("/api/campaigns_v2/<cam_id>/share", methods=["GET"])
def api_campaigns_v2_campaign_share(cam_id):
    """캠페인 단위 셀러 공유 — 셀러용 링크 + 관리자 미리보기(트래킹 제외) 링크."""
    items = _load_campaigns_v2()
    cam = next((c for c in items if c["id"] == cam_id), None)
    if not cam:
        return jsonify({"error": "캠페인 없음"}), 404
    code = _ensure_access_code(cam, items)
    _save_campaigns_v2(items)
    return jsonify({
        "code": code,
        "seller_path": f"/seller/{code}",
        "preview_path": f"/seller/{code}?preview=1",
    })


@app.route("/api/campaigns_v2/<cam_id>/sets/<set_id>/ads/<ad_id>/share", methods=["GET"])
def api_campaigns_v2_share_link(cam_id, set_id, ad_id):
    """(레거시 차수단위 호출) → 캠페인 단위 접속코드로 통합."""
    items = _load_campaigns_v2()
    cam = next((c for c in items if c["id"] == cam_id), None)
    if not cam:
        return jsonify({"error": "캠페인 없음"}), 404
    code = _ensure_access_code(cam, items)
    _save_campaigns_v2(items)
    return jsonify({"token": code, "path": f"/seller/{code}",
                    "preview_path": f"/seller/{code}?preview=1"})


@app.route("/seller/<token>", methods=["GET"])
def seller_view(token):
    """셀러용 캠페인 단위 큐레이션 뷰.
    token = 접속코드(우선) | 레거시 base64 차수토큰(→캠페인 코드로 redirect)."""
    items = _load_campaigns_v2()
    cam = next((c for c in items if c.get("seller_access_code") == token), None)
    if cam:
        return render_template(
            "seller_view.html",
            code=token,
            view=_seller_safe_view(cam),
            is_preview=(request.args.get("preview") == "1"),
        )
    # 레거시 base64 토큰 → 해당 캠페인 코드로 redirect (기존 공유 링크 호환)
    cam_id, _set_id, _ad_id = _decode_seller_token(token)
    if cam_id:
        legacy = next((c for c in items if c["id"] == cam_id), None)
        if legacy:
            code = _ensure_access_code(legacy, items)
            _save_campaigns_v2(items)
            q = "?preview=1" if request.args.get("preview") == "1" else ""
            return redirect(f"/seller/{code}{q}")
    return "잘못된 링크입니다", 404


SELLER_TRACK_FILE = DATA_DIR / "seller_tracking.json"


def _load_seller_track() -> dict:
    if not SELLER_TRACK_FILE.exists():
        return {}
    try:
        return json.loads(SELLER_TRACK_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_seller_track(data: dict) -> None:
    SELLER_TRACK_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _code_track_key(code: str) -> str:
    return f"code:{code}"


def _record_track_event(key: str, p: dict) -> None:
    """event: open(세션 시작) | ping(heartbeat) | close. body: {event, session_id, seconds?}"""
    event = p.get("event")
    sid = p.get("session_id") or "anon"
    now = datetime.now().isoformat(timespec="seconds")
    data = _load_seller_track()
    rec = data.setdefault(key, {"sessions": [], "total_seconds": 0, "visit_count": 0, "first_at": now, "last_at": now})
    rec["last_at"] = now
    if event == "open":
        rec["visit_count"] = (rec.get("visit_count") or 0) + 1
        rec["sessions"].insert(0, {"session_id": sid, "started_at": now, "seconds": 0, "last_ping": now})
        rec["sessions"] = rec["sessions"][:100]  # 최근 100세션
    elif event in ("ping", "close"):
        sess = next((s for s in rec["sessions"] if s.get("session_id") == sid), None)
        if sess:
            add = max(0, min(int(p.get("seconds") or 0), 90))  # heartbeat 클램프
            sess["seconds"] = (sess.get("seconds") or 0) + add
            sess["last_ping"] = now
            rec["total_seconds"] = (rec.get("total_seconds") or 0) + add
    _save_seller_track(data)


@app.route("/api/sv/<code>/track", methods=["POST"])
def api_sv_track(code):
    """셀러 접속 트래킹(코드 기반). 관리자 미리보기(preview)는 기록 제외."""
    p = request.get_json(silent=True) or {}
    if p.get("preview"):
        return jsonify({"ok": True, "skipped": "preview"})
    _record_track_event(_code_track_key(code), p)
    return jsonify({"ok": True})


@app.route("/api/sv/<code>/slot", methods=["PATCH"])
def api_sv_slot(code):
    """셀러 체크/링크 (posted, live_url 만 허용). set_id + raw day/slot 인덱스."""
    items = _load_campaigns_v2()
    cam = next((c for c in items if c.get("seller_access_code") == code), None)
    if not cam:
        return jsonify({"error": "잘못된 코드"}), 404
    p = request.get_json(force=True, silent=True) or {}
    if p.get("preview"):
        return jsonify({"ok": True, "skipped": "preview"})  # 미리보기는 셀러 데이터 변경 X
    field = p.get("field")
    if field not in ("posted", "live_url"):
        return jsonify({"error": "허용되지 않는 필드"}), 400
    st = next((s for s in cam.get("sets", []) if s["id"] == p.get("set_id")), None)
    ads = (st or {}).get("ads") or []
    ad = ads[0] if ads else None
    if not ad:
        return jsonify({"error": "세트 없음"}), 404
    di, si = p.get("day_index"), p.get("slot_index")
    days = ad.get("content_days") or []
    if di is None or si is None or di >= len(days) or si >= len(days[di].get("slots", [])):
        return jsonify({"error": "인덱스 초과"}), 400
    days[di]["slots"][si][field] = p.get("value")
    if field == "posted" and p.get("value"):
        days[di]["slots"][si]["posted_at"] = datetime.now().isoformat(timespec="seconds")
    _save_campaigns_v2(items)
    return jsonify({"ok": True})


@app.route("/api/seller/<token>/track", methods=["POST"])
def api_seller_track(token):
    """(레거시 base64 토큰 호환) 트래킹 — 캠페인 코드로 환산해 통합 기록."""
    p = request.get_json(silent=True) or {}
    if p.get("preview"):
        return jsonify({"ok": True, "skipped": "preview"})
    key = _code_track_key(token)
    cam_id, _s, _a = _decode_seller_token(token)
    if cam_id:
        items = _load_campaigns_v2()
        legacy = next((c for c in items if c["id"] == cam_id), None)
        if legacy and legacy.get("seller_access_code"):
            key = _code_track_key(legacy["seller_access_code"])
    _record_track_event(key, p)
    return jsonify({"ok": True})


@app.route("/api/campaigns_v2/<cam_id>/sets/<set_id>/ads/<ad_id>/tracking", methods=["GET"])
def api_seller_tracking_get(cam_id, set_id, ad_id):
    """내부용 — 캠페인 셀러 접속 현황(코드 기반 · 캠페인 단위 합산)."""
    items = _load_campaigns_v2()
    cam = next((c for c in items if c["id"] == cam_id), None)
    code = (cam or {}).get("seller_access_code")
    data = _load_seller_track()
    rec = (data.get(_code_track_key(code)) if code else None) or {"sessions": [], "total_seconds": 0, "visit_count": 0}
    return jsonify(rec)


# ─── 콘텐츠 슬롯 단건 수정 (셀러뷰/내부 공용) ───
@app.route("/api/campaigns_v2/<cam_id>/sets/<set_id>/ads/<ad_id>/slot", methods=["PATCH"])
def api_campaigns_v2_patch_slot(cam_id, set_id, ad_id):
    """body: {day_index, slot_index, field, value}"""
    items = _load_campaigns_v2()
    cam = next((c for c in items if c["id"] == cam_id), None)
    st = next((s for s in (cam or {}).get("sets", []) if s["id"] == set_id), None) if cam else None
    ad = next((a for a in (st or {}).get("ads", []) if a["id"] == ad_id), None) if st else None
    if not ad:
        return jsonify({"error": "광고 없음"}), 404
    p = request.get_json(force=True, silent=True) or {}
    di, si, field = p.get("day_index"), p.get("slot_index"), p.get("field")
    days = ad.get("content_days") or []
    if di is None or si is None or di >= len(days) or si >= len(days[di].get("slots", [])):
        return jsonify({"error": "인덱스 범위 초과"}), 400
    days[di]["slots"][si][field] = p.get("value")
    if field == "posted" and p.get("value"):
        days[di]["slots"][si]["posted_at"] = datetime.now().isoformat(timespec="seconds")
    _save_campaigns_v2(items)
    return jsonify({"ok": True})


@app.route("/api/campaigns_v2/from_influencer/<inf_id>", methods=["POST"])
def api_campaigns_v2_from_influencer(inf_id):
    """진행 예정 셀러에서 [캠페인 추가] 클릭 시 호출. 인플루언서 정보 자동 채움 + 1차 세트 자동 생성."""
    influencers = _load_influencers()
    inf = next((x for x in influencers if x.get("id") == inf_id), None)
    if not inf:
        return jsonify({"error": "인플루언서 없음"}), 404

    items = _load_campaigns_v2()
    payload = request.get_json(silent=True) or {}
    cam = {
        "id": _next_id(items, "cam"),
        "seller_name": inf.get("seller_name") or inf.get("instagram_id"),
        "brand": payload.get("brand") or "",
        "product": payload.get("product") or "",
        "type": payload.get("type") or ("메가" if "메가" in (inf.get("category") or "") else "마이크로"),
        "market_schedule": payload.get("market_schedule") or "",
        "linked_influencer_id": inf_id,
        "linked_influencer_handle": inf.get("instagram_id"),
        "status": "준비중",
        "sets": [{
            "id": "set_0001",
            "round": 1,
            "label": "1차",
            "ads": [{
                "id": "ad_0001",
                "name": "공동구매 1차",
                "product_sent_date": None,
                "scheduling": {"start_date": None, "end_date": None, "items": []},
                "events": [],
                "drive_links": [],
                "banners": {
                    "openfeed": {"checked": False, "draft_url": "", "final_url": "", "note": ""},
                    "price": {"checked": False, "draft_url": "", "final_url": "", "note": ""},
                    "event": {"checked": False, "draft_url": "", "final_url": "", "note": ""},
                },
                "reels": [],
                "status": "준비중",
                "created_at": datetime.now().isoformat(timespec="seconds"),
            }],
            "created_at": datetime.now().isoformat(timespec="seconds"),
        }],
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    items.append(cam)
    _save_campaigns_v2(items)

    # 인플루언서 단계 자동 전이
    inf["pipeline_stage"] = "캠페인진행중"
    inf["campaign_id"] = cam["id"]
    inf["campaign_name"] = f"{cam['seller_name']} · {cam.get('brand') or ''}".strip(" ·")
    _save_influencers(influencers)

    return jsonify({"ok": True, "campaign": cam})


@app.route("/api/dm/daily_stats", methods=["GET"])
def api_dm_daily_stats():
    """데일리 DM 탭 상단 통계 카드."""
    try:
        from dm_scheduler import build_queue  # type: ignore
    except ImportError:
        return jsonify({"error": "scheduler 미설치"}), 500

    influencers = _load_influencers()
    accounts = _load_our_accounts()
    conversations = _load_inbox()
    today = datetime.now().strftime("%Y-%m-%d")

    q = build_queue(influencers, accounts, max_per_run=10000)
    active_accounts = sum(1 for a in accounts if a.get("status") == "활성")
    sent_today = sum(1 for inf in influencers
                     if (inf.get("last_sent_date") or "")[:10] == today)
    replies_total = sum(1 for c in conversations
                        if any(m.get("from") == "them" for m in c.get("messages", [])))

    return jsonify({
        "candidates": len(q.get("queue", [])),
        "active_accounts": active_accounts,
        "sent_today": sent_today,
        "replies": replies_total,
        "queue_reasons": q.get("reasons", {}),
    })


# ═══════════════════════════════════════════════════════════
# 📊 Phase J — 대시보드 v2 (카페24 스타일) + 함수 셀
# ═══════════════════════════════════════════════════════════

@app.route("/api/dashboard_v2", methods=["GET"])
def api_dashboard_v2():
    """월별 매출 + 일별 매출 + 캠페인별 표.
    매출 = 마켓.revenue 합산, 비용 = 마켓.cost 합산. 일자 = 마켓.scheduling.start_date.
    """
    campaigns = _load_campaigns_v2()
    # 상단 브랜드 바 필터 (브랜드명) — 있으면 그 브랜드 캠페인만 집계
    brand_filter = (request.args.get("brand") or "").strip()
    if brand_filter:
        campaigns = [c for c in campaigns if (c.get("brand") or "") == brand_filter]
    now = datetime.now()
    today_str = now.strftime("%Y-%m-%d")

    # 일별 / 월별 집계
    by_day = {}     # "YYYY-MM-DD" → {revenue, cost, market_count}
    by_month = {}   # "YYYY-MM" → {revenue, cost, market_count}
    by_campaign = []  # 캠페인별 row

    for cam in campaigns:
        cam_rev = 0
        cam_cost = 0
        cam_markets = 0
        latest_start = None
        for st in cam.get("sets", []):
            for ad in st.get("ads", []):
                rev = ad.get("revenue") or 0
                cost = ad.get("cost") or 0
                cam_rev += rev
                cam_cost += cost
                cam_markets += 1
                sd = (ad.get("scheduling") or {}).get("start_date")
                if sd and len(sd) >= 10:
                    day = sd[:10]
                    month = sd[:7]
                    d = by_day.setdefault(day, {"revenue": 0, "cost": 0, "markets": 0})
                    d["revenue"] += rev; d["cost"] += cost; d["markets"] += 1
                    m = by_month.setdefault(month, {"revenue": 0, "cost": 0, "markets": 0})
                    m["revenue"] += rev; m["cost"] += cost; m["markets"] += 1
                    if not latest_start or sd > latest_start:
                        latest_start = sd
        margin = (cam_rev - cam_cost) / cam_rev * 100 if cam_rev > 0 else None
        by_campaign.append({
            "id": cam["id"],
            "seller_name": cam.get("seller_name"),
            "brand": cam.get("brand"),
            "product": cam.get("product"),
            "type": cam.get("type"),
            "status": cam.get("status"),
            "revenue": cam_rev,
            "cost": cam_cost,
            "profit": cam_rev - cam_cost,
            "margin_pct": round(margin, 1) if margin is not None else None,
            "market_count": cam_markets,
            "latest_market_date": latest_start,
            "market_schedule": cam.get("market_schedule"),
            "settlement_done": bool(cam.get("settlement_done")),
        })

    # 정렬 — 매출 큰 순서
    by_campaign.sort(key=lambda x: -(x["revenue"] or 0))

    # 최근 12개월 리스트 (없는 달도 0 으로 박음)
    months_list = []
    cur = now.replace(day=1)
    for i in range(11, -1, -1):
        m = cur.replace(year=cur.year - (1 if cur.month - i <= 0 else 0),
                       month=((cur.month - i - 1) % 12) + 1)
        key = m.strftime("%Y-%m")
        v = by_month.get(key, {"revenue": 0, "cost": 0, "markets": 0})
        months_list.append({
            "month": key,
            "label": m.strftime("%y년 %m월"),
            "revenue": v["revenue"],
            "cost": v["cost"],
            "markets": v["markets"],
        })

    # 최근 7일 리스트
    from datetime import timedelta
    days_list = []
    for i in range(6, -1, -1):
        d = (now - timedelta(days=i)).strftime("%Y-%m-%d")
        v = by_day.get(d, {"revenue": 0, "cost": 0, "markets": 0})
        days_list.append({
            "date": d,
            "label": d[5:],
            "revenue": v["revenue"],
            "cost": v["cost"],
            "markets": v["markets"],
            "is_today": d == today_str,
        })

    total_rev = sum(c["revenue"] for c in by_campaign)
    total_cost = sum(c["cost"] for c in by_campaign)

    # ── 날짜 범위 선택 (start ~ end) → 막대 series + 기간 합계 window ──
    cur_month_key = now.strftime("%Y-%m")

    def _parse(s, default):
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d") if s else default
        except Exception:
            return default

    end_dt = _parse(request.args.get("end"), now)
    start_dt = _parse(request.args.get("start"), datetime(now.year, now.month, 1) - timedelta(days=334))
    if start_dt > end_dt:
        start_dt, end_dt = end_dt, start_dt
    span_days = (end_dt - start_dt).days
    gran = request.args.get("gran")
    if gran not in ("day", "month"):
        gran = "day" if span_days <= 62 else "month"

    s_str, e_str = start_dt.strftime("%Y-%m-%d"), end_dt.strftime("%Y-%m-%d")

    series = []
    if gran == "day":
        d = start_dt
        while d <= end_dt:
            key = d.strftime("%Y-%m-%d")
            v = by_day.get(key, {})
            series.append({"label": d.strftime("%m.%d"), "value": v.get("revenue", 0),
                           "is_current": key == today_str})
            d += timedelta(days=1)
    else:
        yy, mm = start_dt.year, start_dt.month
        while (yy < end_dt.year) or (yy == end_dt.year and mm <= end_dt.month):
            key = f"{yy:04d}-{mm:02d}"
            v = by_month.get(key, {})
            series.append({"label": f"{mm}월", "value": v.get("revenue", 0),
                           "is_current": key == cur_month_key})
            mm += 1
            if mm > 12:
                mm = 1; yy += 1

    def _window_sum(lo, hi):
        rev = cost = mk = 0
        lo_s, hi_s = lo.strftime("%Y-%m-%d"), hi.strftime("%Y-%m-%d")
        for cam in campaigns:
            for stx in cam.get("sets", []):
                for ad in stx.get("ads", []):
                    sd = (ad.get("scheduling") or {}).get("start_date")
                    if sd and len(sd) >= 10 and lo_s <= sd[:10] <= hi_s:
                        rev += ad.get("revenue") or 0
                        cost += ad.get("cost") or 0
                        mk += 1
        return rev, cost, mk

    w_rev, w_cost, w_mk = _window_sum(start_dt, end_dt)
    # 직전 동일 길이 기간
    prev_end = start_dt - timedelta(days=1)
    prev_start = prev_end - timedelta(days=span_days)
    p_rev, p_cost, p_mk = _window_sum(prev_start, prev_end)

    return jsonify({
        "today": today_str,
        "range": {"start": s_str, "end": e_str, "gran": gran},
        "totals": {
            "revenue": total_rev,
            "cost": total_cost,
            "profit": total_rev - total_cost,
            "margin_pct": round((total_rev - total_cost) / total_rev * 100, 1) if total_rev > 0 else None,
            "campaign_count": len(by_campaign),
            "market_count": sum(c["market_count"] for c in by_campaign),
        },
        "window": {
            "revenue": w_rev, "cost": w_cost, "profit": w_rev - w_cost,
            "margin_pct": round((w_rev - w_cost) / w_rev * 100, 1) if w_rev > 0 else None,
            "market_count": w_mk,
        },
        "prev_window": {"revenue": p_rev, "cost": p_cost, "market_count": p_mk},
        "series": series,
        "months": months_list,
        "days": days_list,
        "campaigns": by_campaign,
    })


@app.route("/api/dashboard_v2/cell", methods=["PATCH"])
def api_dashboard_v2_cell():
    """대시보드 셀 인라인 편집 — 캠페인의 첫 광고 revenue/cost 수정."""
    payload = request.get_json(force=True, silent=True) or {}
    cam_id = payload.get("campaign_id")
    field = payload.get("field")  # revenue | cost
    value = payload.get("value")
    if not (cam_id and field in ("revenue", "cost")):
        return jsonify({"error": "campaign_id + field 필요"}), 400
    items = _load_campaigns_v2()
    cam = next((c for c in items if c["id"] == cam_id), None)
    if not cam:
        return jsonify({"error": "캠페인 없음"}), 404
    if not cam.get("sets"):
        return jsonify({"error": "세트 없음"}), 400
    ad = (cam["sets"][0].get("ads") or [None])[0]
    if not ad:
        return jsonify({"error": "광고 없음"}), 400
    try:
        ad[field] = int(value or 0)
    except (TypeError, ValueError):
        return jsonify({"error": "숫자만 가능"}), 400
    _save_campaigns_v2(items)
    return jsonify({"ok": True, "value": ad[field]})


# ═══════════════════════════════════════════════════════════
# 🗂 Phase H — Activity Log + Google Drive 자동 sync
# ═══════════════════════════════════════════════════════════

@app.route("/api/activity/log", methods=["POST"])
def api_activity_log():
    """프론트에서 액션 1건 로깅. body: {action, tab?, target?, detail?}"""
    try:
        from activity_log import log_action  # type: ignore
    except ImportError as e:
        return jsonify({"error": f"activity_log 모듈 실패: {e}"}), 500
    payload = request.get_json(silent=True) or {}
    action = (payload.get("action") or "").strip()
    if not action:
        return jsonify({"error": "action 필수"}), 400
    log_action(
        action,
        tab=payload.get("tab") or "",
        target=payload.get("target") or "",
        detail=payload.get("detail") or {},
    )
    return jsonify({"ok": True})


@app.route("/api/activity/recent", methods=["GET"])
def api_activity_recent():
    try:
        from activity_log import read_recent  # type: ignore
    except ImportError:
        return jsonify({"entries": []})
    limit = int(request.args.get("limit", 200))
    return jsonify({"entries": read_recent(limit=limit)})


@app.route("/api/drive/sync", methods=["POST"])
def api_drive_sync():
    """Drive sync 수동 트리거. force=true 면 쿨다운 무시."""
    try:
        from activity_log import sync_to_drive  # type: ignore
    except ImportError as e:
        return jsonify({"error": f"activity_log 모듈 실패: {e}"}), 500
    force = (request.get_json(silent=True) or {}).get("force") or \
            request.args.get("force", "").lower() in ("1", "true", "yes")
    cfg = load_config() or {}
    result = sync_to_drive(cfg, force=force)
    return jsonify(result)


@app.route("/api/drive/status", methods=["GET"])
def api_drive_status():
    try:
        from activity_log import get_sync_status  # type: ignore
    except ImportError:
        return jsonify({"synced": False})
    return jsonify(get_sync_status())


# 백그라운드 sync 스케줄러 (5분마다 자동)
def _start_drive_sync_scheduler():
    import threading
    import time as _t

    def _loop():
        # 시작 시 30초 대기 (서버 부팅 안정)
        _t.sleep(30)
        while True:
            try:
                from activity_log import sync_to_drive  # type: ignore
                cfg = load_config() or {}
                if cfg.get("env_mode") == "cloud":
                    # 클라우드에서는 credentials.json 없으므로 skip
                    pass
                else:
                    sync_to_drive(cfg)
            except Exception as e:
                log.warning(f"백그라운드 Drive sync 실패: {e}")
            _t.sleep(300)  # 5분

    threading.Thread(target=_loop, daemon=True, name="drive_sync").start()


# ═══════════════════════════════════════════════════════════
# 📨 DM 영업 시스템 — Legacy (구버전, 곧 마이그레이션 예정)
# ═══════════════════════════════════════════════════════════

DM_ACCOUNTS_FILE = DATA_DIR / "dm_accounts.json"
DM_TARGETS_FILE = DATA_DIR / "dm_targets.json"
DM_TEMPLATES_FILE = DATA_DIR / "dm_templates.json"
DM_JOBS_FILE = DATA_DIR / "dm_jobs.json"


def _dm_read(path: Path, key: str) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    try:
        return json.loads(path.read_text(encoding="utf-8")).get(key, [])
    except Exception:
        return []


def _dm_write(path: Path, key: str, items: list[dict[str, Any]]) -> None:
    path.write_text(json.dumps({key: items}, ensure_ascii=False, indent=2), encoding="utf-8")


# ─── 계정 풀 ──────────────────────────────────────────────
@app.route("/api/dm/accounts", methods=["GET"])
def api_dm_accounts_list():
    items = _dm_read(DM_ACCOUNTS_FILE, "accounts")
    # 비밀번호 마스킹
    out = []
    for a in items:
        masked = dict(a)
        pw = masked.get("password", "")
        if pw:
            masked["password"] = "*" * 8
        out.append(masked)
    return jsonify({"accounts": out})


@app.route("/api/dm/accounts", methods=["POST"])
def api_dm_accounts_add():
    payload = request.get_json(force=True)
    items = _dm_read(DM_ACCOUNTS_FILE, "accounts")
    now = datetime.now().isoformat(timespec="seconds")
    new = {
        "id": _next_id(items, "a"),
        "username": (payload.get("username") or "").strip().lstrip("@"),
        "password": payload.get("password") or "",
        "sender_name": (payload.get("sender_name") or "").strip(),
        "status": "active",  # active / warmup / blocked / disabled
        "daily_count": 0,
        "daily_limit": int(payload.get("daily_limit") or 50),
        "total_sent": 0,
        "last_used_at": None,
        "last_reset_date": now[:10],
        "notes": (payload.get("notes") or "").strip(),
        "created_at": now,
        "updated_at": now,
    }
    if not new["username"] or not new["password"]:
        return jsonify({"error": "username과 password 필수"}), 400
    items.append(new)
    _dm_write(DM_ACCOUNTS_FILE, "accounts", items)
    return jsonify({"account": {**new, "password": "*" * 8}})


@app.route("/api/dm/accounts/bulk", methods=["POST"])
def api_dm_accounts_bulk():
    """엑셀로 100개 한 번에 박기. CSV/JSON 둘 다 지원."""
    payload = request.get_json(force=True)
    rows = payload.get("rows") or []
    items = _dm_read(DM_ACCOUNTS_FILE, "accounts")
    now = datetime.now().isoformat(timespec="seconds")
    added = 0
    existing = {a["username"] for a in items}
    for row in rows:
        u = (row.get("username") or "").strip().lstrip("@")
        p = row.get("password") or ""
        if not u or not p or u in existing:
            continue
        existing.add(u)
        items.append({
            "id": _next_id(items, "a"),
            "username": u,
            "password": p,
            "sender_name": (row.get("sender_name") or "").strip(),
            "status": "active",
            "daily_count": 0,
            "daily_limit": int(row.get("daily_limit") or 50),
            "total_sent": 0,
            "last_used_at": None,
            "last_reset_date": now[:10],
            "notes": "",
            "created_at": now,
            "updated_at": now,
        })
        added += 1
    _dm_write(DM_ACCOUNTS_FILE, "accounts", items)
    return jsonify({"added": added, "total": len(items)})


@app.route("/api/dm/accounts/<aid>", methods=["PATCH"])
def api_dm_accounts_patch(aid: str):
    payload = request.get_json(force=True)
    items = _dm_read(DM_ACCOUNTS_FILE, "accounts")
    a = next((x for x in items if x["id"] == aid), None)
    if not a:
        return jsonify({"error": "not found"}), 404
    for k in ("username", "password", "sender_name", "status", "daily_limit", "notes"):
        if k in payload and payload[k] != "":
            v = payload[k]
            if k == "username":
                v = str(v).strip().lstrip("@")
            elif k == "daily_limit":
                v = int(v)
            elif isinstance(v, str):
                v = v.strip()
            a[k] = v
    a["updated_at"] = datetime.now().isoformat(timespec="seconds")
    _dm_write(DM_ACCOUNTS_FILE, "accounts", items)
    return jsonify({"account": {**a, "password": "*" * 8}})


@app.route("/api/dm/accounts/<aid>", methods=["DELETE"])
def api_dm_accounts_delete(aid: str):
    items = [a for a in _dm_read(DM_ACCOUNTS_FILE, "accounts") if a["id"] != aid]
    _dm_write(DM_ACCOUNTS_FILE, "accounts", items)
    return jsonify({"ok": True})


# ─── 영업 명단 (타겟) ─────────────────────────────────────
@app.route("/api/dm/targets", methods=["GET"])
def api_dm_targets_list():
    return jsonify({"targets": _dm_read(DM_TARGETS_FILE, "targets")})


@app.route("/api/dm/targets", methods=["POST"])
def api_dm_targets_add():
    payload = request.get_json(force=True)
    items = _dm_read(DM_TARGETS_FILE, "targets")
    now = datetime.now().isoformat(timespec="seconds")
    new = {
        "id": _next_id(items, "t"),
        "username": (payload.get("username") or "").strip().lstrip("@"),
        "display_name": (payload.get("display_name") or "").strip(),
        "category": (payload.get("category") or "").strip(),
        "followers": int(payload.get("followers") or 0),
        "status": "pending",  # pending / sending / sent / replied / failed
        "notes": (payload.get("notes") or "").strip(),
        "last_sent_at": None,
        "last_sent_account": None,
        "reply_received": False,
        "created_at": now,
    }
    if not new["username"]:
        return jsonify({"error": "username 필수"}), 400
    items.append(new)
    _dm_write(DM_TARGETS_FILE, "targets", items)
    return jsonify({"target": new})


@app.route("/api/dm/targets/bulk", methods=["POST"])
def api_dm_targets_bulk():
    payload = request.get_json(force=True)
    rows = payload.get("rows") or []
    items = _dm_read(DM_TARGETS_FILE, "targets")
    now = datetime.now().isoformat(timespec="seconds")
    added = 0
    existing = {t["username"] for t in items}
    for row in rows:
        u = (row.get("username") or "").strip().lstrip("@")
        if not u or u in existing:
            continue
        existing.add(u)
        items.append({
            "id": _next_id(items, "t"),
            "username": u,
            "display_name": (row.get("display_name") or "").strip(),
            "category": (row.get("category") or "").strip(),
            "followers": int(row.get("followers") or 0),
            "status": "pending",
            "notes": (row.get("notes") or "").strip(),
            "last_sent_at": None,
            "last_sent_account": None,
            "reply_received": False,
            "created_at": now,
        })
        added += 1
    _dm_write(DM_TARGETS_FILE, "targets", items)
    return jsonify({"added": added, "total": len(items)})


@app.route("/api/dm/targets/<tid>", methods=["PATCH"])
def api_dm_targets_patch(tid: str):
    payload = request.get_json(force=True)
    items = _dm_read(DM_TARGETS_FILE, "targets")
    t = next((x for x in items if x["id"] == tid), None)
    if not t:
        return jsonify({"error": "not found"}), 404
    for k in ("display_name", "category", "followers", "status", "notes", "reply_received"):
        if k in payload:
            v = payload[k]
            if k == "followers":
                v = int(v or 0)
            elif isinstance(v, str):
                v = v.strip()
            t[k] = v
    _dm_write(DM_TARGETS_FILE, "targets", items)
    return jsonify({"target": t})


@app.route("/api/dm/targets/<tid>", methods=["DELETE"])
def api_dm_targets_delete(tid: str):
    items = [t for t in _dm_read(DM_TARGETS_FILE, "targets") if t["id"] != tid]
    _dm_write(DM_TARGETS_FILE, "targets", items)
    return jsonify({"ok": True})


# ─── 템플릿 ──────────────────────────────────────────────
@app.route("/api/dm/templates", methods=["GET"])
def api_dm_templates_list():
    return jsonify({"templates": _dm_read(DM_TEMPLATES_FILE, "templates")})


@app.route("/api/dm/templates", methods=["POST"])
def api_dm_templates_add():
    payload = request.get_json(force=True)
    items = _dm_read(DM_TEMPLATES_FILE, "templates")
    new = {
        "id": _next_id(items, "t"),
        "name": (payload.get("name") or "").strip(),
        "brand_id": (payload.get("brand_id") or "").strip(),
        "sender_name": (payload.get("sender_name") or "").strip(),
        "body": (payload.get("body") or "").strip(),
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    if not new["name"] or not new["body"]:
        return jsonify({"error": "name과 body 필수"}), 400
    items.append(new)
    _dm_write(DM_TEMPLATES_FILE, "templates", items)
    return jsonify({"template": new})


@app.route("/api/dm/templates/<tid>", methods=["PATCH"])
def api_dm_templates_patch(tid: str):
    payload = request.get_json(force=True)
    items = _dm_read(DM_TEMPLATES_FILE, "templates")
    t = next((x for x in items if x["id"] == tid), None)
    if not t:
        return jsonify({"error": "not found"}), 404
    for k in ("name", "brand_id", "sender_name", "body"):
        if k in payload:
            t[k] = (payload[k] or "").strip()
    _dm_write(DM_TEMPLATES_FILE, "templates", items)
    return jsonify({"template": t})


@app.route("/api/dm/templates/<tid>", methods=["DELETE"])
def api_dm_templates_delete(tid: str):
    items = [t for t in _dm_read(DM_TEMPLATES_FILE, "templates") if t["id"] != tid]
    _dm_write(DM_TEMPLATES_FILE, "templates", items)
    return jsonify({"ok": True})


# ─── 발송 작업 ────────────────────────────────────────────
DM_JOBS_STATE: dict[str, dict] = {}  # 메모리에 진행률 추적


def _select_available_account(accounts: list[dict]) -> dict | None:
    """발송 가능한 계정 선택 — 오늘 한도 안 차고, 활성 상태."""
    today = datetime.now().date().isoformat()
    candidates = []
    for a in accounts:
        if a.get("status") != "active":
            continue
        # 일일 카운트 리셋
        if a.get("last_reset_date") != today:
            a["daily_count"] = 0
            a["last_reset_date"] = today
        if a.get("daily_count", 0) >= a.get("daily_limit", 50):
            continue
        candidates.append(a)
    if not candidates:
        return None
    # 마지막 사용 시간 오래된 거 우선
    candidates.sort(key=lambda x: x.get("last_used_at") or "")
    return candidates[0]


@app.route("/api/dm/send", methods=["POST"])
def api_dm_send():
    """선택한 target들에게 DM 발송 시작."""
    payload = request.get_json(force=True)
    target_ids = payload.get("target_ids") or []
    template_id = payload.get("template_id")
    if not target_ids:
        return jsonify({"error": "target_ids 필수"}), 400
    if not template_id:
        return jsonify({"error": "template_id 필수"}), 400

    accounts = _dm_read(DM_ACCOUNTS_FILE, "accounts")
    if not [a for a in accounts if a.get("status") == "active"]:
        return jsonify({"error": "활성 계정 없음. 먼저 계정 풀에 등록."}), 400

    templates = _dm_read(DM_TEMPLATES_FILE, "templates")
    template = next((t for t in templates if t["id"] == template_id), None)
    if not template:
        return jsonify({"error": "템플릿 없음"}), 400

    targets = _dm_read(DM_TARGETS_FILE, "targets")
    selected = [t for t in targets if t["id"] in target_ids]
    if not selected:
        return jsonify({"error": "선택된 타겟 없음"}), 400

    # cloud 모드면 실제 발송 X — UI만 제공
    cfg = load_config()
    if cfg.get("env_mode") == "cloud":
        return jsonify({
            "error": "DM 발송은 로컬 PC에서만 가능. 클라우드 모드에서는 명단 관리만.",
            "note": "PC에서 워크스페이스 켜고 발송하세요."
        }), 400

    # 작업 ID 생성
    job_id = uuid.uuid4().hex[:12]
    DM_JOBS_STATE[job_id] = {
        "id": job_id,
        "status": "running",
        "total": len(selected),
        "sent": 0,
        "failed": 0,
        "current": None,
        "log": [],
        "started_at": datetime.now().isoformat(timespec="seconds"),
        "finished_at": None,
    }

    def run():
        try:
            from dm_sender import DMSender  # type: ignore
            sender = DMSender(state=DM_JOBS_STATE[job_id], log_callback=_log_callback(job_id))
            sender.run_batch(
                targets=selected,
                template=template,
                accounts_file=str(DM_ACCOUNTS_FILE),
                targets_file=str(DM_TARGETS_FILE),
            )
            DM_JOBS_STATE[job_id]["status"] = "done"
            DM_JOBS_STATE[job_id]["finished_at"] = datetime.now().isoformat(timespec="seconds")
        except Exception as e:  # noqa: BLE001
            log.exception("DM send job failed")
            DM_JOBS_STATE[job_id]["status"] = "error"
            DM_JOBS_STATE[job_id]["log"].append(f"❌ 에러: {e}")
            DM_JOBS_STATE[job_id]["finished_at"] = datetime.now().isoformat(timespec="seconds")

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"job_id": job_id})


def _log_callback(job_id: str):
    def cb(msg: str):
        if job_id in DM_JOBS_STATE:
            st = DM_JOBS_STATE[job_id]
            st["last_beat"] = datetime.now().isoformat(timespec="seconds")  # 심장박동
            logs = st["log"]
            logs.append(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")
            if len(logs) > 200:
                st["log"] = logs[-200:]
    return cb


@app.route("/api/dm/jobs/<jid>", methods=["GET"])
def api_dm_job_status(jid: str):
    state = DM_JOBS_STATE.get(jid)
    if not state:
        return jsonify({"error": "not found"}), 404
    return jsonify(state)


@app.route("/api/dm/jobs/<jid>/stop", methods=["POST"])
def api_dm_job_stop(jid: str):
    state = DM_JOBS_STATE.get(jid)
    if state:
        state["status"] = "stopping"
    return jsonify({"ok": True})


@app.route("/api/dm/status", methods=["GET"])
def api_dm_status():
    """발송 엔진 실시간 상태 — 링크 어디서나 '지금 발송 중인지' 확인.
    engine: running(가동) | stale(멈춘 듯) | idle(대기) | error."""
    if not DM_JOBS_STATE:
        return jsonify({"engine": "idle", "job": None})
    running = [j for j in DM_JOBS_STATE.values() if j.get("status") in ("running", "stopping")]
    if running:
        job = max(running, key=lambda j: j.get("last_beat") or j.get("started_at") or "")
    else:
        job = max(DM_JOBS_STATE.values(), key=lambda j: j.get("started_at") or "")
    last = job.get("last_beat") or job.get("started_at")
    age = None
    try:
        age = (datetime.now() - datetime.fromisoformat(last)).total_seconds()
    except Exception:
        age = None
    status = job.get("status")
    if status in ("running", "stopping"):
        # 발송 간격/휴식(최대 10분) 고려 — 12분 이상 무신호면 멈춘 것으로 판단
        engine = "stale" if (age is not None and age > 720) else "running"
    elif status == "error":
        engine = "error"
    else:
        engine = "idle"
    return jsonify({
        "engine": engine,
        "age_seconds": int(age) if age is not None else None,
        "last_beat": last,
        "job": {k: job.get(k) for k in ("id", "kind", "status", "total", "sent",
                                        "failed", "held", "current", "started_at", "finished_at")},
    })


# ═══════════════════════════════════════════════════════════
# 🎬 콘텐츠 워크스페이스 — 소재 기획안 AI 자동생성 (AI Studio 이식)
# ═══════════════════════════════════════════════════════════
@app.route("/content")
def content_workspace():
    return render_template("content.html", ver=_asset_ver())


@app.route("/api/content/analyze", methods=["POST"])
def api_content_analyze():
    """광고영상 업로드 → 자막/나레이션/연출 표 추출."""
    f = request.files.get("video")
    if not f:
        return jsonify({"error": "영상 파일을 올려주세요."}), 400
    feedback = request.form.get("feedback") or ""
    try:
        from modules import content_studio
        rows = content_studio.analyze_video(load_config(), f.read(), f.mimetype or "video/mp4", feedback)
        return jsonify({"analysis": rows})
    except Exception as e:  # noqa: BLE001
        log.exception("content analyze failed")
        return jsonify({"error": str(e)}), 500


CONTENT_PLANS_FILE = DATA_DIR / "content_plans.json"


def _load_content_plans() -> list[dict]:
    if not CONTENT_PLANS_FILE.exists():
        return []
    try:
        return json.loads(CONTENT_PLANS_FILE.read_text(encoding="utf-8")).get("plans", [])
    except Exception:
        return []


def _save_content_plans(items: list[dict]) -> None:
    CONTENT_PLANS_FILE.write_text(json.dumps({"plans": items}, ensure_ascii=False, indent=2), encoding="utf-8")


@app.route("/api/content/plan", methods=["POST"])
def api_content_plan():
    """레퍼런스 분석 + 제품(USP) → 새 기획안 생성.
    product_id가 오면 그 제품의 '확정 기획안'들을 학습 참고로 함께 넣는다(누적 학습)."""
    p = request.get_json(force=True, silent=True) or {}
    try:
        from modules import content_studio
        history = []
        pid = (p.get("product_id") or "").strip()
        if pid:
            history = [x for x in _load_content_plans() if x.get("product_id") == pid]
        result = content_studio.generate_plan(
            load_config(), p.get("analysis") or [], p.get("product") or {},
            p.get("feedback") or "", history=history)
        return jsonify({"plan": result.get("plan") or [], "why_watch": result.get("why_watch", ""),
                        "why_buy": result.get("why_buy", ""), "learned_from": len(history)})
    except Exception as e:  # noqa: BLE001
        log.exception("content plan failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/content/plans", methods=["GET", "POST"])
def api_content_plans():
    """확정 기획안 라이브러리 — 제품별로 '초안→최종' 누적 적재.
    GET: ?product_id= / ?brand= / ?op_type= 필터. POST: 확정본 저장."""
    items = _load_content_plans()
    if request.method == "POST":
        p = request.get_json(force=True, silent=True) or {}
        final = p.get("final") or []
        if not final:
            return jsonify({"error": "확정할 기획안이 비어 있습니다."}), 400
        rec = {
            "id": uuid.uuid4().hex[:8],
            "product_id": (p.get("product_id") or "").strip(),
            "product_name": (p.get("product_name") or "").strip(),
            "brand": (p.get("brand") or "").strip(),
            "op_type": p.get("op_type") if p.get("op_type") in ("own", "agency") else "own",
            "title": (p.get("title") or "").strip(),
            "appeals": p.get("appeals") or [],
            "hook_angle": (p.get("hook_angle") or "").strip(),
            "reference": p.get("reference") or [],
            "draft": p.get("draft") or [],
            "final": final,
            "note": (p.get("note") or "").strip(),
            "why_watch": (p.get("why_watch") or "").strip(),
            "why_buy": (p.get("why_buy") or "").strip(),
            "created_at": datetime.now().isoformat(timespec="seconds"),
        }
        items.append(rec)
        _save_content_plans(items)
        return jsonify({"ok": True, "plan": rec})

    pid = request.args.get("product_id")
    brand = request.args.get("brand")
    op = request.args.get("op_type")
    out = items
    if pid:
        out = [x for x in out if x.get("product_id") == pid]
    if brand:
        out = [x for x in out if x.get("brand") == brand]
    if op in ("own", "agency"):
        out = [x for x in out if x.get("op_type") == op]
    out = sorted(out, key=lambda x: x.get("created_at", ""), reverse=True)
    return jsonify({"plans": out})


@app.route("/api/content/plans/<plan_id>", methods=["DELETE"])
def api_content_plan_item_delete(plan_id):
    _save_content_plans([x for x in _load_content_plans() if x.get("id") != plan_id])
    return jsonify({"ok": True})


@app.route("/api/content/shoot", methods=["POST"])
def api_content_shoot():
    """기획안 → 컷별 촬영 콘티(샷 리스트)."""
    p = request.get_json(force=True, silent=True) or {}
    try:
        from modules import content_studio
        shots = content_studio.generate_shoot_plan(load_config(), p.get("plan") or [], p.get("product") or {})
        return jsonify({"shots": shots})
    except Exception as e:  # noqa: BLE001
        log.exception("content shoot failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/content/shoot/schedule", methods=["POST"])
def api_content_shoot_schedule():
    """여러 기획안 → 장소별 동선 촬영 스케줄."""
    p = request.get_json(force=True, silent=True) or {}
    try:
        from modules import content_studio
        sched = content_studio.generate_shoot_schedule(load_config(), p.get("plans") or [], p.get("product") or {})
        return jsonify(sched)
    except Exception as e:  # noqa: BLE001
        log.exception("content shoot schedule failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/content/shoot/docx", methods=["POST"])
def api_content_shoot_docx():
    """촬영 스케줄(JSON) → Word(.docx) 다운로드."""
    p = request.get_json(force=True, silent=True) or {}
    try:
        from modules import shoot_docx
        data = shoot_docx.build_docx(p.get("schedule") or {}, p.get("meta") or {})
        fname = (p.get("filename") or "촬영스케줄") + ".docx"
        resp = make_response(data)
        resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        from urllib.parse import quote
        resp.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(fname)}"
        return resp
    except Exception as e:  # noqa: BLE001
        log.exception("content shoot docx failed")
        return jsonify({"error": str(e)}), 500


# ─── 제작 관리 (제목·날짜·사용자·브랜드·제품·분류·비고, 누적·공유) ───
CONTENT_PRODUCTIONS_FILE = DATA_DIR / "content_productions.json"


def _load_content_productions() -> list[dict]:
    if not CONTENT_PRODUCTIONS_FILE.exists():
        return []
    try:
        return json.loads(CONTENT_PRODUCTIONS_FILE.read_text(encoding="utf-8")).get("rows", [])
    except Exception:
        return []


def _save_content_productions(rows: list[dict]) -> None:
    CONTENT_PRODUCTIONS_FILE.write_text(json.dumps({"rows": rows}, ensure_ascii=False, indent=2), encoding="utf-8")


@app.route("/api/content/productions", methods=["GET", "POST"])
def api_content_productions():
    """제작 관리 테이블 — 제목/날짜/사용자/브랜드/제품/분류(shoot|noshoot)/비고."""
    rows = _load_content_productions()
    if request.method == "POST":
        p = request.get_json(force=True, silent=True) or {}
        rid = p.get("id")
        rec = next((x for x in rows if x.get("id") == rid), None) if rid else None
        if not rec:
            rec = {"id": uuid.uuid4().hex[:8], "created_at": datetime.now().isoformat(timespec="seconds")}
            rows.append(rec)
        for k in ("title", "date", "user", "brand", "product", "product_id", "note"):
            if k in p:
                rec[k] = (p.get(k) or "").strip()
        if "category" in p:
            rec["category"] = p.get("category") if p.get("category") in ("shoot", "noshoot") else "noshoot"
        if "projects" in p and isinstance(p["projects"], list):  # 항목별 기획안(분석·플랜·탭) 영속 저장
            rec["projects"] = p["projects"]
        if not rec.get("title"):
            return jsonify({"error": "제목을 입력하세요."}), 400
        rec.setdefault("category", "noshoot")
        rec.setdefault("date", datetime.now().strftime("%Y-%m-%d"))
        _save_content_productions(rows)
        return jsonify({"ok": True, "row": rec})
    rows = sorted(rows, key=lambda x: (x.get("date", ""), x.get("created_at", "")), reverse=True)
    return jsonify({"rows": rows})


@app.route("/api/content/productions/<row_id>", methods=["DELETE"])
def api_content_production_delete(row_id):
    _save_content_productions([x for x in _load_content_productions() if x.get("id") != row_id])
    return jsonify({"ok": True})


# ─── 효율 분석 (메타 마케팅 API 연동) ───
# 토큰은 사용자가 직접 입력 → 이 파일에만 저장(.gitignore). 응답에 토큰 원문은 절대 노출 X.
META_CONFIG_FILE = DATA_DIR / "meta_config.json"


def _load_meta_config() -> dict:
    if not META_CONFIG_FILE.exists():
        return {"token": "", "accounts": []}
    try:
        d = json.loads(META_CONFIG_FILE.read_text(encoding="utf-8"))
        return {"token": d.get("token", ""), "accounts": d.get("accounts", [])}
    except Exception:
        return {"token": "", "accounts": []}


def _save_meta_config(cfg: dict) -> None:
    META_CONFIG_FILE.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")


@app.route("/api/content/meta/config", methods=["GET", "POST"])
def api_meta_config():
    """메타 연결 설정. GET은 토큰 원문 대신 연결여부만 반환. POST로 토큰/계정 저장."""
    cfg = _load_meta_config()
    if request.method == "POST":
        p = request.get_json(force=True, silent=True) or {}
        if "token" in p and (p.get("token") or "").strip():
            cfg["token"] = p["token"].strip()  # 새 토큰 들어오면 교체
        if p.get("clear_token"):
            cfg["token"] = ""
        if "accounts" in p and isinstance(p["accounts"], list):
            cfg["accounts"] = [
                {"id": (a.get("id") or "").strip().replace("act_", ""),
                 "brand": (a.get("brand") or "").strip(),
                 "name": (a.get("name") or "").strip()}
                for a in p["accounts"] if (a.get("id") or "").strip()
            ]
        _save_meta_config(cfg)
    return jsonify({"connected": bool(cfg.get("token")), "accounts": cfg.get("accounts", [])})


@app.route("/api/content/meta/verify", methods=["POST"])
def api_meta_verify():
    """현재 저장된 토큰(또는 요청에 담긴 토큰)으로 접근 가능한 광고계정 목록 확인."""
    from modules import meta_ads
    p = request.get_json(force=True, silent=True) or {}
    token = (p.get("token") or "").strip() or _load_meta_config().get("token", "")
    try:
        return jsonify(meta_ads.verify_token(token))
    except Exception as e:  # noqa: BLE001
        return jsonify({"error": str(e)}), 400


@app.route("/api/content/perf", methods=["POST"])
def api_content_perf():
    """메타 광고 성과 조회 (지출·ROAS·CTR·구매·CPA 등)."""
    from modules import meta_ads
    p = request.get_json(force=True, silent=True) or {}
    cfg = _load_meta_config()
    try:
        rows = meta_ads.fetch_insights(
            cfg.get("token", ""), p.get("account_id") or "",
            p.get("date_preset") or "last_7d", p.get("level") or "campaign")
        return jsonify({"rows": rows})
    except Exception as e:  # noqa: BLE001
        return jsonify({"error": str(e)}), 400


CONTENT_PRODUCTS_FILE = DATA_DIR / "content_products.json"


def _load_content_products() -> list[dict]:
    if not CONTENT_PRODUCTS_FILE.exists():
        return []
    try:
        return json.loads(CONTENT_PRODUCTS_FILE.read_text(encoding="utf-8")).get("products", [])
    except Exception:
        return []


def _save_content_products(items: list[dict]) -> None:
    CONTENT_PRODUCTS_FILE.write_text(json.dumps({"products": items}, ensure_ascii=False, indent=2), encoding="utf-8")


@app.route("/api/content/products", methods=["GET", "POST"])
def api_content_products():
    """콘텐츠용 제품 정보 — 브랜드·제품·소구점(USP)·특이사항."""
    items = _load_content_products()
    if request.method == "POST":
        p = request.get_json(force=True, silent=True) or {}
        pid = p.get("id")
        rec = next((x for x in items if x.get("id") == pid), None) if pid else None
        if not rec:
            rec = {"id": uuid.uuid4().hex[:8], "op_type": "own",
                   "created_at": datetime.now().isoformat(timespec="seconds")}
            items.append(rec)
        for k in ("brand", "product", "usp", "notes", "op_type"):
            if k in p:
                rec[k] = (p.get(k) or "").strip()
        if "appeals" in p:  # 소구점(여러 개)
            ap = p.get("appeals") or []
            if isinstance(ap, str):
                ap = ap.replace(",", "\n").split("\n")
            rec["appeals"] = [a.strip() for a in ap if a and a.strip()]
        if rec.get("op_type") not in ("own", "agency"):
            rec["op_type"] = "own"
        if not rec.get("product") and not rec.get("brand"):
            return jsonify({"error": "브랜드 또는 제품명을 입력하세요."}), 400
        _save_content_products(items)
        return jsonify({"ok": True, "product": rec})
    return jsonify({"products": items})


@app.route("/api/content/products/<pid>", methods=["DELETE"])
def api_content_product_delete(pid):
    _save_content_products([x for x in _load_content_products() if x.get("id") != pid])
    return jsonify({"ok": True})


@app.route("/api/content/usp", methods=["POST"])
def api_content_usp():
    """상세페이지 URL 또는 파일(PDF/이미지) → 제품명·USP 자동추출."""
    try:
        from modules import content_studio
        f = request.files.get("file")
        if f:
            info = content_studio.extract_usp_file(load_config(), f.read(), f.mimetype or "application/pdf")
        else:
            url = (request.get_json(silent=True) or {}).get("url") or request.form.get("url") or ""
            if not url.startswith("http"):
                return jsonify({"error": "URL 또는 파일을 입력하세요."}), 400
            info = content_studio.extract_usp_url(load_config(), url)
        return jsonify({"product": info})
    except Exception as e:  # noqa: BLE001
        log.exception("content usp failed")
        return jsonify({"error": str(e)}), 500


# ═══════════════════════════════════════════════════════════
# 📨 엑셀 기반 DM 자동발송 (회사 DM_Sender_GUI v2.2.2 양식 그대로)
#   입력 6열: 발신ID / 발신PW / 발신자이름[name] / 타겟ID / 타겟이름[targetname] / 내용
#   결과 +2열: 발송상태(성공/실패) / 실패사유
#   ⚠ 로컬 PC 전용 (인스타가 클라우드 IP 차단 + Playwright 브라우저 필요)
# ═══════════════════════════════════════════════════════════
DM_EXCEL_COLS = ["sender_id", "sender_pw", "sender_name", "target_id", "target_name", "message"]
DM_EXCEL_HEADERS = ["발신계정아이디", "발신계정비밀번호", "발신자 이름[=name]",
                    "DM받을사람 계정", "DM받을사람 이름[=targetname]", "DM내용"]
DM_EXCEL_ROWS: dict[str, list[dict]] = {}  # job_id → rows (비밀번호 포함 → 상태응답엔 안 실림)


def _parse_dm_excel(file_storage) -> list[dict]:
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_storage.read()), read_only=True, data_only=True)
    ws = wb.active
    rows: list[dict] = []
    for i, vals in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # 헤더
        v = list(vals or []) + [None] * 6
        if all(v[j] is None or str(v[j]).strip() == "" for j in range(6)):
            continue
        row = {DM_EXCEL_COLS[j]: ("" if v[j] is None else str(v[j])) for j in range(6)}
        if not row["sender_id"].strip() or not row["target_id"].strip():
            continue
        row["target_id"] = row["target_id"].strip().lstrip("@")
        rows.append(row)
    return rows


@app.route("/api/dm/excel/run", methods=["POST"])
def api_dm_excel_run():
    """엑셀 업로드 → 자동 DM 발송 시작 (로컬 전용)."""
    if (load_config() or {}).get("env_mode") == "cloud":
        return jsonify({"error": "DM 발송은 로컬 PC에서만 됩니다. PC에서 워크스페이스를 켜고 실행하세요."}), 400
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "엑셀 파일을 올려주세요."}), 400
    auto_follow = request.form.get("auto_follow") in ("1", "true", "True", "on")

    def _intf(name, default):
        try:
            return int(request.form.get(name) or default)
        except Exception:
            return default
    # 안전(밴 회피) 옵션 — 기본 = 안전 모드
    opts = {
        "daily_limit": _intf("daily_limit", 30),
        "batch_limit": _intf("batch_limit", 10),
        "gap_min": _intf("gap_min", 60),
        "gap_max": _intf("gap_max", 300),
        "break_every": _intf("break_every", 6),
        "break_min": _intf("break_min", 180),
        "break_max": _intf("break_max", 600),
    }
    try:
        rows = _parse_dm_excel(f)
    except Exception as e:  # noqa: BLE001
        return jsonify({"error": f"엑셀 읽기 실패: {e}"}), 400
    if not rows:
        return jsonify({"error": "발송할 행이 없습니다. (1행=헤더, 2행부터 데이터 / 발신ID·타겟ID 필수)"}), 400

    job_id = uuid.uuid4().hex[:12]
    DM_EXCEL_ROWS[job_id] = rows
    DM_JOBS_STATE[job_id] = {
        "id": job_id, "kind": "excel", "status": "running",
        "total": len(rows), "sent": 0, "failed": 0, "held": 0, "current": None, "log": [],
        "accounts": len({r["sender_id"] for r in rows}), "auto_follow": auto_follow, "opts": opts,
        "started_at": datetime.now().isoformat(timespec="seconds"), "finished_at": None,
    }

    def run():
        try:
            from dm_sender import DMSender  # type: ignore
            sender = DMSender(state=DM_JOBS_STATE[job_id], log_callback=_log_callback(job_id))
            sender.run_excel_rows(rows, auto_follow=auto_follow, opts=opts)
            DM_JOBS_STATE[job_id]["status"] = "done"
        except Exception as e:  # noqa: BLE001
            log.exception("DM excel job failed")
            DM_JOBS_STATE[job_id]["status"] = "error"
            DM_JOBS_STATE[job_id]["log"].append(f"❌ 에러: {e}")
        finally:
            DM_JOBS_STATE[job_id]["finished_at"] = datetime.now().isoformat(timespec="seconds")

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"job_id": job_id, "total": len(rows),
                    "accounts": DM_JOBS_STATE[job_id]["accounts"]})


@app.route("/api/dm/excel/result/<jid>", methods=["GET"])
def api_dm_excel_result(jid):
    """결과 엑셀 다운로드 (입력 6열 + 발송상태/실패사유)."""
    import io
    import openpyxl
    from flask import send_file
    rows = DM_EXCEL_ROWS.get(jid)
    if rows is None:
        return jsonify({"error": "결과 없음 (작업 ID 확인)"}), 404
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(DM_EXCEL_HEADERS + ["발송상태", "실패사유"])
    for r in rows:
        ws.append([r.get("sender_id", ""), r.get("sender_pw", ""), r.get("sender_name", ""),
                   r.get("target_id", ""), r.get("target_name", ""), r.get("message", ""),
                   r.get("status", ""), r.get("reason", "")])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"DMresult_{datetime.now().strftime('%y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/dm/excel/sample", methods=["GET"])
def api_dm_excel_sample():
    """입력 샘플 엑셀 다운로드."""
    import io
    import openpyxl
    from flask import send_file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(DM_EXCEL_HEADERS)
    ws.append(["sampleID", "samplePW", "하루픽스", "받는사람_인스타ID", "받는분이름",
               "안녕하세요 [targetname]님! 하루픽스 [name]입니다 :)\n"
               "공구 제안 드리고 싶어 연락드렸어요. 잠깐 얘기 나눠볼 수 있을까요?"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="DM발송_샘플양식.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ─── DM 발신 계정 저장 + 수동 발송 (테스트/소량용) ───
DM_SENDER_ACCOUNTS_FILE = DATA_DIR / "dm_sender_accounts.json"


def _load_sender_accounts() -> list[dict]:
    if not DM_SENDER_ACCOUNTS_FILE.exists():
        return []
    try:
        return json.loads(DM_SENDER_ACCOUNTS_FILE.read_text(encoding="utf-8")).get("accounts", [])
    except Exception:
        return []


def _save_sender_accounts(items: list[dict]) -> None:
    DM_SENDER_ACCOUNTS_FILE.write_text(
        json.dumps({"accounts": items}, ensure_ascii=False, indent=2), encoding="utf-8")


@app.route("/api/dm/sender-accounts", methods=["GET", "POST"])
def api_dm_sender_accounts():
    """내 인스타 발신 계정 저장. POST{username,password,name}. GET은 비번 마스킹."""
    items = _load_sender_accounts()
    if request.method == "POST":
        p = request.get_json(force=True, silent=True) or {}
        username = (p.get("username") or "").strip().lstrip("@")
        password = (p.get("password") or "").strip()
        name = (p.get("name") or "").strip()
        if not username or not password:
            return jsonify({"error": "아이디와 비밀번호를 입력하세요."}), 400
        ex = next((a for a in items if (a.get("username") or "").lower() == username.lower()), None)
        if ex:
            ex["password"] = password
            if name:
                ex["name"] = name
            acc = ex
        else:
            acc = {"id": uuid.uuid4().hex[:8], "username": username, "password": password,
                   "name": name, "created_at": datetime.now().isoformat(timespec="seconds")}
            items.append(acc)
        _save_sender_accounts(items)
        return jsonify({"ok": True, "account": {"id": acc["id"], "username": acc["username"], "name": acc.get("name", "")}})
    return jsonify({"accounts": [{"id": a["id"], "username": a["username"], "name": a.get("name", ""),
                                  "has_pw": bool(a.get("password"))} for a in items]})


@app.route("/api/dm/sender-accounts/<aid>", methods=["DELETE"])
def api_dm_sender_account_delete(aid):
    _save_sender_accounts([a for a in _load_sender_accounts() if a.get("id") != aid])
    return jsonify({"ok": True})


@app.route("/api/dm/manual/run", methods=["POST"])
def api_dm_manual_run():
    """수동 발송 — 저장계정 선택(or 인라인) + 받는사람 + 메시지 → 안전엔진 재사용."""
    if (load_config() or {}).get("env_mode") == "cloud":
        return jsonify({"error": "DM 발송은 로컬 PC에서만 됩니다. PC에서 워크스페이스를 켜고 실행하세요."}), 400
    p = request.get_json(force=True, silent=True) or {}
    # 발신 계정 — 저장된 것 또는 인라인
    username = (p.get("username") or "").strip().lstrip("@")
    password = (p.get("password") or "").strip()
    name = (p.get("name") or "").strip()
    if p.get("account_id"):
        acc = next((a for a in _load_sender_accounts() if a.get("id") == p["account_id"]), None)
        if not acc:
            return jsonify({"error": "저장된 계정을 찾을 수 없습니다."}), 400
        username, password = acc["username"], acc["password"]
        name = name or acc.get("name", "")
    if not username or not password:
        return jsonify({"error": "발신 계정을 선택하거나 입력하세요."}), 400
    # 받는 사람 — 단일 또는 리스트
    targets = p.get("targets")
    if not targets:
        targets = [{"id": p.get("target_id") or "", "name": p.get("target_name") or ""}]
    message = (p.get("message") or "").strip()
    if not message:
        return jsonify({"error": "메시지를 입력하세요."}), 400
    rows = [{"sender_id": username, "sender_pw": password, "sender_name": name,
             "target_id": (t.get("id") or "").strip().lstrip("@"),
             "target_name": (t.get("name") or "").strip(), "message": message}
            for t in targets if (t.get("id") or "").strip()]
    if not rows:
        return jsonify({"error": "받는 사람 ID를 입력하세요."}), 400

    def _i(k, d):
        try:
            return int(p.get(k) or d)
        except Exception:
            return d
    opts = {"daily_limit": _i("daily_limit", 30), "batch_limit": _i("batch_limit", 10),
            "gap_min": _i("gap_min", 60), "gap_max": _i("gap_max", 300),
            "break_every": _i("break_every", 6)}
    auto_follow = bool(p.get("auto_follow"))

    job_id = uuid.uuid4().hex[:12]
    DM_EXCEL_ROWS[job_id] = rows
    DM_JOBS_STATE[job_id] = {
        "id": job_id, "kind": "manual", "status": "running",
        "total": len(rows), "sent": 0, "failed": 0, "held": 0, "current": None, "log": [],
        "accounts": 1, "auto_follow": auto_follow, "opts": opts,
        "started_at": datetime.now().isoformat(timespec="seconds"), "finished_at": None,
    }

    def run():
        try:
            from dm_sender import DMSender  # type: ignore
            sender = DMSender(state=DM_JOBS_STATE[job_id], log_callback=_log_callback(job_id))
            sender.run_excel_rows(rows, auto_follow=auto_follow, opts=opts)
            DM_JOBS_STATE[job_id]["status"] = "done"
        except Exception as e:  # noqa: BLE001
            log.exception("DM manual job failed")
            DM_JOBS_STATE[job_id]["status"] = "error"
            DM_JOBS_STATE[job_id]["log"].append(f"❌ 에러: {e}")
        finally:
            DM_JOBS_STATE[job_id]["finished_at"] = datetime.now().isoformat(timespec="seconds")

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"job_id": job_id, "total": len(rows)})


def main():
    cfg = load_config()
    host = cfg.get("server", {}).get("host", "127.0.0.1")
    port = int(cfg.get("server", {}).get("port", 5000))

    # Cloudflare Tunnel 자동 시작 (로컬 모드에서만)
    auto_tunnel = cfg.get("tunnel", {}).get("auto_start", True)
    if cfg.get("env_mode") == "cloud":
        log.info("Cloud 모드 — Cloudflare Tunnel 건너뜀")
    elif auto_tunnel and TUNNEL_EXE.exists():
        log.info("Cloudflare Tunnel 자동 시작 중...")
        try:
            _start_tunnel()
        except Exception as e:
            log.warning(f"Tunnel 자동 시작 실패 (무시): {e}")

    log.info(f"서버 시작 → http://{host}:{port}")
    try:
        _start_drive_sync_scheduler()
        log.info("Drive sync 스케줄러 시작 (5분 주기)")
    except Exception as e:
        log.warning(f"Drive sync 스케줄러 시작 실패 (무시): {e}")
    app.run(host=host, port=port, debug=False, use_reloader=False)


if __name__ == "__main__":
    main()
