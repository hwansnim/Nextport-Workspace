"""
인플루언서 / 우리 계정 엑셀 자동 임포트.

지원 엑셀 형식 (자동 감지):
  - 인플루언서: URL, 인스타ID, 셀러명, 현재상태, 최종발송일, 마지막사용계정ID, 발송차수, 히스토리, 비고
  - 우리 계정: 계정생성일, 연결기기, 로그인ID, 로그인PW, 인스타ID, 계정주인, 구글연결계정,
              구글연결계정PW, 연결전화번호, 상태, 누적발송량, 비고
"""
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any


# ─── 컬럼 자동 매핑 ─────────────────────────────────────
INFLUENCER_COLUMN_MAP = {
    "url": ["url", "URL", "링크", "인스타URL", "instagram_url"],
    "instagram_id": ["인스타ID", "인스타 ID", "instagram_id", "instagram", "IG", "아이디"],
    "seller_name": ["셀러명", "셀러 이름", "이름", "셀러", "seller_name", "name"],
    "status": ["현재상태", "상태", "status", "발송상태", "상황"],
    "follower_count": ["팔로워수", "팔로워", "follower", "followers"],
    "category": ["카테고리", "category"],
    "owner": ["담당자", "owner", "manager"],
    "first_reply_date": ["첫회신일", "최초회신일", "first_reply"],
    "reply_account": ["회신계정", "수신계정", "reply_account"],
    "email": ["이메일", "email", "메일"],
    "phone": ["전화번호", "phone", "휴대폰", "tel"],
    "kakao_id": ["카카오톡ID", "카카오톡 ID", "카톡ID", "카톡", "kakao"],
    "last_sent_date": ["최종발송일", "최근발송일", "마지막발송일", "last_sent"],
    "last_sent_account_id": ["마지막사용계정ID", "사용계정", "last_account", "마지막계정"],
    "send_count": ["발송차수", "차수", "발송수", "count"],
    "history": ["히스토리", "기록", "history"],
    "notes": ["비고", "메모", "notes", "memo"],
}

ACCOUNT_COLUMN_MAP = {
    "created_date": ["계정생성일", "생성일", "created"],
    "device": ["연결기기", "기기", "device", "디바이스"],
    "login_id": ["로그인ID", "로그인 ID", "이메일", "login_id", "email"],
    "login_pw": ["로그인PW", "로그인 PW", "비밀번호", "password", "pw"],
    "instagram_id": ["인스타ID", "인스타 ID", "instagram", "IG"],
    "account_owner": ["계정주인", "주인", "owner", "구글계정"],
    "linked_email": ["구글연결계정", "백업이메일", "linked_email", "백업"],
    "linked_email_pw": ["구글연결계정PW", "백업PW", "linked_pw"],
    "phone": ["연결전화번호", "전화번호", "phone", "번호"],
    "status": ["상태", "status"],
    "total_sent": ["누적발송량", "누적발송", "total_sent", "발송수"],
    "notes": ["비고", "메모", "notes"],
}


def _norm_header(s: str) -> str:
    return re.sub(r"[\s\-_/\\]+", "", str(s).strip().lower())


def _find_col_index(headers: list[str], candidates: list[str]) -> int:
    norm_headers = [_norm_header(h) for h in headers]
    for cand in candidates:
        norm_cand = _norm_header(cand)
        for i, h in enumerate(norm_headers):
            if h == norm_cand or norm_cand in h or h in norm_cand:
                return i
    return -1


def _cell(row: list, idx: int, default: Any = "") -> Any:
    if idx < 0 or idx >= len(row):
        return default
    v = row[idx]
    if v is None:
        return default
    return v


def _str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _int(v: Any, default: int = 0) -> int:
    try:
        if v is None or v == "":
            return default
        return int(float(v))
    except (ValueError, TypeError):
        return default


def _date(v: Any) -> str:
    """다양한 날짜 포맷 → YYYY-MM-DD. 1899-12-30 등은 빈값으로."""
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        # 1899-12-30 (엑셀 기본값) 제외
        if v.year < 1950:
            return ""
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s or "1899" in s:
        return ""
    # YYYY-MM-DD, YYYY.MM.DD, YYYY/MM/DD 등
    s2 = re.sub(r"[./]", "-", s)
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s2)
    if m:
        y, mo, d = m.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    return s


def _next_inf_id(items: list[dict]) -> str:
    nums = []
    for it in items:
        m = re.match(r"inf(\d+)", str(it.get("id", "")))
        if m:
            nums.append(int(m.group(1)))
    return f"inf{(max(nums) if nums else 0) + 1:05d}"


def _next_acc_id(items: list[dict]) -> str:
    nums = []
    for it in items:
        m = re.match(r"acc(\d+)", str(it.get("id", "")))
        if m:
            nums.append(int(m.group(1)))
    return f"acc{(max(nums) if nums else 0) + 1:04d}"


def import_influencers(xlsx_path: Path, existing: list[dict], mode: str = "add") -> dict:
    """엑셀 → 인플루언서 리스트.
    mode='add' → 중복 instagram_id 자동 skip (기본)
    mode='update' → 중복이면 기존 항목 필드 갱신 (history/used_account_ids 등 누적 데이터는 보존)
    """
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl 미설치. pip install openpyxl"}

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    # 헤더 행 찾기 (앞 10행 중 'URL' 또는 '인스타ID' 있는 행)
    header_row_idx = None
    headers = []
    for ridx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), start=1):
        norm = [_norm_header(c) for c in row if c is not None]
        if any("url" in n or "인스타" in n or "instagram" in n for n in norm):
            header_row_idx = ridx
            headers = [_str(c) for c in row]
            break

    if header_row_idx is None:
        wb.close()
        return {"error": "헤더 행을 찾지 못함 (URL/인스타ID 컬럼 필요)"}

    col = {}
    for key, candidates in INFLUENCER_COLUMN_MAP.items():
        col[key] = _find_col_index(headers, candidates)

    by_handle = {it.get("instagram_id", "").lower(): it for it in existing}
    now = datetime.now().isoformat(timespec="seconds")
    added = []
    updated = 0
    skipped_dup = 0
    skipped_empty = 0
    UPDATABLE_FIELDS = ["seller_name", "status", "notes", "history_text", "url"]

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        ig_id = _str(_cell(row, col["instagram_id"]))
        url = _str(_cell(row, col["url"]))

        if not ig_id and url:
            m = re.search(r"instagram\.com/([^/?\s]+)", url)
            if m:
                ig_id = m.group(1).strip("_/").strip()

        if not ig_id:
            skipped_empty += 1
            continue
        existing_item = by_handle.get(ig_id.lower())
        if existing_item:
            if mode != "update":
                skipped_dup += 1
                continue
            # 갱신 — 누적 데이터(history, used_account_ids, send_count 등)는 건드리지 않음
            new_values = {
                "seller_name": _str(_cell(row, col["seller_name"])),
                "status": _str(_cell(row, col["status"])),
                "notes": _str(_cell(row, col["notes"])),
                "history_text": _str(_cell(row, col["history"])),
                "url": url,
            }
            for k in UPDATABLE_FIELDS:
                v = new_values.get(k, "")
                if v:
                    existing_item[k] = v
            updated += 1
            continue
        by_handle[ig_id.lower()] = True  # placeholder for dedup within this import

        new_item = {
            "id": _next_inf_id(existing + added),
            "instagram_id": ig_id,
            "url": url or f"https://www.instagram.com/{ig_id}/",
            "seller_name": _str(_cell(row, col["seller_name"])),
            "status": _str(_cell(row, col["status"])) or "미발송",
            # ─── 확장 필드 (스크린샷 컬럼 반영) ───
            "follower_count": _str(_cell(row, col["follower_count"])),
            "category": _str(_cell(row, col["category"])),
            "owner": _str(_cell(row, col["owner"])),
            "first_reply_date": _date(_cell(row, col["first_reply_date"])),
            "reply_account": _str(_cell(row, col["reply_account"])),
            "email": _str(_cell(row, col["email"])),
            "phone": _str(_cell(row, col["phone"])),
            "kakao_id": _str(_cell(row, col["kakao_id"])),
            # ─── 발송 추적 ───
            "last_sent_date": _date(_cell(row, col["last_sent_date"])),
            "last_sent_account_id": _str(_cell(row, col["last_sent_account_id"])),
            "send_count": _int(_cell(row, col["send_count"])),
            "history_text": _str(_cell(row, col["history"])),
            "notes": _str(_cell(row, col["notes"])),
            "history": [],
            "used_account_ids": [],
            "reply_received": False,
            "last_reply_at": None,
            # ─── 파이프라인 (진행 예정 셀러) ───
            "pipeline_stage": "",  # "" | "진행예정" | "미팅예약" | "미팅완료" | "캠페인진행중" | "종료"
            "meetings": [],        # [{date, round, note}]
            "imported_at": now,
        }
        added.append(new_item)

    wb.close()
    return {
        "ok": True,
        "added": len(added),
        "updated": updated,
        "skipped_duplicate": skipped_dup,
        "skipped_empty": skipped_empty,
        "items": added,
        "header_row": header_row_idx,
        "columns_found": {k: headers[v] if v >= 0 else None for k, v in col.items()},
    }


def import_accounts(xlsx_path: Path, existing: list[dict], mode: str = "add") -> dict:
    """엑셀 → 우리 계정 리스트.
    mode='add' → 중복 instagram_id 자동 skip
    mode='update' → 중복이면 기존 항목 필드 갱신 (total_sent/daily_* 등 누적 데이터 보존)
    """
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl 미설치"}

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    header_row_idx = None
    headers = []
    for ridx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), start=1):
        norm = [_norm_header(c) for c in row if c is not None]
        if any("로그인" in n or "loginid" in n or "인스타" in n for n in norm):
            header_row_idx = ridx
            headers = [_str(c) for c in row]
            break

    if header_row_idx is None:
        wb.close()
        return {"error": "헤더 행 못 찾음"}

    col = {}
    for key, candidates in ACCOUNT_COLUMN_MAP.items():
        col[key] = _find_col_index(headers, candidates)

    by_ig = {it.get("instagram_id", "").lower(): it for it in existing}
    now = datetime.now().isoformat(timespec="seconds")
    added = []
    updated = 0
    skipped_dup = 0
    skipped_empty = 0
    UPDATABLE = ["device", "login_id", "login_pw", "account_owner", "linked_email",
                 "linked_email_pw", "phone", "status", "notes", "created_date"]

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        ig_id = _str(_cell(row, col["instagram_id"]))
        login_id = _str(_cell(row, col["login_id"]))
        if not ig_id and not login_id:
            skipped_empty += 1
            continue
        if not ig_id:
            ig_id = login_id.split("@")[0]

        status_raw = _str(_cell(row, col["status"])) or "활성"
        status = status_raw
        if "활성" in status_raw:
            status = "활성"
        elif "차단" in status_raw or "정지" in status_raw:
            status = "차단"
        elif "휴식" in status_raw or "휴면" in status_raw:
            status = "휴식"
        elif "인증" in status_raw or "사람" in status_raw:
            status = "사람인증"

        existing_item = by_ig.get(ig_id.lower())
        if existing_item:
            if mode != "update":
                skipped_dup += 1
                continue
            new_values = {
                "device": _str(_cell(row, col["device"])),
                "login_id": login_id,
                "login_pw": _str(_cell(row, col["login_pw"])),
                "account_owner": _str(_cell(row, col["account_owner"])),
                "linked_email": _str(_cell(row, col["linked_email"])),
                "linked_email_pw": _str(_cell(row, col["linked_email_pw"])),
                "phone": _str(_cell(row, col["phone"])),
                "status": status,
                "notes": _str(_cell(row, col["notes"])),
                "created_date": _date(_cell(row, col["created_date"])),
            }
            for k in UPDATABLE:
                v = new_values.get(k, "")
                if v:
                    existing_item[k] = v
            updated += 1
            continue
        by_ig[ig_id.lower()] = True

        added.append({
            "id": _next_acc_id(existing + added),
            "instagram_id": ig_id,
            "created_date": _date(_cell(row, col["created_date"])),
            "device": _str(_cell(row, col["device"])),
            "login_id": login_id,
            "login_pw": _str(_cell(row, col["login_pw"])),
            "account_owner": _str(_cell(row, col["account_owner"])),
            "linked_email": _str(_cell(row, col["linked_email"])),
            "linked_email_pw": _str(_cell(row, col["linked_email_pw"])),
            "phone": _str(_cell(row, col["phone"])),
            "status": status,
            "total_sent": _int(_cell(row, col["total_sent"])),
            "daily_sent_today": 0,
            "daily_reset_date": now[:10],
            "daily_limit": 50,
            "last_used_at": None,
            "notes": _str(_cell(row, col["notes"])),
            "imported_at": now,
        })

    wb.close()
    return {
        "ok": True,
        "added": len(added),
        "updated": updated,
        "skipped_duplicate": skipped_dup,
        "skipped_empty": skipped_empty,
        "items": added,
        "columns_found": {k: headers[v] if v >= 0 else None for k, v in col.items()},
    }
