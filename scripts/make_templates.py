"""
인플루언서/계정 엑셀 양식 생성기.
data/imports/templates/ 에 두 파일 박음.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "data" / "imports" / "templates"
OUT_DIR.mkdir(parents=True, exist_ok=True)

HEAD_FILL = PatternFill("solid", fgColor="FF6B35")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
NOTE_FILL = PatternFill("solid", fgColor="FFF5E8")
NOTE_FONT = Font(italic=True, color="888888", size=9)

INF_HEADERS = [
    ("인스타ID", "필수 · @ 빼고", "thefashion_kr"),
    ("URL", "선택 · 비우면 자동생성", "https://www.instagram.com/thefashion_kr/"),
    ("셀러명", "선택", "더패션"),
    ("상태", "미발송/발송중/답장받음/컨펌/거절/비공개", "미발송"),
    ("발송차수", "숫자", 0),
    ("최종발송일", "YYYY-MM-DD", ""),
    ("마지막사용계정ID", "우리 계정의 인스타ID", ""),
    ("히스토리", "메모", ""),
    ("비고", "메모", ""),
]

ACC_HEADERS = [
    ("인스타ID", "필수 · @ 빼고", "next_official"),
    ("생성일", "YYYY-MM-DD", "2025-12-01"),
    ("기기", "예: 갤럭시1, 아이폰3", "갤럭시1"),
    ("로그인ID", "이메일/전화/유저네임", "next.official@gmail.com"),
    ("로그인PW", "비번", ""),
    ("계정주인", "사람 이름", "환님"),
    ("구글연결계정", "이메일", "next.official@gmail.com"),
    ("구글PW", "구글 계정 비번", ""),
    ("전화번호", "", "010-1234-5678"),
    ("상태", "활성/사람인증/휴식/차단", "활성"),
    ("비고", "메모", ""),
]


def make_sheet(path, title, headers, sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14, color="FF6B35")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    note = ws.cell(row=2, column=1, value="① 헤더 이름은 한글/영문 모두 자동 인식  ② instagram_id 중복은 자동 skip  ③ 예시 행은 임포트 전 삭제")
    note.font = NOTE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    for col_idx, (name, hint, sample) in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=name)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")

        h = ws.cell(row=5, column=col_idx, value=hint)
        h.font = NOTE_FONT
        h.fill = NOTE_FILL
        h.alignment = Alignment(horizontal="center", wrap_text=True)

        ws.cell(row=6, column=col_idx, value=sample).font = Font(color="888888", italic=True)

        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(name) * 2.2)

    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 36
    ws.freeze_panes = "A7"

    wb.save(path)
    print(f"  OK {path.name} ({len(headers)} columns)")


import sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

print(f"[templates] out: {OUT_DIR}")
make_sheet(OUT_DIR / "influencers_template.xlsx", "인플루언서 명단 (5,000명 마스터)", INF_HEADERS, "인플루언서")
make_sheet(OUT_DIR / "accounts_template.xlsx", "우리 인스타 계정 (100+ 마스터)", ACC_HEADERS, "계정")
print("[templates] done")
